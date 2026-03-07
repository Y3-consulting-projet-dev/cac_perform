from genericpath import exists
from mongoengine import Document, StringField, EmbeddedDocument, EmbeddedDocumentField, ListField, DictField
import bcrypt
import openpyxl
from bson import ObjectId
import json
from io import BytesIO
import os
import unicodedata

# Fallback global pour éviter les NameError si du code legacy
# référence une variable 'db' sans l'initialiser. Chaque méthode
# doit continuer à appeler get_db() pour obtenir une connexion.
db = None

# Utilisation de la configuration centralisée
from src.utils.database import get_database

# Fonction helper pour obtenir la base de données de manière lazy
# (ne pas obtenir au moment de l'import, mais seulement quand nécessaire)
def get_db():
    """
    Obtient la base de données via le gestionnaire centralisé
    Tente de reconnecter automatiquement si la connexion est perdue
    """
    try:
        return get_database()
    except RuntimeError as e:
        # Si la connexion est perdue, essayer de reconnecter
        from config import db_manager
        
        # Vérifier si db_manager a une configuration
        # Si oui, tenter de reconnecter
        if hasattr(db_manager, 'config') and db_manager.config is not None:
            try:
                print("Connexion perdue, tentative de reconnexion automatique...")
                db_manager.connect()
                return db_manager.get_db()
            except Exception as reconnect_error:
                print(f"Échec de la reconnexion: {reconnect_error}")
                raise RuntimeError(f"Base de données non connectée: {e}") from reconnect_error
        else:
            # Si pas de configuration, la base de données n'a pas été initialisée
            # Cela peut arriver si on appelle get_db() avant l'initialisation de Flask
            error_msg = f"Base de données non initialisée. Assurez-vous que Flask est démarré et que la base de données est connectée. Erreur originale: {e}"
            print(f"{error_msg}")
            raise RuntimeError(error_msg) from e


# ==============================
#  Clients
# ==============================
class Client(Document):

    @classmethod
    def afficher_clients(cls):
        try:
            db = get_db()
            cust_data = list(db.Client.find().sort({"_id": -1}))
            for client in cust_data:
                client['_id'] = str(client['_id'])
            return cust_data
        except Exception as e:
            print(f"An exception : {str(e)}")

    @classmethod
    def afficher_missions(cls, id):
        db = get_db()
        results = []
        final = []
        
        # Normaliser l'id (s'assurer que c'est une string)
        id_str = str(id).strip()
        
        print(f"Recherche des missions pour client ID: '{id_str}'")
        
        # Chercher avec l'id tel quel
        query = {"id_client": id_str}
        results = db.Mission1.find(query)
        
        count = db.Mission1.count_documents(query)
        print(f"Missions trouvées avec id_client='{id_str}': {count}")
        
        # Si aucune mission trouvée, essayer avec ObjectId si l'id ressemble à un ObjectId
        if count == 0:
            try:
                # Essayer avec ObjectId
                query_objid = {"id_client": ObjectId(id_str)}
                count_objid = db.Mission1.count_documents(query_objid)
                print(f"Tentative avec ObjectId: {count_objid} missions")
                if count_objid > 0:
                    results = db.Mission1.find(query_objid)
                    count = count_objid
            except:
                pass
        
        # Récupérer toutes les missions pour debug
        total_missions = db.Mission1.count_documents({})
        print(f"Total de missions dans la base: {total_missions}")
        
        if total_missions > 0 and count == 0:
            # Afficher quelques exemples pour debug
            sample_missions = list(db.Mission1.find({}).limit(3))
            print(f"Exemples d'id_client dans la base:")
            for sample in sample_missions:
                sample_id_client = sample.get('id_client', 'NON DÉFINI')
                print(f"   - Mission {sample['_id']}: id_client='{sample_id_client}' (type: {type(sample_id_client)})")
        
        for result in results:
            result['_id'] = str(result['_id'])
            final.append(result)
        
        print(f"Retour de {len(final)} missions pour le client {id_str}")
        return final

    @classmethod
    def ajouter_client(cls, data):
        try:
            db = get_db()
            cust_data = db.Client.insert_one(data)
            inserted_id = str(cust_data.inserted_id)
            return inserted_id
        except Exception as e:
            print(f'An exception occurred: {str(e)}')

    @classmethod
    def modifier_client(cls, data):
        try:
            db = get_db()
            _id = str(data['_id'])
            id_client = ObjectId(_id)

            nom = data['nom']
            activite = data['activite']
            referentiel = data['referentiel']
            forme_juridique = data['forme_juridique']
            capital = data['capital']
            siege_social = data['siege_social']
            adresse = data['adresse']
            n_cc = data['n_cc']

            if id_client:
                db.Client.update_one(
                    {"_id": id_client},
                    {"$set": {
                        "nom": nom,
                        "activite": activite,
                        "referentiel": referentiel,
                        "forme_juridique": forme_juridique,
                        "capital": capital,
                        "siege_social": siege_social,
                        "adresse": adresse,
                        "n_cc": n_cc,
                    }}
                )
                return "Updated succeeded"
        except Exception as e:
            print(f'An exception occurred: {str(e)}')

    @classmethod
    def info_client(cls, id_clit):
        try:
            db = get_db()
            id_client = id_clit
            info = db.Client.find_one({"_id": id_client})

            valeurs = {}
            if info:
                valeurs = {
                    "_id": str(id_client),
                    "nom": info['nom'],
                    "activite": info['activite'],
                    "siege_social": info['siege_social'],
                    "adresse": info['adresse'],
                    "referentiel": info['referentiel'],
                    "forme_juridique": info['forme_juridique'],
                    "capital": info['capital'],
                    "n_cc": info['n_cc'],
                }
            return valeurs
        except Exception as e:
            print(f"An exception : {str(e)}")
            return None

    @classmethod
    def supprimer_client(cls, id_client):
        """Supprime un client et toutes ses missions associées"""
        try:
            db = get_db()
            # Convertir l'ID en ObjectId
            client_id = ObjectId(str(id_client))
            
            # Vérifier que le client existe
            client_info = db.Client.find_one({"_id": client_id})
            if not client_info:
                return {"success": False, "message": "Client non trouvé"}
            
            # Supprimer toutes les missions associées au client
            missions_deleted = db.Mission1.delete_many({"id_client": str(id_client)})
            
            # Supprimer le client
            client_deleted = db.Client.delete_one({"_id": client_id})
            
            if client_deleted.deleted_count > 0:
                return {
                    "success": True, 
                    "message": f"Client supprimé avec succès. {missions_deleted.deleted_count} mission(s) supprimée(s).",
                    "client_name": client_info.get('nom', 'Inconnu')
                }
            else:
                return {"success": False, "message": "Erreur lors de la suppression du client"}
                
        except Exception as e:
            print(f"Erreur lors de la suppression du client: {str(e)}")
            return {"success": False, "message": f"Erreur: {str(e)}"}

    def choix_referentiel(self):
        try:
            db = get_db()
            ref = db.Grouping.find()
            return ref
        except Exception as e:
            print(f"An exception as occured: {str(e)}")


# ==============================
#  Authentification supprimée
# ==============================
# L'ancien système Manager a été remplacé par le nouveau système
# d'authentification dans src/services/user_services.py


# ==============================
#  Modèles embarqués
# ==============================
class LigneComptable(EmbeddedDocument):
    num_compte = StringField(required=True, unique=True)
    libelle = StringField()
    solde = DictField()


class BalanceComptable(EmbeddedDocument):
    lignes = ListField(EmbeddedDocumentField(LigneComptable))


# ==============================
#  Mission
# ==============================
class Mission(Document):
    # champs (non utilisés par mongoengine ici mais conservés)
    balance = EmbeddedDocumentField(BalanceComptable)
    annee_auditee = ListField()

    def get_referentiel_groupes(self, referentiel="syscohada"):
        """
        Retourne la liste des groupes pour un référentiel depuis grouping.json
        """
        try:
            grouping_path = os.path.join(os.path.dirname(__file__), "..", "grouping.json")
            with open(grouping_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            return data.get(referentiel)
        except Exception as e:
            print(f"Erreur chargement referentiel '{referentiel}': {str(e)}")
            return None

    # ---------- Revue analytique ----------
    def revue_analytique(self, id_mission):
        db = get_db()
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        variations = mission.get("balance_variation", [])

        # mapping_efi.json (un niveau au-dessus de ce fichier)
        mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")
        with open(mapping_path, "r", encoding="utf-8") as f:
            mapping_root = json.load(f)
            mapping = mapping_root["structure"]

        # seuils (matérialité choisie si existe)
        materiality = mission.get("materiality", [])
        choice = next((m for m in materiality if m.get("choice")), None)
        perf_mat = abs(int(choice.get("performance_materiality", 0))) if choice else 0

        def collect_prefixes(bloc):
            prefixes = []
            for key in ("brut_cpt", "amor_cpt", "net_cpt", "brut_except_cpt", "amor_except_cpt", "net_except_cpt"):
                if key in bloc and bloc[key]:
                    if isinstance(bloc[key], str):
                        prefixes += [p.strip() for p in bloc[key].split(",") if p.strip()]
                    elif isinstance(bloc[key], list):
                        prefixes += [str(p).strip() for p in bloc[key] if str(p).strip()]
            return prefixes

        def map_efi(numero):
            refs = []
            for bloc in mapping:
                prefixes = collect_prefixes(bloc)
                if any(str(numero).startswith(p) for p in prefixes):
                    refs.append(bloc["ref"])
            return sorted(list(set(refs)))

        out = []
        for row in variations:
            n = int(row.get("solde_n", 0) or 0)
            n1 = int(row.get("solde_n1", 0) or 0)
            delta = n - n1
            delta_pct = (delta / (abs(n1) if n1 else 1.0))
            efi_refs = map_efi(row.get("numero_compte", ""))

            # Commentaire automatique (EFI + Matérialité)
            commentaire_auto = (
                f"EFI: {', '.join(efi_refs) if efi_refs else 'Aucun'}"
                f"Matérialité: {'Oui' if (perf_mat and abs(delta) >= perf_mat) else 'Non'}"
            )

            out.append({
                "numero_compte": row.get("numero_compte"),
                "libelle": row.get("libelle"),
                "solde_n": n,
                "solde_n1": n1,
                "delta_abs": delta,
                "delta_pct": delta_pct,
                "commentaire_auto": commentaire_auto,
                "commentaire_perso": ""  # Commentaire personnalisé vide par défaut
            })

        db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"revue_analytique": out}})
        return out

    def update_commentaire_perso(self, id_mission, numero_compte, commentaire_perso):
        """
        Met à jour (ou initialise) le champ 'commentaire_perso' pour une ligne
        de la revue analytique correspondant au numero_compte.
        Retourne True si l'opération aboutit, False sinon.
        """
        db = get_db()  # Obtenir la connexion à la base de données
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return False
            revue = mission.get("revue_analytique", [])
            if not revue:
                return False
            index = next((i for i, it in enumerate(revue) if it.get("numero_compte") == numero_compte), None)
            if index is None:
                return False
            path = f"revue_analytique.{index}.commentaire_perso"
            result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {path: commentaire_perso or ""}})
            return result.acknowledged
        except Exception:
            return False

    def update_commentaire_perso(self, id_mission, numero_compte, commentaire_perso):
        """
        Met à jour le commentaire personnalisé pour un compte spécifique
        """
        try:
            print(f"Début de mise à jour du commentaire pour le compte {numero_compte}")
            
            # Vérifier d'abord que la mission existe
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                print(f"Mission {id_mission} non trouvée")
                return False
            
            # Vérifier que la revue analytique existe
            if "revue_analytique" not in mission:
                print(f"Revue analytique non trouvée pour la mission {id_mission}")
                return False
            
            # Trouver l'index du compte dans la revue analytique
            revue_analytique = mission["revue_analytique"]
            compte_index = None
            
            for i, item in enumerate(revue_analytique):
                if item.get("numero_compte") == numero_compte:
                    compte_index = i
                    break
            
            if compte_index is None:
                print(f"Compte {numero_compte} non trouvé dans la revue analytique")
                return False
            
            # Vérifier si le commentaire a réellement changé
            current_comment = revue_analytique[compte_index].get("commentaire_perso", "")
            if current_comment == commentaire_perso:
                print(f"Commentaire identique pour le compte {numero_compte}, aucune modification nécessaire")
                return True  # Retourner True car c'est un succès (pas d'erreur)
            
            print(f"Mise à jour du commentaire: '{current_comment}' -> '{commentaire_perso}'")
            
            # Mettre à jour le commentaire personnalisé
            result = db.Mission1.update_one(
                {"_id": ObjectId(id_mission)},
                {"$set": {f"revue_analytique.{compte_index}.commentaire_perso": commentaire_perso}}
            )
            
            success = result.modified_count > 0
            if success:
                print(f"Commentaire mis à jour avec succès pour le compte {numero_compte}")
            else:
                print(f"Aucune modification effectuée pour le compte {numero_compte} (modified_count: {result.modified_count})")
            
            return success
            
        except Exception as e:
            print(f"Erreur lors de la mise à jour du commentaire: {e}")
            import traceback
            traceback.print_exc()
            return False

    # ---------- Nouvelle mission ----------
    def nouvelle_mission(self, balances, annee_auditee, id_client, date_debut, date_fin, date_debut_mandat=None, date_fin_mandat=None, responsable_nom=None, responsable_grade=None, responsable_role=None, responsable_id=None):
        try:
            db = get_db()  # Obtenir la connexion à la base de données
            
            # Si la connexion est absente, tenter une reconnexion propre
            if db is None:
                from src.utils.database import ensure_connection
                print("Aucune connexion DB détectée, tentative de reconnexion...")
                ensure_connection()
                db = get_db()
                if db is None:
                    raise Exception("Base de données non connectée (get_db a renvoyé None)")
        except Exception as e:
            print(f"ERREUR: Impossible d'obtenir la connexion à la base de données: {e}")
            raise Exception(f"Erreur de connexion à la base de données: {str(e)}")
        
        balance_ids = []
        les_balance_n_n1 = []
        annee_balance = annee_auditee

        # S'assurer que db est défini avant la boucle
        try:
            _ = db  # Vérifier que db existe
        except NameError:
            # Si db n'est pas défini, le redéfinir
            db = get_db()

        for balance in balances:
            try:
                balance_created = self.creation_balance(balance, int(annee_auditee), id_client)
                annee_auditee = int(annee_auditee) - 1

                tuple_en_tableau = list(*[balance_created])
                balance_ids.append(tuple_en_tableau[0])
                les_balance_n_n1.append(tuple_en_tableau[1])
            except Exception as e:
                # S'assurer que db est défini avant de lever l'exception
                try:
                    _ = db
                except NameError:
                    db = get_db()
                # Re-lever l'exception avec un message plus clair
                error_msg = f"Erreur lors de la création de la balance: {str(e)}"
                print(f"{error_msg}")
                raise Exception(error_msg) from e

        balance_variation = self.rapprochement_des_balances(
            les_balance_n_n1[0], les_balance_n_n1[1]
        )

        # Récupérer le référentiel (par défaut "syscohada")
        referentiel = "syscohada"  # Vous pouvez le récupérer depuis les données si nécessaire
        grouping_model = self.create_grouping(balance_variation, referentiel)
        etats = self.prod_efi(les_balance_n_n1[0], les_balance_n_n1[1], balance_variation)

        # S'assurer que id_client est bien une string
        id_client_str = str(id_client).strip()
        
        print(f"Création de mission avec id_client='{id_client_str}'")
        
        # Préparer les données de la mission
        mission_data = {
            "id_client": id_client_str,
            "annee_auditee": str(annee_balance),
            "date_debut": date_debut,
            "date_fin": date_fin,
            "balances": balance_ids,
            "balance_variation": balance_variation,
            "grouping": grouping_model,
            "efi": etats,
            "materiality": []
        }

        if responsable_nom:
            mission_data["responsable_nom"] = responsable_nom
        if responsable_grade:
            mission_data["responsable_grade"] = responsable_grade
        if responsable_role:
            mission_data["responsable_role"] = responsable_role
        if responsable_id:
            mission_data["responsable_id"] = responsable_id
        
        # Ajouter les dates du mandat si fournies
        if date_debut_mandat:
            mission_data["date_debut_mandat"] = date_debut_mandat
            print(f"Date de début du mandat: {date_debut_mandat}")
        if date_fin_mandat:
            mission_data["date_fin_mandat"] = date_fin_mandat
            print(f"Date de fin du mandat: {date_fin_mandat}")
        
        result = db.Mission1.insert_one(mission_data)

        insert_id = str(result.inserted_id)
        print(f"Mission créée avec ID: {insert_id}")
        
        # Vérifier que la mission a bien été sauvegardée
        mission_saved = db.Mission1.find_one({"_id": result.inserted_id})
        if mission_saved:
            print(f"Mission vérifiée en base: id_client='{mission_saved.get('id_client')}'")
        else:
            print(f"ATTENTION: Mission créée mais non trouvée en base!")
        res = {
            "id_client": id_client,
            "annee_auditee": str(annee_balance),
            "date_debut": date_debut,
            "date_fin": date_fin,
            "balances": balance_ids,
            "balance_variation": balance_variation,
            "grouping": grouping_model,
            "efi": etats,
            "materiality": []
        }

        format_id = {"_id": insert_id, "mission": res}
        self.audit_trail(format_id['_id'])
        return format_id

    # ---------- Lecture d'une balance Excel -> Mongo (Excel uniquement) ----------
    def creation_balance(self, balance_data, annee_auditee, id_client):
        try:
            db = get_db()  # Obtenir la connexion à la base de données
        except Exception as e:
            print(f"ERREUR: Impossible d'obtenir la connexion à la base de données: {e}")
            raise Exception(f"Erreur de connexion à la base de données: {str(e)}")
        
        balance = balance_data
        data = []

        try:
            # Traitement des fichiers Excel uniquement
            workbook = openpyxl.load_workbook(balance)
            
            # Détection automatique de la feuille de balance
            sheet = None
            sheet_name = None
            
            # Noms de feuilles acceptés (par ordre de priorité)
            accepted_sheet_names = [
                'Balance des comptes',  # Nom avec espaces (détecté dans vos fichiers)
                'Balance_des_comptes',  # Nom standard
                'BALANCE_2023',         # Votre format
                'BALANCE__2024',        # Votre format
                'Sage',                 # Format Sage
                'Sheet1',               # Format générique
                'Feuil1',               # Format français Excel (détecté dans vos fichiers)
                'Balance',              # Format court
                'Comptes',              # Format alternatif
            ]
            
            # Chercher la première feuille disponible
            for name in accepted_sheet_names:
                if name in workbook.sheetnames:
                    sheet = workbook[name]
                    sheet_name = name
                    print(f"Feuille trouvée: '{name}' dans {balance.filename if hasattr(balance, 'filename') else 'fichier'}")
                    break
            
            # Si aucune feuille standard n'est trouvée, prendre la première
            if sheet is None and workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
                sheet_name = workbook.sheetnames[0]
                print(f"Feuille non standard utilisée: '{sheet_name}' dans {balance.filename if hasattr(balance, 'filename') else 'fichier'}")
            
            if sheet is None:
                raise Exception("Aucune feuille trouvée dans le fichier Excel")

            print(f"Traitement de la feuille '{sheet_name}' avec {sheet.max_row} lignes et {sheet.max_column} colonnes")

            def _parse_number(value):
                """Conversion robuste des montants Excel (nombres ou textes formatés)."""
                try:
                    if value is None:
                        return 0.0
                    if isinstance(value, (int, float)):
                        return float(value)

                    s = str(value).strip()
                    if not s:
                        return 0.0
                    s = s.replace('\u00a0', '').replace(' ', '')
                    s = s.replace('', '-').replace('', '-')
                    if s in ('-', '--'):
                        return 0.0

                    neg = False
                    if s.startswith('(') and s.endswith(')'):
                        neg = True
                        s = s[1:-1]

                    # Harmoniser séparateurs décimaux/milliers
                    if ',' in s and '.' in s:
                        if s.rfind(',') > s.rfind('.'):
                            s = s.replace('.', '').replace(',', '.')
                        else:
                            s = s.replace(',', '')
                    elif ',' in s:
                        # Si une seule virgule et 1-2 décimales -> décimal, sinon millier
                        parts = s.split(',')
                        if len(parts) == 2 and len(parts[1]) <= 2:
                            s = s.replace(',', '.')
                        else:
                            s = s.replace(',', '')

                    val = float(s)
                    return -val if neg else val
                except Exception:
                    return 0.0
            
            # Détection automatique du format et des colonnes
            print(f"Analyse du format de la feuille '{sheet_name}'...")
            
            # Analyser les premières lignes pour détecter le format
            format_detecte = None
            colonnes_map = {}
            
            # Formats supportés
            formats_supportes = {
                'standard': {
                    'pattern': ['numero', 'libelle', 'debit_initial', 'credit_initial', 'debit_mvt', 'credit_mvt', 'debit_fin', 'credit_fin'],
                    'detection': lambda row: len(row) >= 8 and any('compte' in str(cell).lower() for cell in row if cell)
                },
                'balance_simple': {
                    'pattern': ['numero', 'libelle', 'debit_initial', 'credit_initial', 'debit_fin', 'credit_fin'],
                    'detection': lambda row: len(row) >= 6 and (
                        any('compte' in str(cell).lower() for cell in row if cell) or
                        (str(row[0]).isdigit() and len(str(row[0])) >= 3)
                    )
                },
                'sage': {
                    'pattern': ['numero', 'libelle', 'solde_n1', 'mouvement_debit', 'mouvement_credit', 'solde_n'],
                    'detection': lambda row: len(row) >= 6 and str(row[0]).isdigit() and len(str(row[0])) >= 3
                },
                'simple': {
                    'pattern': ['numero', 'libelle', 'montant'],
                    'detection': lambda row: len(row) >= 3 and str(row[0]).isdigit()
                }
            }
            
            # Détecter le format (par ordre de priorité)
            # IMPORTANT: Si le fichier a 8 colonnes avec mouvements, utiliser "standard" en priorité
            # Ordre de détection: standard avant balance_simple si 8 colonnes détectées
            if sheet.max_column >= 8:
                format_priority = ['standard', 'balance_simple', 'sage', 'simple']
            else:
                format_priority = ['balance_simple', 'standard', 'sage', 'simple']
            
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                row_data = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(10, sheet.max_column + 1))]
                
                # Tester les formats par ordre de priorité
                for format_name in format_priority:
                    if format_name in formats_supportes:
                        format_info = formats_supportes[format_name]
                        if format_info['detection'](row_data):
                            format_detecte = format_name
                            print(f"Format détecté: {format_name} à la ligne {row_idx}")
                            break
                
                if format_detecte:
                    break
            
            if not format_detecte:
                print("Format non reconnu, utilisation du format par défaut")
                format_detecte = 'standard'
            
            print(f"\n{'='*60}")
            print(f"FORMAT FINAL DÉTECTÉ: {format_detecte}")
            print(f"   Nombre de colonnes dans le fichier: {sheet.max_column}")
            print(f"{'='*60}\n")
            
            # Traitement selon le format détecté
            if format_detecte == 'sage':
                # Format Sage: [numero, libelle, solde_n1, mouvement_debit, mouvement_credit, solde_n]
                print("Traitement format Sage...")
                
                # Compteurs pour le diagnostic
                lignes_traitees = 0
                lignes_ignorees_colonnes = 0
                lignes_ignorees_vides = 0
                lignes_ignorees_numero_invalide = 0
                lignes_erreur = 0
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                    # Vérifier que la ligne a au moins 6 colonnes
                    if len(row) < 6:
                        lignes_ignorees_colonnes += 1
                        continue
                        
                    if not row[0] or str(row[0]).strip() == '':
                        lignes_ignorees_vides += 1
                        continue
                    
                    try:
                        numero = str(row[0]).strip()
                        if not numero.isdigit() or len(numero) < 3:
                            lignes_ignorees_numero_invalide += 1
                            continue
                            
                        ligne = {}
                        ligne['numero_compte'] = numero
                        ligne['libelle'] = str(row[1] or '').strip()
                        
                        # Format Sage: solde_n1, mouvement_debit, mouvement_credit, solde_n
                        solde_n1 = _parse_number(row[2])
                        mouvement_debit = _parse_number(row[3])
                        mouvement_credit = _parse_number(row[4])
                        solde_n = _parse_number(row[5])
                        
                        # Calculer les soldes d'ouverture et de clôture
                        ligne['debit_initial'] = max(0, solde_n1) if solde_n1 > 0 else 0
                        ligne['credit_initial'] = abs(solde_n1) if solde_n1 < 0 else 0
                        ligne['debit_mvt'] = mouvement_debit
                        ligne['credit_mvt'] = mouvement_credit
                        ligne['debit_fin'] = max(0, solde_n) if solde_n > 0 else 0
                        ligne['credit_fin'] = abs(solde_n) if solde_n < 0 else 0
                        
                        ligne['solde_reel'] = ligne['debit_fin'] - ligne['credit_fin']
                        ligne['solde'] = abs(ligne['solde_reel'])
                        ligne['sign_solde'] = "D" if ligne['debit_fin'] >= ligne['credit_fin'] else "C"
                        
                        data.append(ligne)
                        lignes_traitees += 1
                        
                    except Exception as e:
                        lignes_erreur += 1
                        print(f"Erreur ligne {row_idx} (Sage): {e}")
                        continue
                
                # Afficher le résumé du traitement
                print(f"Résumé du traitement (format Sage):")
                print(f"Lignes traitées avec succès: {lignes_traitees}")
                if lignes_ignorees_colonnes > 0:
                    print(f"Lignes ignorées (moins de 6 colonnes): {lignes_ignorees_colonnes}")
                if lignes_ignorees_vides > 0:
                    print(f"Lignes ignorées (numéro de compte vide): {lignes_ignorees_vides}")
                if lignes_ignorees_numero_invalide > 0:
                    print(f"Lignes ignorées (numéro de compte invalide, doit être numérique avec 3+ chiffres): {lignes_ignorees_numero_invalide}")
                if lignes_erreur > 0:
                    print(f"Lignes avec erreur: {lignes_erreur}")
            
            elif format_detecte == 'balance_simple':
                # Format balance simple: [numero, libelle, debit_initial, credit_initial, debit_fin, credit_fin]
                # Sans les colonnes de mouvement
                print(f"\n{'='*60}")
                print("Traitement format balance simple...")
                print(f"{'='*60}\n")
                
                # Compteurs pour le diagnostic
                lignes_traitees = 0
                lignes_ignorees_colonnes = 0
                lignes_ignorees_vides = 0
                lignes_ignorees_numero_vide = 0
                lignes_erreur = 0
                
                header_row = 1
                
                # Trouver la ligne d'en-tête
                for row_idx in range(1, min(5, sheet.max_row + 1)):
                    row_data = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
                    if any('compte' in str(cell).lower() for cell in row_data if cell):
                        header_row = row_idx
                        print(f"En-tête détecté à la ligne {row_idx}: {row_data}")
                        break
                
                # DÉTECTION DE L'ORDRE DES COLONNES (si 8 colonnes)
                colonnes_inversees_balance_simple = False
                if sheet.max_column >= 8:
                    header_row_data = [sheet.cell(row=header_row, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
                    col7_header = str(header_row_data[6] if len(header_row_data) > 6 else '').strip()
                    col8_header = str(header_row_data[7] if len(header_row_data) > 7 else '').strip()
                    col7_val = col7_header.lower()
                    col8_val = col8_header.lower()
                    
                    print(f"DÉTECTION DE L'ORDRE DES COLONNES (format balance_simple):")
                    print(f"   Colonne 7 (index 6): '{col7_header}'")
                    print(f"   Colonne 8 (index 7): '{col8_header}'")
                    
                    has_solde_credit_col7 = (
                        ('solde' in col7_val and 'credit' in col7_val) or
                        ('solde' in col7_val and 'crédit' in col7_val) or
                        (col7_val and 'credit' in col7_val and 'debit' not in col7_val and 'débit' not in col7_val)
                    )
                    
                    has_solde_debit_col8 = (
                        ('solde' in col8_val and 'debit' in col8_val) or
                        ('solde' in col8_val and 'débit' in col8_val) or
                        (col8_val and 'debit' in col8_val and 'credit' not in col8_val and 'crédit' not in col8_val)
                    )
                    
                    if has_solde_credit_col7 and has_solde_debit_col8:
                        colonnes_inversees_balance_simple = True
                        print(f"Format détecté: Solde crédit (col 7) puis Solde débit (col 8) - COLONNES INVERSÉES")
                    else:
                        print(f"Format détecté: Débit (col 7) puis Crédit (col 8) - FORMAT STANDARD")
                
                # Traitement des lignes de données
                for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    # Vérifier que la ligne a au moins 6 colonnes
                    if len(row) < 6:
                        lignes_ignorees_colonnes += 1
                        continue
                        
                    if row[0] is None and row[1] is None:
                        lignes_ignorees_vides += 1
                        break  # Fin des données
                        
                    # Vérifier que la ligne contient des données valides
                    if not row[0] or str(row[0]).strip() == '':
                        lignes_ignorees_numero_vide += 1
                        continue
                    
                    try:
                        ligne = {}
                        # Convertir le numéro de compte en string, en gérant les cas None, 0, etc.
                        num_compte_raw = row[0]
                        if num_compte_raw is None:
                            raise ValueError("Numéro de compte est None")
                        num_compte_str = str(num_compte_raw).strip()
                        if not num_compte_str or num_compte_str.lower() == 'none' or num_compte_str.lower() == 'nan':
                            raise ValueError(f"Numéro de compte invalide: '{num_compte_raw}'")
                        ligne['numero_compte'] = num_compte_str
                        ligne['libelle'] = row[1] or ''
                        
                        # Format balance : on vérifie d'abord si les mouvements sont dans le fichier
                        ligne['debit_initial'] = _parse_number(row[2])
                        ligne['credit_initial'] = _parse_number(row[3])
                        
                        # Vérifier si le fichier Excel contient les colonnes de mouvements
                        # Format 1 (avec mouvements) : debit_initial, credit_initial, debit_mvt, credit_mvt, debit_fin, credit_fin
                        # Format 2 (sans mouvements) : debit_initial, credit_initial, debit_fin, credit_fin
                        try:
                            # Essayer de lire les mouvements depuis le fichier (colonnes 4 et 5)
                            if len(row) > 5 and (row[4] is not None or row[5] is not None):
                                ligne['debit_mvt'] = _parse_number(row[4])
                                ligne['credit_mvt'] = _parse_number(row[5])
                                # Utiliser l'ordre détecté pour les colonnes finales
                                if colonnes_inversees_balance_simple and len(row) > 7:
                                    ligne['credit_fin'] = _parse_number(row[6])
                                    ligne['debit_fin'] = _parse_number(row[7])
                                else:
                                    ligne['debit_fin'] = _parse_number(row[6])
                                    ligne['credit_fin'] = _parse_number(row[7])
                            else:
                                # Format sans mouvements : utiliser solde_final = solde_initial + mouvements
                                # Pour la vérification d'identité, on utilise l'équation : 
                                # (debit_fin - credit_fin) = (debit_initial - credit_initial) + (debit_mvt - credit_mvt)
                                ligne['debit_fin'] = _parse_number(row[4])
                                ligne['credit_fin'] = _parse_number(row[5])
                                
                                # Calculer le solde initial et final
                                solde_initial = ligne['debit_initial'] - ligne['credit_initial']
                                solde_final = ligne['debit_fin'] - ligne['credit_fin']
                                mouvement_net = solde_final - solde_initial
                                
                                # Approximer les mouvements débit/crédit à partir du mouvement net
                                # Note: Cette approximation n'est pas parfaite car on ne connaît pas
                                # le détail des mouvements débit/crédit séparément
                                if mouvement_net > 0:
                                    # Le solde a augmenté, donc plus de débits que de crédits
                                    ligne['debit_mvt'] = abs(mouvement_net)
                                    ligne['credit_mvt'] = 0
                                elif mouvement_net < 0:
                                    # Le solde a diminué, donc plus de crédits que de débits
                                    ligne['debit_mvt'] = 0
                                    ligne['credit_mvt'] = abs(mouvement_net)
                                else:
                                    # Pas de changement, mouvements à zéro ou équilibrés
                                    # Approximation : répartir selon la variation des totaux débit/crédit
                                    debit_mvt_calc = max(0, ligne['debit_fin'] - ligne['debit_initial'])
                                    credit_mvt_calc = max(0, ligne['credit_fin'] - ligne['credit_initial'])
                                    ligne['debit_mvt'] = debit_mvt_calc
                                    ligne['credit_mvt'] = credit_mvt_calc
                        except (IndexError, ValueError, TypeError):
                            # Si erreur, utiliser le format simple sans mouvements
                            ligne['debit_mvt'] = 0
                            ligne['credit_mvt'] = 0
                            ligne['debit_fin'] = _parse_number(row[4])
                            ligne['credit_fin'] = _parse_number(row[5])
                        
                        ligne['solde_reel'] = ligne['debit_fin'] - ligne['credit_fin']
                        ligne['solde'] = abs(ligne['solde_reel'])
                        ligne['sign_solde'] = "D" if ligne['debit_fin'] >= ligne['credit_fin'] else "C"
                        
                        data.append(ligne)
                        lignes_traitees += 1
                    except Exception as e:
                        lignes_erreur += 1
                        print(f"Erreur ligne {row_idx} (balance simple): {e}")
                        continue
                
                # Afficher le résumé du traitement
                print(f"Résumé du traitement (format balance simple):")
                print(f"Lignes traitées avec succès: {lignes_traitees}")
                if lignes_ignorees_colonnes > 0:
                    print(f"   Lignes ignorées (moins de 6 colonnes): {lignes_ignorees_colonnes}")
                if lignes_ignorees_numero_vide > 0:
                    print(f"   Lignes ignorées (numéro de compte vide): {lignes_ignorees_numero_vide}")
                if lignes_ignorees_vides > 0:
                    print(f"   Lignes vides détectées: {lignes_ignorees_vides}")
                if lignes_erreur > 0:
                    print(f"Lignes avec erreur: {lignes_erreur}")
                        
            else:
                # Format standard ou autre
                print("Traitement format standard...")
                header_row = 1
                
                # Trouver la ligne d'en-tête
                for row_idx in range(1, min(5, sheet.max_row + 1)):
                    row_data = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(8, sheet.max_column + 1))]
                    if any('compte' in str(cell).lower() for cell in row_data if cell):
                        header_row = row_idx
                        print(f"En-tête détecté à la ligne {row_idx}: {row_data}")
                        break
                
                # DÉTECTION DE L'ORDRE DES COLONNES (une seule fois avant la boucle)
                # Format 1 (standard) : debit_fin, credit_fin (colonnes 7, 8)
                # Format 2 (balance avec solde crédit/débit) : credit_fin, debit_fin (colonnes 7, 8)
                header_row_data = [sheet.cell(row=header_row, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
                
                # Vérifier les colonnes 7 et 8 (index 6 et 7) pour détecter l'ordre
                col7_header = str(header_row_data[6] if len(header_row_data) > 6 else '').strip()
                col8_header = str(header_row_data[7] if len(header_row_data) > 7 else '').strip()
                col7_val = col7_header.lower()
                col8_val = col8_header.lower()
                
                print(f"DÉTECTION DE L'ORDRE DES COLONNES:")
                print(f"   Colonne 7 (index 6): '{col7_header}'")
                print(f"   Colonne 8 (index 7): '{col8_header}'")
                
                # Détecter si "solde crédit" ou "credit" est en colonne 7 et "solde débit" ou "debit" en colonne 8
                # Patterns possibles : "Solde crédit", "Solde Crédit", "SOLDE CREDIT", "Crédit", etc.
                has_solde_credit_col7 = (
                    ('solde' in col7_val and 'credit' in col7_val) or
                    ('solde' in col7_val and 'crédit' in col7_val) or
                    (col7_val and 'credit' in col7_val and 'debit' not in col7_val and 'débit' not in col7_val)
                )
                
                has_solde_debit_col8 = (
                    ('solde' in col8_val and 'debit' in col8_val) or
                    ('solde' in col8_val and 'débit' in col8_val) or
                    (col8_val and 'debit' in col8_val and 'credit' not in col8_val and 'crédit' not in col8_val)
                )
                
                # Détecter aussi le format inverse (débit puis crédit)
                has_debit_col7 = (
                    ('debit' in col7_val and 'credit' not in col7_val and 'crédit' not in col7_val) or
                    ('débit' in col7_val and 'credit' not in col7_val and 'crédit' not in col7_val)
                )
                
                has_credit_col8 = (
                    ('credit' in col8_val and 'debit' not in col8_val and 'débit' not in col8_val) or
                    ('crédit' in col8_val and 'debit' not in col8_val and 'débit' not in col8_val)
                )
                
                # Déterminer l'ordre
                colonnes_inversees = False
                if has_solde_credit_col7 and has_solde_debit_col8:
                    colonnes_inversees = True
                    print(f"Format détecté: Solde crédit (col 7) puis Solde débit (col 8) - COLONNES INVERSÉES")
                    print(f"   Mapping: credit_fin = row[6] (col 7), debit_fin = row[7] (col 8)")
                elif has_debit_col7 and has_credit_col8:
                    colonnes_inversees = False
                    print(f"Format détecté: Débit (col 7) puis Crédit (col 8) - FORMAT STANDARD")
                    print(f"   Mapping: debit_fin = row[6] (col 7), credit_fin = row[7] (col 8)")
                else:
                    # Par défaut, utiliser le format standard
                    colonnes_inversees = False
                    print(f"Format non reconnu, utilisation du format standard par défaut")
                    print(f"   Mapping: debit_fin = row[6] (col 7), credit_fin = row[7] (col 8)")
                
                # Compteurs pour le diagnostic
                lignes_traitees = 0
                lignes_ignorees_colonnes = 0
                lignes_ignorees_vides = 0
                lignes_ignorees_numero_vide = 0
                lignes_erreur = 0
                
                # Traitement des lignes de données
                for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    # Vérifier que la ligne a au moins 8 colonnes
                    if len(row) < 8:
                        lignes_ignorees_colonnes += 1
                        continue
                        
                    if row[0] is None and row[1] is None:
                        lignes_ignorees_vides += 1
                        break  # Fin des données
                        
                    # Vérifier que la ligne contient des données valides
                    if not row[0] or str(row[0]).strip() == '':
                        lignes_ignorees_numero_vide += 1
                        continue
                    
                    try:
                        ligne = {}
                        ligne['numero_compte'] = str(row[0])
                        ligne['libelle'] = row[1] or ''
                        ligne['debit_initial'] = _parse_number(row[2])
                        ligne['credit_initial'] = _parse_number(row[3])
                        ligne['debit_mvt'] = _parse_number(row[4])
                        ligne['credit_mvt'] = _parse_number(row[5])
                        
                        # Utiliser l'ordre détecté précédemment
                        if colonnes_inversees:
                            # Format avec "Solde crédit" (col 7) puis "Solde débit" (col 8)
                            ligne['credit_fin'] = _parse_number(row[6])
                            ligne['debit_fin'] = _parse_number(row[7])
                        else:
                            # Format standard avec "Débit fin" puis "Crédit fin"
                            ligne['debit_fin'] = _parse_number(row[6])
                            ligne['credit_fin'] = _parse_number(row[7])
                        
                        ligne['solde_reel'] = ligne['debit_fin'] - ligne['credit_fin']
                        ligne['solde'] = abs(ligne['solde_reel'])
                        ligne['sign_solde'] = "D" if ligne['debit_fin'] >= ligne['credit_fin'] else "C"
                        data.append(ligne)
                        lignes_traitees += 1
                    except Exception as e:
                        lignes_erreur += 1
                        print(f"Erreur ligne {row_idx} (standard): {e}")
                        continue
                
                # Afficher le résumé du traitement
                print(f"Résumé du traitement:")
                print(f"Lignes traitées avec succès: {lignes_traitees}")
                if lignes_ignorees_colonnes > 0:
                    print(f"   Lignes ignorées (moins de 8 colonnes): {lignes_ignorees_colonnes}")
                if lignes_ignorees_numero_vide > 0:
                    print(f"   Lignes ignorées (numéro de compte vide): {lignes_ignorees_numero_vide}")
                if lignes_ignorees_vides > 0:
                    print(f"   Lignes vides détectées: {lignes_ignorees_vides}")
                if lignes_erreur > 0:
                    print(f"Lignes avec erreur: {lignes_erreur}")
            
            print(f"{len(data)} lignes de données extraites de la feuille '{sheet_name}'")
            
            # Vérifier qu'au moins une ligne de données a été extraite
            if len(data) == 0:
                filename = balance.filename if hasattr(balance, 'filename') else 'fichier inconnu'
                format_info = ""
                if format_detecte == 'sage':
                    format_info = "Format Sage détecté: attendu 6 colonnes minimum (numero, libelle, solde_n1, mouvement_debit, mouvement_credit, solde_n)"
                elif format_detecte == 'balance_simple':
                    format_info = "Format balance simple détecté: attendu 6 colonnes (numero, libelle, debit_initial, credit_initial, debit_fin, credit_fin)"
                else:
                    format_info = "Format standard détecté: attendu 8 colonnes minimum (numero, libelle, debit_initial, credit_initial, debit_mvt, credit_mvt, debit_fin, credit_fin)"
                
                error_msg = (
                    f"Aucune donnée extraite du fichier '{filename}' (feuille '{sheet_name}'). "
                    f"Vérifiez que:\n"
                    f"- La feuille contient bien des données (au moins une ligne avec un numéro de compte)\n"
                    f"- {format_info}\n"
                    f"- Les numéros de compte ne sont pas vides et commencent par un chiffre\n"
                    f"- Les lignes vides ou sans numéro de compte sont ignorées"
                )
                print(f"{error_msg}")
                pending_error_msg = error_msg
                    
        except Exception as e:
            print(f"Erreur lors du traitement du fichier Excel: {str(e)}")
            # S'assurer que db est défini avant de lever l'exception
            # (au cas où l'exception contient une référence à db dans sa traceback)
            try:
                _ = db  # Vérifier que db existe
            except NameError:
                # Si db n'est pas défini, le redéfinir
                db = get_db()
            raise e

        # Double v?rification avant insertion
        if len(data) == 0:
            # Fallback: parsing tr?s permissif sur toutes les feuilles
            def _to_float(v):
                try:
                    return float(str(v).replace(' ', '').replace('Â ', '').replace(',', '.'))
                except Exception:
                    return 0.0

            def _looks_like_account(v):
                if v is None:
                    return False
                s = str(v).strip()
                if not s:
                    return False
                if 'compte' in s.lower():
                    return False
                return s.isdigit() and len(s) >= 2

            def _flex_parse_sheet(ws):
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return []
                max_cols = max(len(r) for r in rows if r) if rows else 0
                if max_cols == 0:
                    return []

                # detect account column (max numeric account-like values)
                col_scores = []
                for c in range(max_cols):
                    score = 0
                    for r in rows[:200]:
                        if c < len(r) and _looks_like_account(r[c]):
                            score += 1
                    col_scores.append(score)
                if not col_scores or max(col_scores) == 0:
                    return []
                acc_col = col_scores.index(max(col_scores))

                # detect libelle column (stringy)
                lib_scores = []
                for c in range(max_cols):
                    if c == acc_col:
                        lib_scores.append(-1)
                        continue
                    score = 0
                    for r in rows[:200]:
                        if c < len(r) and isinstance(r[c], str) and r[c].strip():
                            score += 1
                    lib_scores.append(score)
                lib_col = lib_scores.index(max(lib_scores)) if lib_scores and max(lib_scores) > 0 else None

                # detect numeric columns
                num_cols = []
                for c in range(max_cols):
                    score = 0
                    for r in rows[:200]:
                        if c < len(r) and isinstance(r[c], (int, float)):
                            score += 1
                        elif c < len(r) and isinstance(r[c], str):
                            s = r[c].strip().replace(' ', '').replace(' ', '')
                            if s.replace(',', '.').replace('-', '').isdigit():
                                score += 1
                    if score > 0:
                        num_cols.append((c, score))
                num_cols = [c for c, _ in sorted(num_cols, key=lambda x: x[0])]

                debit_initial_col = credit_initial_col = debit_mvt_col = credit_mvt_col = None
                debit_fin_col = credit_fin_col = None

                if len(num_cols) >= 2:
                    debit_fin_col, credit_fin_col = num_cols[-2], num_cols[-1]
                if len(num_cols) >= 4:
                    debit_initial_col, credit_initial_col = num_cols[0], num_cols[1]
                if len(num_cols) >= 6:
                    debit_mvt_col, credit_mvt_col = num_cols[2], num_cols[3]

                out = []
                for r in rows:
                    if acc_col >= len(r):
                        continue
                    if not _looks_like_account(r[acc_col]):
                        continue
                    numero = str(r[acc_col]).strip()
                    libelle = ''
                    if lib_col is not None and lib_col < len(r) and r[lib_col] is not None:
                        libelle = str(r[lib_col]).strip()

                    ligne = {
                        'numero_compte': numero,
                        'libelle': libelle,
                        'debit_initial': _to_float(r[debit_initial_col]) if debit_initial_col is not None and debit_initial_col < len(r) else 0.0,
                        'credit_initial': _to_float(r[credit_initial_col]) if credit_initial_col is not None and credit_initial_col < len(r) else 0.0,
                        'debit_mvt': _to_float(r[debit_mvt_col]) if debit_mvt_col is not None and debit_mvt_col < len(r) else 0.0,
                        'credit_mvt': _to_float(r[credit_mvt_col]) if credit_mvt_col is not None and credit_mvt_col < len(r) else 0.0,
                        'debit_fin': _to_float(r[debit_fin_col]) if debit_fin_col is not None and debit_fin_col < len(r) else 0.0,
                        'credit_fin': _to_float(r[credit_fin_col]) if credit_fin_col is not None and credit_fin_col < len(r) else 0.0,
                    }
                    ligne['solde_reel'] = ligne['debit_fin'] - ligne['credit_fin']
                    ligne['solde'] = abs(ligne['solde_reel'])
                    ligne['sign_solde'] = "D" if ligne['debit_fin'] >= ligne['credit_fin'] else "C"
                    out.append(ligne)
                return out

            best = []
            best_name = None
            for ws_name in workbook.sheetnames:
                ws = workbook[ws_name]
                cand = _flex_parse_sheet(ws)
                if len(cand) > len(best):
                    best = cand
                    best_name = ws_name

            if best:
                print(f"? Fallback: {len(best)} lignes extraites depuis '{best_name}' (parsing permissif)")
                data = best

        if len(data) == 0:
            filename = balance.filename if hasattr(balance, 'filename') else 'fichier inconnu'
            raise Exception(pending_error_msg or f"Aucune donn?e valide trouv?e dans le fichier '{filename}'. V?rifiez le format du fichier Excel.")

        # S'assurer que db est défini avant d'utiliser
        try:
            _ = db  # Vérifier que db existe
        except NameError:
            # Si db n'est pas défini, le redéfinir
            db = get_db()

        result = db.Balance.insert_one(
            {"id_client": id_client, "annee_balance": annee_auditee, "balance": data}
        )

        inserted_id = str(result.inserted_id)
        print(f"Balance insérée en base avec {len(data)} lignes (ID: {inserted_id})")
        return inserted_id, data

    # ---------- Grouping ----------
    # ---------- Grouping ----------
    # ---------- Grouping ----------
    def create_grouping(self, balances_rapprochee, referentiel="syscohada"):
        """
        Regroupe les comptes de la balance selon le référentiel SYSCOHADA.
        
        Les règles de regroupement sont définies par des préfixes de comptes.
        Pour les comptes présents à la fois à l'actif et au passif (ex: 42, 43, 44...),
        le classement dépend du signe du solde (débiteur = actif, créditeur = passif).
        
        Chaque entrée du résultat représente un grand groupe avec :
        - ref        : identifiant court du groupe
        - libelle    : intitulé du groupe en majuscules
        - section    : "actif", "passif" ou "pnl"
        - solde_n    : total période N
        - solde_n1   : total période N-1
        - variation  : solde_n - solde_n1
        - variation_percent
        - comptes    : liste des comptes détaillés (numero_compte, libelle, solde_n, solde_n1, variation)
        """
        if not isinstance(balances_rapprochee, list):
            raise TypeError("balances_rapprochee doit être une liste")

        print(f"Balances reçues : {len(balances_rapprochee)} comptes")

        # =====================================================================
        # DÉFINITION DES GROUPES SYSCOHADA
        # Chaque groupe : (ref, libelle, section, règles)
        # Une règle est un callable(numero, solde_n) -> bool
        # L'ordre est important : on prend le PREMIER groupe qui correspond.
        # =====================================================================

        def starts(numero, *prefixes):
            """Vrai si numero commence par l'un des préfixes donnés."""
            return any(numero.startswith(p) for p in prefixes)

        def _solde_pour_signe(solde):
            """
            Solde utilisé pour départager actif/passif:
            - si tuple/list => (solde_n, solde_n1) avec priorité à N puis N-1
            - sinon => valeur fournie (comportement legacy)
            """
            if isinstance(solde, (tuple, list)):
                solde_n = solde[0] if len(solde) > 0 else 0
                solde_n1 = solde[1] if len(solde) > 1 else 0
                return solde_n if (solde_n or 0) != 0 else (solde_n1 or 0)
            return solde or 0

        def is_debiteur(solde):
            return _solde_pour_signe(solde) >= 0

        def is_crediteur(solde):
            return _solde_pour_signe(solde) < 0

        GROUPES_RULES = [
            # ---- ACTIF ----
            {
                "libelle": "IMMOBILISATIONS INCORPORELLES",
                "section": "actif",
                "match": lambda n, s: starts(n, "21", "281", "291"),
            },
            {
                "libelle": "IMMOBILISATIONS CORPORELLES",
                "section": "actif",
                "match": lambda n, s: starts(n, "22", "23", "24", "25", "282", "283", "284", "292", "293", "294", "295"),
            },
            {
                "libelle": "IMMOBILISATIONS FINANCIÈRES",
                "section": "actif",
                "match": lambda n, s: starts(n, "26", "27", "296", "297"),
            },
            {
                "libelle": "STOCKS",
                "section": "actif",
                "match": lambda n, s: starts(n, "31", "32", "33", "34", "35", "36", "37", "38", "39"),
            },
            {
                "libelle": "FOURNISSEUR AVANCES VERSEES",
                "section": "actif",
                "match": lambda n, s: starts(n, "409"),
            },
            {
                "libelle": "CLIENTS",
                "section": "actif",
                "match": lambda n, s: starts(n, "41") and not starts(n, "419") and is_debiteur(s),
            },
            {
                "libelle": "AUTRES CRÉANCES",
                "section": "actif",
                # 42, 43, 44, 185, 45, 46, 47 sauf 47 débiteur
                "match": lambda n, s: starts(n, "42", "43", "44", "45", "46", "47", "185") and not starts(n, "478", "479") and is_debiteur(s),
            },
            {
                "libelle": "ACTIF CIRCULANT HAO",
                "section": "actif",
                "match": lambda n, s: starts(n, "485", "488", "498"),
            },
            {
                "libelle": "TITRES DE PLACEMENT",
                "section": "actif",
                "match": lambda n, s: starts(n, "50", "590"),
            },
            {
                "libelle": "VALEUR À ENCAISSER",
                "section": "actif",
                "match": lambda n, s: starts(n, "51", "591"),
            },
            {
                "libelle": "BANQUES, CHÈQUES POSTAUX, CAISSES ET ASSIMILÉ",
                "section": "actif",
                # 52, 53, 54, 55, 57, 58 débiteurs
                "match": lambda n, s: starts(n, "52", "53", "54", "55", "57", "58") and is_debiteur(s),
            },
            {
                "libelle": "ÉCART DE CONVERSION ACTIF",
                "section": "actif",
                "match": lambda n, s: starts(n, "478"),
            },

            # ---- PASSIF ----
            {
                "libelle": "FOURNISSEURS D'EXPLOITATION",
                "section": "passif",
                # Tous les 40* (hors 409*) appartiennent à ce grand groupe,
                # même si le solde est débiteur, créditeur ou nul.
                "match": lambda n, s: starts(n, "40") and not starts(n, "409"),
            },
            {
                "libelle": "DETTES FISCALES ET SOCIALES",
                "section": "passif",
                # 42, 43, 44 créditeurs
                "match": lambda n, s: starts(n, "42", "43", "44") and is_crediteur(s),
            },
            {
                "libelle": "AUTRES DETTES",
                "section": "passif",
                # 185, 45, 46, 47 sauf 479 créditeurs
                "match": lambda n, s: starts(n, "185", "45", "46", "47") and not starts(n, "478","479") and is_crediteur(s),
            },
            {
                "libelle": "DETTES CIRCULANT HAO",
                "section": "passif",
                "match": lambda n, s: starts(n, "481", "482", "484", "4998"),
            },
            {
                "libelle": "PROVISIONS POUR RISQUES À COURT TERME",
                "section": "passif",
                # 499 sauf 4998, 599
                "match": lambda n, s: starts(n, "499", "599") and not starts(n, "4998"),
            },
            {
                "libelle": "BANQUES, CHÈQUES POSTAUX, CAISSES ET ASSIMILÉ",
                "section": "passif",
                # 52, 53, 54, 55, 566, 57, 58 créditeurs
                "match": lambda n, s: starts(n, "52", "53", "54", "55", "566", "57", "58") and is_crediteur(s),
            },
            {
                "libelle": "BANQUES, CRÉDIT D'ESCOMPTE ET DE TRÉSORERIE",
                "section": "passif",
                "match": lambda n, s: starts(n, "564", "565"),
            },
            {
                "libelle": "CLIENTS AVANCES REÇUES",
                "section": "passif",
                # 419 = avances reçues clients (passif), quel que soit le signe importé.
                "match": lambda n, s: starts(n, "419"),
            },
            {
                "libelle": "ÉCART DE CONVERSION PASSIF",
                "section": "passif",
                "match": lambda n, s: starts(n, "479"),
            },
            {
                "libelle": "CAPITAUX PROPRES",
                "section": "passif",
                "match": lambda n, s: starts(n, "10", "11", "12", "13"),
            },
            {
                "libelle": "SUBVENTIONS",
                "section": "passif",
                "match": lambda n, s: starts(n, "14"),
            },
            {
                "libelle": "PROVISIONS RÉGLEMENTÉES",
                "section": "passif",
                "match": lambda n, s: starts(n, "15"),
            },
            {
                "libelle": "EMPRUNTS ET DETTES FINANCIÈRES DIVERSES",
                "section": "passif",
                "match": lambda n, s: starts(n, "16", "18"),
            },
            {
                "libelle": "DETTES DE LOCATIONS ACQUISITION",
                "section": "passif",
                "match": lambda n, s: starts(n, "17"),
            },
            {
                "libelle": "PROVISIONS FINANCIÈRES POUR RISQUES ET CHARGES",
                "section": "passif",
                "match": lambda n, s: starts(n, "19"),
            },

            # ---- COMPTE DE RÉSULTAT ----
            {
                "libelle": "ACHAT DE MARCHANDISES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "601"),
            },
            {
                "libelle": "ACHAT DE MATIÈRE PREMIÈRE ET FOURNITURES LIÉES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "602"),
            },
            {
                "libelle": "VARIATION DE STOCKS DE MARCHANDISES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "6031"),
            },
            {
                "libelle": "VARIATION DE STOCKS DE MATIÈRE PREMIÈRE ET FOURNITURES LIÉES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "6032"),
            },
            {
                "libelle": "VARIATION DE STOCKS D'AUTRES APPROVISIONNEMENT",
                "section": "pnl",
                "match": lambda n, s: starts(n, "6033"),
            },
            {
                "libelle": "AUTRES ACHATS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "604", "605", "608"),
            },
            {
                "libelle": "TRANSPORTS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "61"),
            },
            {
                "libelle": "SERVICES EXTÉRIEURS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "62", "63"),
            },
            {
                "libelle": "IMPÔTS ET TAXES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "64"),
            },
            {
                "libelle": "AUTRES CHARGES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "65"),
            },
            {
                "libelle": "CHARGES DE PERSONNEL",
                "section": "pnl",
                "match": lambda n, s: starts(n, "66"),
            },
            {
                "libelle": "FRAIS FINANCIERS ET CHARGES ASSIMILÉS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "67"),
            },
            {
                "libelle": "DOTATIONS AUX AMORTISSEMENTS ET AUX PROVISIONS ET DÉPRÉCIATIONS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "681", "691"),
            },
            {
                "libelle": "DOTATIONS AUX PROVISIONS ET AUX DÉPRÉCIATIONS FINANCIÈRES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "697"),
            },
            {
                "libelle": "VENTES DE MARCHANDISES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "701"),
            },
            {
                "libelle": "VENTE DE PRODUITS FABRIQUÉS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "702", "703", "704"),
            },
            {
                "libelle": "TRAVAUX, SERVICES VENDUS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "705", "706"),
            },
            {
                "libelle": "PRODUITS ACCESSOIRES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "707"),
            },
            {
                "libelle": "SUBVENTIONS D'EXPLOITATION",
                "section": "pnl",
                "match": lambda n, s: starts(n, "71"),
            },
            {
                "libelle": "PRODUCTION IMMOBILISÉE",
                "section": "pnl",
                "match": lambda n, s: starts(n, "72"),
            },
            {
                "libelle": "PRODUCTION STOCKÉE (OU DESTOCKAGE)",
                "section": "pnl",
                "match": lambda n, s: starts(n, "73"),
            },
            {
                "libelle": "AUTRES PRODUITS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "75"),
            },
            {
                "libelle": "REVENUS FINANCIERS ET ASSIMILÉS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "77"),
            },
            {
                "libelle": "TRANSFERT DE CHARGES D'EXPLOITATION",
                "section": "pnl",
                "match": lambda n, s: starts(n, "781"),
            },
            {
                "libelle": "TRANSFERT DE CHARGES FINANCIÈRES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "787"),
            },
            {
                "libelle": "REPRISES D'AMORTISSEMENT, DE PROVISIONS ET DÉPRÉCIATION",
                "section": "pnl",
                "match": lambda n, s: starts(n, "791", "798", "799"),
            },
            {
                "libelle": "REPRISE DE PROVISIONS ET DÉPRÉCIATIONS FINANCIÈRES",
                "section": "pnl",
                "match": lambda n, s: starts(n, "797"),
            },
            {
                "libelle": "VALEURS COMPTABLES DES CESSIONS D'IMMOBILISATIONS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "81"),
            },
            {
                "libelle": "PRODUITS DES CESSIONS D'IMMOBILISATIONS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "82"),
            },
            {
                "libelle": "AUTRES CHARGES HAO",
                "section": "pnl",
                "match": lambda n, s: starts(n, "83", "85"),
            },
            {
                "libelle": "AUTRES PRODUITS HAO",
                "section": "pnl",
                "match": lambda n, s: starts(n, "84", "86", "88"),
            },
            {
                "libelle": "PARTICIPATION DES TRAVAILLEURS",
                "section": "pnl",
                "match": lambda n, s: starts(n, "87"),
            },
            {
                "libelle": "IMPÔTS SUR LE RÉSULTAT",
                "section": "pnl",
                "match": lambda n, s: starts(n, "89"),
            },
        ]

        # =====================================================================
        # Initialiser un dictionnaire de groupes vide
        # =====================================================================
        groupes_data = {}
        for g in GROUPES_RULES:
            key = g["libelle"] + "|" + g["section"]  # clé interne unique
            groupes_data[key] = {
                "libelle": g["libelle"],
                "section": g["section"],
                "solde_n": 0,
                "solde_n1": 0,
                "comptes": [],
                "_match": g["match"],
            }

        # =====================================================================
        # Classifier chaque compte dans son groupe
        # =====================================================================
        non_classes = []

        def sign_of(value):
            v = value or 0
            if v > 0:
                return 1
            if v < 0:
                return -1
            return 0

        def is_sign_sensitive_account(numero_compte):
            # Comptes dont la bascule actif/passif dépend du sens du solde.
            return (
                starts(numero_compte, "41", "42", "43", "44", "45", "46", "47", "185", "52", "53", "54", "55", "57", "58")
                and not starts(numero_compte, "419", "478", "479")
            )

        def classify_and_append(numero_compte, libelle_compte, match_solde, add_solde_n, add_solde_n1):
            for _, g in groupes_data.items():
                try:
                    if g["_match"](numero_compte, match_solde):
                        g["solde_n"] += add_solde_n
                        g["solde_n1"] += add_solde_n1
                        g["comptes"].append({
                            "numero_compte": numero_compte,
                            "libelle": libelle_compte,
                            "solde_n": add_solde_n,
                            "solde_n1": add_solde_n1,
                            "variation": add_solde_n - add_solde_n1,
                        })
                        return True
                except Exception:
                    continue
            return False

        # for item in balances_rapprochee:
        #     numero = str(item.get("numero_compte", "")).strip()
        #     if not numero or not numero[0].isdigit():
        #         continue

        #     solde_n = item.get("solde_n", 0) or 0
        #     solde_n1 = item.get("solde_n1", 0) or 0
        #     libelle = item.get("libelle", "")

        for item in balances_rapprochee:
            numero = str(item.get("numero_compte", "")).strip()
            if len(numero) < 2:
                continue

            prefixe = numero[:2]
            if not prefixe.isdigit():
                continue

            solde_n = item.get("solde_n", 0) or 0
            solde_n1 = item.get("solde_n1", 0) or 0
            libelle = item.get("libelle", "")

            # CORRECTION BUG 419 : réaffectation des comptes 419 créditeurs
            # 419 = avances reçues de client passif circulant (DI)
            # Si solde créditeur (négatif en SYSCOHADA débiteur-normal), on le bascule
            if numero.startswith("419") and solde_n < 0:
                target_key = "DI_419"  # groupe spécial à créer, voir ci-dessous
                # OU plus simple : l'affecter à un groupe dédié dans le référentiel

            # Si le sens change entre N et N-1 pour un compte sensible au signe,
            # on classe chaque période séparément pour refléter correctement
            # actif/passif sur chaque colonne.
            sign_n = sign_of(solde_n)
            sign_n1 = sign_of(solde_n1)
            if is_sign_sensitive_account(numero) and sign_n != 0 and sign_n1 != 0 and sign_n != sign_n1:
                matched_n = classify_and_append(numero, libelle, solde_n, solde_n, 0)
                matched_n1 = classify_and_append(numero, libelle, solde_n1, 0, solde_n1)
                if not (matched_n and matched_n1):
                    non_classes.append(numero)
                continue

            if not classify_and_append(numero, libelle, (solde_n, solde_n1), solde_n, solde_n1):
                non_classes.append(numero)

        if non_classes:
            print(f"{len(non_classes)} compte(s) non classés : {non_classes[:20]}")

        # =====================================================================
        # Finaliser et construire le résultat
        # =====================================================================
        result = []
        for g in groupes_data.values():
            if not g["comptes"]:
                continue

            g["comptes"].sort(key=lambda x: x["numero_compte"])

            variation = g["solde_n"] - g["solde_n1"]
            variation_percent = (variation / g["solde_n1"] * 100) if g["solde_n1"] else 0

            result.append({
                "libelle": g["libelle"],
                "section": g["section"],
                "solde_n": g["solde_n"],
                "solde_n1": g["solde_n1"],
                "variation": variation,
                "variation_percent": variation_percent,
                "comptes": g["comptes"],
            })

        print(f"Grouping généré : {len(result)} groupes")
        return result


    # ---------- Variation N vs N-1 ----------
    def rapprochement_des_balances(self, balance_n, balance_n1):
        variation_des_balances = []

        def _index_balance(balance_rows):
            """
            Indexe une balance par (numéro de compte, libellé) pour conserver
            les lignes distinctes d'un même compte avec des libellés différents
            (ex: 409100 + "FOURNISSEURS DIVERS" / "F/SSEURS AVANCE ET ACOMPTE").
            """
            index = {}
            for item in balance_rows or []:
                numero = str(item.get('numero_compte', '')).strip()
                if not numero:
                    continue

                solde = item.get('solde_reel', 0) or 0
                libelle = str(item.get('libelle', '') or '').strip()
                key = (numero, libelle)

                if key not in index:
                    index[key] = {
                        'numero_compte': numero,
                        'libelle': libelle,
                        'solde_reel': solde
                    }
                else:
                    index[key]['solde_reel'] += solde
            return index

        idx_n = _index_balance(balance_n)
        idx_n1 = _index_balance(balance_n1)

        # Inclure toutes les lignes présentes en N ou en N-1 par couple
        # (numero_compte, libelle), sans fusionner des libellés différents.
        all_keys = list(dict.fromkeys(list(idx_n.keys()) + list(idx_n1.keys())))

        for numero, libelle in all_keys:
            item_n = idx_n.get((numero, libelle), {})
            item_n1 = idx_n1.get((numero, libelle), {})
            ligne = {
                'numero_compte': numero,
                'libelle': item_n.get('libelle') or item_n1.get('libelle') or libelle or '',
                'solde_n': item_n.get('solde_reel', 0) or 0,
                'solde_n1': item_n1.get('solde_reel', 0) or 0
            }
            variation_des_balances.append(ligne)

        return variation_des_balances

    # ---------- Charges ----------
    def total_charges(self, id_mission):
        db = get_db()  # Obtenir la connexion à la base de données
        balance = db.Mission1.find_one({"_id": ObjectId(id_mission)})['balance_variation']
        charges = sum(item['solde_n'] for item in balance if str(item['numero_compte']).startswith('6'))
        return abs(charges)

    # ---------- Benchmarks ----------
    def get_benchmarks(self, id_mission):
        db = get_db()  # Obtenir la connexion ? la base de donn?es
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            efi = mission.get('efi') or mission.get('etats_financiers_preliminaires', {}).get('efi') or {}

            def _find_efi_value(section, keywords):
                try:
                    items = efi.get(section, [])
                    for it in items:
                        lib = str(it.get('libelle', '')).lower()
                        if any(k in lib for k in keywords):
                            return int(it.get('net_solde_n', 0) or 0)
                except Exception:
                    return None
                return None

            def _find_efi_value_by_ref(section, refs):
                try:
                    items = efi.get(section, [])
                    refs_set = {str(r).strip().upper() for r in refs}
                    for it in items:
                        ref = str(it.get('ref', '')).strip().upper()
                        if ref in refs_set:
                            return int(it.get('net_solde_n', 0) or 0)
                except Exception:
                    return None
                return None

            bench = {}
            # Chaque benchmark est alimenté par une référence EFI dédiée (mapping_efi.json),
            # avec fallback conservateur si la ref n'est pas trouvée.
            total_assets = _find_efi_value_by_ref('actif', ['BZ'])
            if total_assets is None:
                try:
                    total_assets = int(efi.get('actif', [])[-1].get('net_solde_n', 0) or 0)
                except Exception:
                    total_assets = 0
            bench['total_assets'] = int(total_assets or 0)

            revenue = _find_efi_value_by_ref('pnl', ['XB'])
            if revenue is None:
                try:
                    revenue = int(efi.get('pnl', [])[7].get('net_solde_n', 0) or 0)
                except Exception:
                    revenue = 0
            bench['revenue'] = int(revenue or 0)

            ebitda = _find_efi_value_by_ref('pnl', ['XD'])
            if ebitda is None:
                ebitda = _find_efi_value('pnl', ['excedent brut d\'exploitation', 'ebe', 'ebitda'])
            bench['ebitda'] = int(ebitda or 0)

            net_result = _find_efi_value_by_ref('pnl', ['XI'])
            taxes_result = _find_efi_value_by_ref('pnl', ['RS'])
            if net_result is not None and taxes_result is not None:
                profit_before_tax = int(net_result) - int(taxes_result)
            else:
                try:
                    profit_before_tax = int(efi.get('pnl', [])[33].get('net_solde_n', 0) or 0) + int(efi.get('pnl', [])[38].get('net_solde_n', 0) or 0)
                except Exception:
                    profit_before_tax = 0
            bench['profit_before_tax'] = int(profit_before_tax or 0)

            bench['expenses'] = self.total_charges(id_mission)

            total_equity = _find_efi_value_by_ref('passif', ['CP'])
            if total_equity is None:
                total_equity = _find_efi_value('passif', ['capitaux propres', 'equity', 'net assets'])
            bench['total_equity_net_assets'] = int(total_equity or 0)

            # Proxy "cash flows from operations" basé sur le résultat d'exploitation (ref XE),
            # distinct de l'EBITDA/EBE (ref XD).
            cfo_value = _find_efi_value_by_ref('pnl', ['XE'])
            if cfo_value is None:
                cfo_value = _find_efi_value('pnl', ['cash flow', 'flux de tresorerie', 'flux de trésorerie'])
            bench['cash_flows_from_operations'] = int(cfo_value or 0)
            return bench
        except Exception as e:
            print(f"An error there: {str(e)}")
            return None

    # ---------- Materiality ----------
    def save_materiality(self, id_mission, materialities):
        db = get_db()  # Obtenir la connexion à la base de données
        query = {"_id": ObjectId(id_mission)}
        verify = db.Mission1.find_one(query)

        if not verify:
            return 0

        # Validation des facteurs par benchmark (ex: 5-10% => exclut 1-4)
        factor_ranges = {
            "ebitda": (3, 5),
            "expenses": (3, 5),
            "profit_before_tax": (5, 10),
            "revenue": (0.8, 2),
            "total_assets": (1, 2),
            "total_equity_net_assets": (1.0, 3.0),
            "cash_flows_from_operations": (3.0, 5.0)
        }
        try:
            benchmark = materialities.get("benchmark")
            factor_raw = materialities.get("factor")
            if benchmark in factor_ranges and factor_raw is not None:
                factor = float(factor_raw)
                min_f, max_f = factor_ranges[benchmark]
                if factor < min_f or factor > max_f:
                    print(f"WARNING: Facteur {factor}% hors plage [{min_f}, {max_f}] pour {benchmark}")
                    return 0
        except Exception:
            pass

        # Vérifier si le benchmark existe déjà
        existing_materiality = verify.get('materiality', [])
        benchmark_exists = any(item.get('benchmark') == materialities['benchmark'] for item in existing_materiality)

        # Assurer un champ commentaire (optionnel)
        if 'commentaire' not in materialities:
            materialities['commentaire'] = ""

        # Calcul de la performance materiality selon le niveau de risque (si fourni)
        try:
            risk_level = (materialities.get('risk_level') or '').lower().strip()
            mat_val = float(materialities.get('materiality', 0) or 0)
            if risk_level in ('?lev?', 'eleve'):
                materialities['performance_materiality'] = mat_val * 0.6
            elif risk_level in ('mod?r?', 'modere'):
                materialities['performance_materiality'] = mat_val * 0.7
            elif risk_level == 'faible':
                materialities['performance_materiality'] = mat_val * 0.8
        except Exception:
            pass

        # Forcer des valeurs toujours positives pour les seuils (r?gle m?tier)
        try:
            materialities['materiality'] = abs(int(materialities.get('materiality', 0) or 0))
            materialities['performance_materiality'] = abs(int(materialities.get('performance_materiality', 0) or 0))
            materialities['trivial_misstatements'] = abs(int(materialities.get('trivial_misstatements', 0) or 0))
        except Exception:
            pass

        if benchmark_exists:
            # Mettre ? jour le benchmark existant
            update = db.Mission1.update_one(
                {"_id": ObjectId(id_mission), "materiality.benchmark": materialities['benchmark']},
                {"$set": {"materiality.$": materialities}}
            )
            # Consid?rer comme succ?s m?me si aucune modification effective
            return 1 if update.matched_count > 0 else 0
        else:
            # Ajouter un nouveau benchmark
            update = db.Mission1.update_one(query, {"$push": {"materiality": materialities}})
            return 1 if update.matched_count > 0 else 0

    def validate_materiality(self, id_mission, benchmark):
        db = get_db()  # Obtenir la connexion à la base de données
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        materialities = mission['materiality']

        exist_choice = next((item for item in materialities if item.get('choice')), None)
        if not exist_choice:
            for item in materialities:
                if item['benchmark'] == benchmark:
                    item['choice'] = True
        else:
            exist_choice['choice'] = ""
            for item in materialities:
                if item['benchmark'] == benchmark:
                    item['choice'] = True

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"materiality": materialities}})
        return result.modified_count

    def get_materiality(self, id_mission):
        db = get_db()  # Obtenir la connexion à la base de données
        materiality = db.Mission1.find_one({"_id": ObjectId(id_mission)}, {"_id": 0, "materiality": 1})
        return materiality

    def materialite(self, id_mission):
        """
        Méthode pour récupérer et calculer les données de matérialité
        """
        db = get_db()  # Obtenir la connexion à la base de données
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "materiality": []}

            # Récupérer les données de matérialité existantes
            materiality = mission.get("materiality", [])
            
            # Si aucune donnée de matérialité, calculer les benchmarks
            if not materiality:
                print(f"Calcul des benchmarks de matérialité pour la mission {id_mission}")
                materiality = self._calculate_materiality_benchmarks(mission)
                
                # Sauvegarder les calculs si des benchmarks ont été trouvés
                if materiality:
                    db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"materiality": materiality}})
                    print(f"{len(materiality)} benchmark(s) sauvegardé(s) dans la base de données")

            if not materiality:
                # Vérifier la cause du problème pour donner un message plus explicite
                efi = mission.get("efi", {})
                if not efi:
                    message = "Impossible de calculer les seuils de matérialité. Les états financiers préliminaires n'ont pas été générés. Veuillez d'abord compléter l'étape 5 (États financiers préliminaires)."
                elif not isinstance(efi, dict) or not any(section in efi for section in ["actif", "passif", "pnl"]):
                    message = "Impossible de calculer les seuils de matérialité. Les états financiers sont incomplets ou mal formatés. Veuillez régénérer les états financiers préliminaires."
                else:
                    message = "Impossible de calculer les seuils de matérialité. Les données financières nécessaires (profit_before_tax, ebitda, revenue, total_assets, total_expenses) n'ont pas été trouvées dans les états financiers. Vérifiez que les balances contiennent les informations nécessaires."
                
                return {"ok": False, "message": message, "materiality": []}

            return {
                "ok": True,
                "message": "Données de matérialité récupérées avec succès",
                "materiality": materiality
            }

        except Exception as e:
            import traceback
            print(f"Erreur dans materialite: {e}")
            traceback.print_exc()
            return {"ok": False, "message": f"Erreur lors du calcul de la matérialité: {str(e)}", "materiality": []}

    def _calculate_materiality_benchmarks(self, mission):
        """
        Calcule les benchmarks de matérialité basés sur les données financières
        """
        try:
            # Récupérer les données financières
            efi = mission.get("efi", {})
            if not efi:
                print("Aucune donnée EFI trouvée dans la mission")
                print("Conseil: Veuillez générer les états financiers préliminaires (Étape 5) avant de calculer les matérialités")
                return []

            # Vérifier que les données EFI sont structurées correctement
            if not isinstance(efi, dict):
                print(f"Format de données EFI invalide: {type(efi)}")
                return []
            
            # Vérifier que les sections principales existent
            sections_presentes = [section for section in ["actif", "passif", "pnl"] if section in efi and isinstance(efi[section], list)]
            if not sections_presentes:
                print("Aucune section EFI valide trouvée (actif, passif, pnl)")
                return []
            
            print(f"Sections EFI trouvées: {sections_presentes}")
            
            benchmarks = []
            valeurs_trouvees = []
            
            # Benchmark 1: Profit Before Tax
            profit_before_tax = self._get_financial_value(efi, "profit_before_tax")
            if profit_before_tax and profit_before_tax != 0:
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                profit_before_tax = profit_before_tax * -1
                materiality = profit_before_tax * 0.05  # 5% du bénéfice avant impôt
                # Le seuil de matérialité est toujours positif (valeur absolue)
                abs_materiality = abs(materiality)
                is_negative = materiality < 0
                
                benchmarks.append({
                    "benchmark": "profit_before_tax",
                    "factor": "5%",
                    "materiality": int(abs_materiality),  # Toujours positif
                    "performance_materiality": int(abs_materiality * 0.8),
                    "trivial_misstatements": int(abs_materiality * 0.05),
                    "choice": "",
                    "warning": "ATTENTION: Seuil de matérialité négatif !" if is_negative else "",
                    "original_value": profit_before_tax
                })
                valeurs_trouvees.append(f"profit_before_tax: {profit_before_tax}")
            else:
                print(f"Valeur profit_before_tax non trouvée ou nulle: {profit_before_tax}")

            # Benchmark 2: EBITDA
            ebitda = self._get_financial_value(efi, "ebitda")
            if ebitda and ebitda != 0:
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                ebitda = ebitda * -1
                materiality = ebitda * 0.05  # 5% de l'EBITDA
                # Le seuil de matérialité est toujours positif (valeur absolue)
                abs_materiality = abs(materiality)
                is_negative = materiality < 0
                
                benchmarks.append({
                    "benchmark": "ebitda",
                    "factor": "5%",
                    "materiality": int(abs_materiality),  # Toujours positif
                    "performance_materiality": int(abs_materiality * 0.8),
                    "trivial_misstatements": int(abs_materiality * 0.05),
                    "choice": "",
                    "warning": "ATTENTION: Seuil de matérialité négatif !" if is_negative else "",
                    "original_value": ebitda
                })
                valeurs_trouvees.append(f"ebitda: {ebitda}")
            else:
                print(f"Valeur ebitda non trouvée ou nulle: {ebitda}")

            # Benchmark 3: Revenue
            revenue = self._get_financial_value(efi, "revenue")
            if revenue and revenue != 0:
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                revenue = revenue * -1
                materiality = revenue * 0.01  # 1% du chiffre d'affaires
                # Le seuil de matérialité est toujours positif (valeur absolue)
                abs_materiality = abs(materiality)
                is_negative = materiality < 0
                
                benchmarks.append({
                    "benchmark": "revenue",
                    "factor": "1%",
                    "materiality": int(abs_materiality),  # Toujours positif
                    "performance_materiality": int(abs_materiality * 0.8),
                    "trivial_misstatements": int(abs_materiality * 0.05),
                    "choice": "",
                    "warning": "ATTENTION: Seuil de matérialité négatif !" if is_negative else "",
                    "original_value": revenue
                })
                valeurs_trouvees.append(f"revenue: {revenue}")
            else:
                print(f"Valeur revenue non trouvée ou nulle: {revenue}")

            # Benchmark 4: Total Assets
            total_assets = self._get_financial_value(efi, "total_assets")
            if total_assets and total_assets > 0:  # Les actifs ne peuvent pas être négatifs
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                total_assets = total_assets * -1
                materiality = total_assets * 0.01  # 1% du total des actifs
                benchmarks.append({
                    "benchmark": "total_assets",
                    "factor": "1%",
                    "materiality": max(int(materiality), 0),
                    "performance_materiality": max(int(materiality * 0.8), 0),
                    "trivial_misstatements": max(int(materiality * 0.05), 0),
                    "choice": "",
                    "warning": "",
                    "original_value": total_assets
                })
                valeurs_trouvees.append(f"total_assets: {total_assets}")
            else:
                print(f"Valeur total_assets non trouvée ou nulle: {total_assets}")

            # Benchmark 5: Total Expenses (Charges)
            total_expenses = self._get_financial_value(efi, "total_expenses")
            if total_expenses and total_expenses > 0:  # Les charges doivent toujours être positives
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                total_expenses = total_expenses * -1
                materiality = total_expenses * 0.02  # 2% du total des charges
                benchmarks.append({
                    "benchmark": "total_expenses",
                    "factor": "2%",
                    "materiality": int(materiality),  # Toujours positif
                    "performance_materiality": int(materiality * 0.8),
                    "trivial_misstatements": int(materiality * 0.05),
                    "choice": "",
                    "warning": "",
                    "original_value": total_expenses
                })
                valeurs_trouvees.append(f"total_expenses: {total_expenses}")
            else:
                print(f"Valeur total_expenses non trouvée ou nulle: {total_expenses}")

            if benchmarks:
                print(f"{len(benchmarks)} benchmark(s) calculé(s) avec succès: {valeurs_trouvees}")
            else:
                print("Aucun benchmark n'a pu être calculé. Vérifiez que les états financiers contiennent les données nécessaires.")
                print("Les valeurs recherchées sont: profit_before_tax, ebitda, revenue, total_assets, total_expenses")

            return benchmarks

        except Exception as e:
            import traceback
            print(f"Erreur lors du calcul des benchmarks: {e}")
            traceback.print_exc()
            return []

    def _get_financial_value(self, efi, key):
        """
        Récupère une valeur financière depuis les états financiers
        """
        try:
            # Mapping des clés vers les libellés français ou calculs
            key_mappings = {
                "profit_before_tax": ["bénéfice avant impôt", "résultat avant impôt", "profit", "bénéfice"],
                "ebitda": ["ebitda", "excédent brut d'exploitation", "ebe"],
                "revenue": ["chiffre d'affaires", "revenus", "recettes", "ventes", "ca"],
                "total_assets": ["total actif", "total de l'actif"],
                "total_expenses": ["total charges", "charges totales", "total des charges"]
            }
            
            # Chercher dans les différents états financiers
            for section in ["actif", "passif", "pnl"]:
                if section in efi and isinstance(efi[section], list):
                    for item in efi[section]:
                        ref = item.get("ref", "").lower()
                        libelle = item.get("libelle", "").lower()
                        net_solde = item.get("net_solde_n", 0)
                        
                        # Vérifier si la ref correspond
                        if ref == key or ref == key.replace("_", ""):
                            # Retourner la valeur brute (avec son signe) pour pouvoir multiplier par -1
                            return net_solde if net_solde else None
                        
                        # Vérifier si le libellé correspond aux mappings
                        if key in key_mappings:
                            for search_term in key_mappings[key]:
                                if search_term in libelle:
                                    # Retourner la valeur brute (avec son signe) pour pouvoir multiplier par -1
                                    return net_solde if net_solde else None
            
            # Calculs spécifiques si les valeurs directes ne sont pas trouvées
            if key == "total_assets" and "actif" in efi:
                # Calculer le total actif en sommant tous les actifs (avec leur signe)
                # Les actifs peuvent être négatifs dans la balance, donc on garde le signe
                total = sum(item.get("net_solde_n", 0) for item in efi.get("actif", []) if item.get("net_solde_n", 0))
                return total if total != 0 else None
                
            if key == "total_expenses" and "pnl" in efi:
                # Chercher les charges dans le compte de résultat (avec leur signe)
                charges_keywords = ["charge", "dépense", "coût"]
                total = 0
                for item in efi.get("pnl", []):
                    libelle = item.get("libelle", "").lower()
                    if any(kw in libelle for kw in charges_keywords):
                        solde = item.get("net_solde_n", 0)
                        # Garder le signe pour pouvoir multiplier par -1 ensuite
                        total += solde
                return abs(total) if total != 0 else None
            
            return None
        except Exception as e:
            print(f"Erreur dans _get_financial_value pour {key}: {e}")
            return None

    # ---------- Analyses grouping ----------
    def make_quantitative_analysis(self, id_mission):
        db = get_db()  # Obtenir la connexion à la base de données
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']
        materiality = next((mat for mat in mission['materiality'] if mat.get('choice')), None)

        if materiality is not None:
            for item in grouping:
                value = False
                if int(item['solde_n']) >= int(materiality['materiality']):
                    value = True
                item['materiality'] = value

            result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
            return result.modified_count
        else:
            return 0

    def make_qualitative_analysis(self, id_mission, significant):
        db = get_db()  # Obtenir la connexion à la base de données
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']

        for item in grouping:
            value = next((elt['significant'] for elt in significant if elt.get('compte') == item.get('compte')), None)
            if value is None:
                item['significant'] = False
            else:
                item['significant'] = value

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
        return result.modified_count

    def make_final_sm(self, id_mission):
        db = get_db()  # Obtenir la connexion à la base de données
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']

        for item in grouping:
            mat = item.get('materiality', None)
            sign = item.get('significant', None)

            if (mat is not None) and (sign is not None):
                value = ""
                if mat is False and sign is True:
                    value = "non matériel significatif"
                elif mat is False and sign is False:
                    value = "non matériel non significatif"
                elif mat is True and sign is True:
                    value = "matériel significatif"
                elif mat is True and sign is False:
                    value = "matériel non significatif"
                else:
                    value = None
                item['mat_sign'] = value

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
        return result.modified_count, grouping

    # ---------- Récup mission ----------
    def afficher_informations_missions(self, id_client):
        db = get_db()  # Obtenir la connexion à la base de données
        _id = id_client
        query = db.Mission1.find_one({"_id": _id})
        query['_id'] = str(query['_id'])
        return query

    # ---------- Production COMMENTAIRE ----------
    def prod_efi(self, balance_n, balance_n1, balance_variation):
        """
        Produit les états financiers préliminaires (EFI) SYSCOHADA en se basant
        sur des règles de préfixes de comptes codées en dur.
        Plus aucune dépendance au fichier mapping_efi.json.

        Structure de chaque ligne retournée :
        ref, libelle, note, brut_solde_n, amor_solde_n, net_solde_n, net_solde_n1,
        compte_to_be_used, compte_to_be_used_off

        Les lignes agrégées (AD, AI, AZ) sont calculées en fin de méthode par
        compute_etats_financiers().
        """

        # ------------------------------------------------------------------
        # CATALOGUE DES LIGNES SYSCOHADA
        # Chaque entrée : (ref, libelle, nature, note, brut_cpts, amor_cpts,
        #                  net_cpts, brut_except, amor_except, net_except)
        # brut + amor  net = brut + amor   (amortissements sont négatifs)
        # net seul     net = somme des comptes net_cpts
        # except_cpts  comptes exclus du calcul principal puis ré-inclus
        #                    séparément (voir logique de prod_efi originale)
        # ------------------------------------------------------------------
        EFI_CATALOGUE = [
            # ACTIF 
            # Immobilisations incorporelles (lignes feuilles)
            ("AE", "Frais de développement et de prospection",               "actif", None,
            ["211","2181","2191"], ["2811","2818","2911","2918","2919"],      [], ["2181"], [],  ["2181"]),
            ("AF", "Brevets, licences, logiciels et droits similaires",       "actif", None,
            ["212","213","214","2193"], ["2812","2813","2814","2912","2913","2914","2919"], [], [], [], []),
            ("AG", "Fonds commercial et droit au bail",                       "actif", None,
            ["215","216"], ["2815","2816","2915","2916"],                     [], [], [], []),
            ("AH", "Autres immobilisations incorporelles",                    "actif", None,
            ["217","218","2198"], ["2817","2818","2917","2918","2919"],       [], ["2181"], [], ["2181"]),
            # Ligne agrégée AD calculée par compute_etats_financiers
            ("AD", "IMMOBILISATIONS INCORPORELLES",                           "actif", 3,
            ["211","2191","212","213","214","2193","215","216","217","218","2198"],
            ["2811","2818","2911","2918","2919","2812","2813","2814","2912","2913","2914",
            "2815","2816","2915","2916","2817","2818","2917","2918"],
            [], [], [], []),

            # Immobilisations corporelles (lignes feuilles)
            ("AJ", "Terrains (1) dont Placement Net :",                       "actif", None,
            ["22"], ["282","292"],                                            [], [], [], []),
            ("AK", "BÂtiments(1)dont Placement Net :",                        "actif", None,
            ["231","232","233","237","2391"],
            ["2831","2832","2833","2837","2931","2932","2933","2937","2939"],
            [], [], [], []),
            ("AL", "Aménagements, agencements et installations",              "actif", None,
            ["234","235","238","2392","2393"],
            ["2834","2835","2838","2934","2935","2938","2939"],               [], [], [], []),
            ("AM", "Matériel, mobilier et actifs biologiques",                "actif", None,
            ["241", "242", "243", "244", "246", "247", "248"],
            ["2841", "2842", "2843", "2844", "2846", "2847", "2848",
            "2941", "2942", "2943", "2944", "2946", "2947", "2948"],
            [], [], [], []),
            ("AN", "Matériel de transport",                                   "actif", None,
            ["245","2495"], ["2845","2945","2949"],                           [], [], [], []),
            # Ligne agrégée AI
            ("AI", "IMMOBILISATIONS CORPORELLES",                             "actif", 3,
            ["22", "231", "232", "233", "237", "2391",
            "234", "235", "238", "2392", "2393",
            "241", "242", "243", "244", "246", "247", "248",
            "245", "2495"],
            ["282", "292",
            "2831", "2832", "2833", "2837", "2931", "2932", "2933", "2937", "2939",
            "2834", "2835", "2838", "2934", "2935", "2938",
            "2841", "2842", "2843", "2844", "2846", "2847", "2848",
            "2941", "2942", "2943", "2944", "2946", "2947", "2948",
            "2845", "2945", "2949"],
            [], [], [], []),

            # Avances sur immobilisations
            ("AP", "AVANCES ET ACOMPTES VERSES SUR IMMOBILISATIONS",          "actif", None,
            ["251","252"], ["2951","2952"],                                   [], [], [], []),

            # Immobilisations financières (lignes feuilles)
            ("AR", "Titres de participation",                                 "actif", None,
            ["26"], ["296"],                                                  [], [], [], []),
            ("AS", "Autres immobilisations financières",                      "actif", None,
            ["27"], ["297"],                                                  [], [], [], []),
            # Ligne agrégée AQ
            ("AQ", "IMMOBILISATIONS FINANCIERES",                             "actif", None,
            ["26","27"], ["296","297"],                                       [], [], [], []),

            # Total actif immobilisé (agrégé)
            ("AZ", "TOTAL ACTIF IMMOBILISE",                                  "actif", None,
            ["211","2181","2191","212","213","214","2193","215","216","217","218","2198",
            "22","231","232","233","237","2391","234","235","238","2392","2393","24",
            "245","2495","251","252","26","27"],
            ["2811","2818","2911","2918","2919","2812","2813","2814","2912","2913","2914",
            "2815","2816","2915","2916","2817","2817","2917","2918","282","292",
            "2831","2832","2833","2837","2931","2932","2933","2937","2939",
            "2834","2835","2838","2934","2935","2938","284","2845","2945","2949",
            "2951","2952","296","297"],
            [], [], [], []),

            # Actif circulant HAO
            ("BA", "ACTIF CIRCULANT HAO",                                     "actif", None,
            ["485","488"], ["498"],                                           [], [], [], []),

            # Stocks
            ("BB", "STOCKS ET ENCOURS",                                       "actif", 4,
            ["31","32","33","34","35","36","37","38"], ["39"],                [], [], [], []),

            # Créances (lignes feuilles)
            ("BH", "Fournisseurs avances versées",                            "actif", None,
            ["409"], ["490"],                                                 [], [], [], []),
            ("BI", "Clients",                                                  "actif", 5,
            ["41"], ["491"],                                                  [], ["419"], [], ["419"]),
            ("BJ", "Autres créances",                                         "actif", 6,
            ["185","42","43","44","45","46","47"],
            ["492","493","494","495","496","497"],                            [], ["478"], [], ["478"]),
            # Ligne agrégée BG
            ("BG", "CREANCES ET EMPLOIS ASSIMILES",                           "actif", None,
            ["485","488","31","32","33","34","35","36","37","38",
            "409","41","185","42","43","44","45","46","47"],
            ["498","39","490","491","492","493","494","495","496","497"],
            [], ["419","478"], [], ["419","478"]),

            # Total actif circulant (agrégé)
            ("BK", "TOTAL ACTIF CIRCULANT",                                   "actif", None,
            ["409","41","185","42","43","44","45","46","47"],
            ["490","491","492","493","494","495","496","497"],
            [], ["419","478"], [], ["419","478"]),

            # Trésorerie actif (lignes feuilles)
            ("BQ", "Titres de placement",                                     "actif", None,
            ["50"], ["590"],                                                  [], [], [], []),
            ("BR", "Valeurs à encaisser",                                     "actif", None,
            ["51"], ["591"],                                                  [], [], [], []),
            ("BS", "Banques, chèques postaux, caisse et assimilés",           "actif", None,
            ["52","53","54","55","57","581","582"], ["592","593","594"],       [], [], [], []),
            # Ligne agrégée BT
            ("BT", "TOTAL TRESORERIE-ACTIF",                                  "actif", None,
            ["50","51","52","53","54","55","57","581","582"],
            ["590","591","592","593","594"],                                  [], [], [], []),

            # Écart de conversion actif
            ("BU", "Ecart de conversion-Actif",                               "actif", None,
            ["478"], ["0"],                                                   [], [], [], []),

            # Total général actif
            ("BZ", "TOTAL GENERAL",                                           "actif", None,
            ["211","2181","2191","212","213","214","2193","215","216","217","218","2198",
            "22","231","232","233","237","2391","234","235","238","2392","2393","24",
            "245","2495","251","252","26","27","485","488","31","32","33","34","35","36",
            "37","38","409","41","185","42","43","44","45","46","47","50","51","52","53",
            "54","55","57","581","582","478"],
            ["2811","2818","2911","2918","2919","2812","2813","2814","2912","2913","2914",
            "2815","2816","2915","2916","2817","2917","2918","282","292",
            "2831","2832","2833","2837","2931","2932","2933","2937","2939",
            "2834","2835","2838","2934","2935","2938","284","2845","2945","2949",
            "2951","2952","296","297","498","39","490","491","492","493","494","495",
            "496","497","590","591","592","593","594"],
            [], ["2181","245","2495","419","478"], ["2845","2945","2949"],
            ["2181","245","2495","2845","2945","2949","419","478"]),

            # PASSIF
            # Capitaux propres (lignes feuilles)
            ("CA", "Capital",                                                 "passif", None,
            [], [], ["101","102","103","104"],                                [], [], []),
            ("CB", "Apporteurs capital non appelé (-)",                       "passif", None,
            [], [], ["109"],                                                  [], [], []),
            ("CD", "Primes liées au capital social",                          "passif", None,
            [], [], ["105"],                                                  [], [], []),
            ("CE", "Ecarts de réévaluation",                                  "passif", None,
            [], [], ["106"],                                                  [], [], []),
            ("CF", "Réserves indisponibles",                                  "passif", None,
            [], [], ["111","112","113"],                                      [], [], []),
            ("CG", "Réserves libres",                                         "passif", None,
            [], [], ["118"],                                                  [], [], []),
            ("CH", "Report à nouveau (+ ou -)",                               "passif", None,
            [], [], ["121","129"],                                            [], [], []),
            ("CJ", "Résultat net de l'exercice (bénéfice + ou perte -)",                               "passif", None,
            [], [], ["701","702","703","704","705","706","707","601","6031",
                    "602","6032","604","605","608","6033","61","62","63","64","65","73","72","71","75","781",
                    "66","791","798","799","681","691","77","797","787","67","697",
                    "82","84","86","88","81","83","85","87","89"],
            [], [], []),
            ("CL", "Subventions d'investissement",                            "passif", None,
            [], [], ["14"],                                                   [], [], []),
            ("CM", "Provisions réglementées",                                 "passif", None,
            [], [], ["15"],                                                   [], [], []),
            # Ligne agrégée CP
            ("CP", "TOTAL CAPITAUX PROPRES ET RESSOURCES ASSIMILEES",         "passif", None,
            [], [],
            ["101","102","103","104","109","105","106","111","112","113","118",
            "121","129","131","139","14","15","701","702","703","704","705","706","707","601","6031",
                    "602","6032","604","605","608","6033","61","62","63","64","65","73","72","71","75","781",
                    "66","791","798","799","681","691","77","797","787","67","697",
                    "82","84","86","88","81","83","85","87","89"],
            [], [], []),

            # Dettes financières (lignes feuilles)
            ("DA", "Emprunts et dettes financières diverses",                 "passif", None,
            [], [], ["16","181","182","183","184"],                           [], [], []),
            ("DB", "Dettes de location-acquisition",                         "passif", None,
            [], [], ["17"],                                                   [], [], []),
            ("DC", "Provisions pour risques et charges",                      "passif", None,
            [], [], ["19"],                                                   [], [], []),
            # Ligne agrégée DD
            ("DD", "TOTAL DETTES FINANCIERES ET RESSOURCES ASSIMILEES",       "passif", None,
            [], [], ["16","181","182","183","184","17","19"],                 [], [], []),
            # Ligne agrégée DF
            ("DF", "TOTAL RESSOURCES STABLES",                                "passif", None,
            [], [],
            ["101","102","103","104","109","105","106","111","112","113","118",
            "121","129","131","139","14","15","16","181","182","183","184","17","19"],
            [], [], []),

            # Passif circulant (lignes feuilles)
            ("DH", "Dettes circulantes HAO",                                  "passif", None,
            [], [], ["481","482","484","4998"],                               [], [], []),
            ("DI", "Clients, avances reçues",                                 "passif", None,
            [], [], ["419"],                                                  [], [], []),
            ("DJ", "Fournisseurs d'exploitation",                             "passif", None,
            [], [], ["40"],                                                   [], [], ["409"]),
            ("DK", "Dettes fiscales et sociales",                             "passif", None,
            [], [], ["42","43","44"],                                         [], [], []),
            ("DM", "Autres dettes",                                           "passif", None,
            [], [], ["185","45","46","47"],                                   [], [], ["479"]),
            ("DN", "Provisions pour risques et charges à court terme",        "passif", None,
            [], [], ["499","599"],                                            [], [], ["4998"]),
            # Ligne agrégée DP
            ("DP", "TOTAL PASSIF CIRCULANT",                                  "passif", None,
            [], [],
            ["481","482","484","4998","419","40","42","43","44","185","45","46","47","499","599"],
            [], [], ["409","479"]),

            # Trésorerie passif (lignes feuilles)
            ("DQ", "Banques, crédits d'escompte",                             "passif", None,
            [], [], ["564","565"],                                            [], [], []),
            ("DR", "Banques, établissements financiers et crédits de trésorerie", "passif", None,
            [], [], ["52","53","561","566"],                                  [], [], []),
            # Ligne agrégée DT
            ("DT", "TOTAL TRESORERIE-PASSIF",                                 "passif", None,
            [], [], ["564","565","52","53","561","566"],                      [], [], []),

            # Écart de conversion passif
            ("DV", "Ecart de conversion-Passif",                              "passif", None,
            [], [], ["479"],                                                  [], [], []),

            # Total général passif
            ("DZ", "TOTAL GENERAL",                                           "passif", None,
            [], [],
            ["101","102","103","104","109","105","106","111","112","113","118",
            "121","129","131","139","14","15","16","181","182","183","184","17","19",
            "481","482","484","4998","419","40","42","43","44","185","45","46","47",
            "499","599","564","565","52","53","561","566","479"],
            [], [], ["409"]),

            # COMPTE DE RÉSULTAT
            ("TA", "Ventes de marchandises",                                   "pnl", None,
            [], [], ["701"],                                                  [], [], []),
            ("RA", "Achats de marchandises",                                   "pnl", None,
            [], [], ["601"],                                                  [], [], []),
            ("RB", "Variation de stocks de marchandises",                      "pnl", None,
            [], [], ["6031"],                                                 [], [], []),
            ("XA", "MARGE COMMERCIALE (Somme TA à RB)",                        "pnl", None,
            [], [], ["701","601","6031"],                                     [], [], []),
            ("TB", "Ventes de produits fabriqués",                             "pnl", None,
            [], [], ["702","703","704"],                                      [], [], []),
            ("TC", "Travaux, services vendus",                                 "pnl", None,
            [], [], ["705","706"],                                            [], [], []),
            ("TD", "Produits accessoires",                                     "pnl", None,
            [], [], ["707"],                                                  [], [], []),
            ("XB", "CHIFFRE D'AFFAIRES (A + B + C + D)",                       "pnl", None,
            [], [], ["701","702","703","704","705","706","707"],              [], [], []),
            ("TE", "Production stockée (ou déstockage)",                       "pnl", None,
            [], [], ["73"],                                                   [], [], []),
            ("TF", "Production immobilisée",                                   "pnl", None,
            [], [], ["72"],                                                   [], [], []),
            ("TG", "Subventions d'exploitation",                               "pnl", None,
            [], [], ["71"],                                                   [], [], []),
            ("TH", "Autres produits",                                          "pnl", None,
            [], [], ["75"],                                                   [], [], []),
            ("TI", "Transferts de charges d'exploitation",                     "pnl", None,
            [], [], ["781"],                                                  [], [], []),
            ("RC", "Achats de matières premières et fournitures liées",        "pnl", None,
            [], [], ["602"],                                                  [], [], []),
            ("RD", "Variation de stocks de matières premières et fournitures liées", "pnl", None,
            [], [], ["6032"],                                                 [], [], []),
            ("RE", "Autres achats",                                            "pnl", None,
            [], [], ["604","605","608"],                                      [], [], []),
            ("RF", "Variation de stocks d'autres approvisionnements",          "pnl", None,
            [], [], ["6033"],                                                 [], [], []),
            ("RG", "Transports",                                               "pnl", None,
            [], [], ["61"],                                                   [], [], []),
            ("RH", "Services extérieurs",                                      "pnl", None,
            [], [], ["62","63"],                                              [], [], []),
            ("RI", "Impôts et taxes",                                          "pnl", None,
            [], [], ["64"],                                                   [], [], []),
            ("RJ", "Autres charges",                                           "pnl", None,
            [], [], ["65"],                                                   [], [], []),
            ("XC", "VALEUR AJOUTEE (XB +RA+RB) + (somme TE à RJ)",            "pnl", None,
            [], [], ["701","702","703","704","705","706","707","601","6031",
                    "602","6032","604","605","608","6033","61","62","63","64","65","73","72","71","75","781"],
            [], [], []),
            ("RK", "Charges de personnel",                                     "pnl", None,
            [], [], ["66"],                                                   [], [], []),
            ("XD", "EXCEDENT BRUT D'EXPLOITATION (XC+RK)",                     "pnl", None,
            [], [], ["701","702","703","704","705","706","707","601","6031",
                    "602","6032","604","605","608","6033","61","62","63","64","65","73","72","71","75","781","66"],
            [], [], []),
            ("TJ", "Reprises d'amortissements, provisions et dépréciations",   "pnl", None,
            [], [], ["791","798","799"],                                      [], [], []),
            ("RL", "Dotations aux amortissements, aux provisions et dépréciations", "pnl", None,
            [], [], ["681","691"],                                            [], [], []),
            ("XE", "RESULTAT D'EXPLOITATION (XD+TJ+ RL)",                      "pnl", None,
            [], [], ["701","702","703","704","705","706","707","601","6031",
                    "602","6032","604","605","608","6033","61","62","63","64","65","73","72","71","75","781",
                    "66","791","798","799","681","691"],
            [], [], []),
            ("TK", "Revenus financiers et assimilés",                          "pnl", None,
            [], [], ["77"],                                                   [], [], []),
            ("TL", "Reprises de provisions et dépréciations financières",      "pnl", None,
            [], [], ["797"],                                                  [], [], []),
            ("TM", "Transferts de charges financières",                        "pnl", None,
            [], [], ["787"],                                                  [], [], []),
            ("RM", "Frais financiers et charges assimilées",                   "pnl", None,
            [], [], ["67"],                                                   [], [], []),
            ("RN", "Dotations aux provisions et aux dépréciations financières","pnl", None,
            [], [], ["697"],                                                  [], [], []),
            ("XF", "RESULTAT FINANCIER (somme TK à RN)",                       "pnl", None,
            [], [], ["77","797","787","67","697"],                            [], [], []),
            ("XG", "RESULTAT DES ACTIVITES ORDINAIRES (XE+XF)",                "pnl", None,
            [], [], ["701","702","703","704","705","706","707","601","6031",
                    "602","6032","604","605","608","6033","61","62","63","64","65","73","72","71","75","781",
                    "66","791","798","799","681","691","77","797","787","67","697"],
            [], [], []),
            ("TN", "Produits des cessions d'immobilisations",                  "pnl", None,
            [], [], ["82"],                                                   [], [], []),
            ("TO", "Autres Produits HAO",                                      "pnl", None,
            [], [], ["84","86","88"],                                         [], [], []),
            ("RO", "Valeurs comptables des cessions d'immobilisations",        "pnl", None,
            [], [], ["81"],                                                   [], [], []),
            ("RP", "Autres Charges HAO",                                       "pnl", None,
            [], [], ["83","85"],                                              [], [], []),
            ("XH", "RESULTAT HORS ACTIVITES ORDINAIRES (somme TN à RP)",       "pnl", None,
            [], [], ["82","84","86","88","81","83","85"],                     [], [], []),
            ("RQ", "Participation des travailleurs",                           "pnl", None,
            [], [], ["87"],                                                   [], [], []),
            ("RS", "Impôts sur le résultat",                                   "pnl", None,
            [], [], ["89"],                                                   [], [], []),
            ("XI", "RESULTAT NET (XG+XH+RQ+RS)",                               "pnl", None,
            [], [], ["701","702","703","704","705","706","707","601","6031",
                    "602","6032","604","605","608","6033","61","62","63","64","65","73","72","71","75","781",
                    "66","791","798","799","681","691","77","797","787","67","697",
                    "82","84","86","88","81","83","85","87","89"],
            [], [], []),
        ]

        # ------------------------------------------------------------------
        # Helpers
        # ------------------------------------------------------------------
        def _sum_prefixes(balance, prefixes):
            if not prefixes:
                return 0
            return sum(
                item["solde_reel"]
                for item in balance
                if any(str(item["numero_compte"]).startswith(p) for p in prefixes)
            )

        def _collect_prefixes(*lists):
            """Retourne la liste dédupliquée de tous les préfixes utilisés."""
            seen = set()
            out = []
            for lst in lists:
                for p in lst:
                    if p and p not in seen:
                        seen.add(p)
                        out.append(p)
            return out

        # ------------------------------------------------------------------
        # Calcul
        # ------------------------------------------------------------------
        datum = {"actif": [], "passif": [], "pnl": []}

        for (ref, libelle, nature, note,
            brut_cpts, amor_cpts, net_cpts,
            brut_except, amor_except, net_except) in EFI_CATALOGUE:

            row = {
                "ref":            ref,
                "libelle":        libelle,
                "note":           note,
                "brut_solde_n":   None,
                "amor_solde_n":   None,
                "net_solde_n":    0,
                "net_solde_n1":   0,
                "compte_to_be_used":     "",
                "compte_to_be_used_off": _collect_prefixes(brut_cpts, amor_cpts, net_cpts, brut_except, amor_except, net_except),
            }

            if brut_cpts or amor_cpts:
                # Cas avec brut + amortissements
                brut_n   = _sum_prefixes(balance_n,  brut_cpts)
                amor_n   = _sum_prefixes(balance_n,  amor_cpts)
                bex_n    = _sum_prefixes(balance_n,  brut_except)
                aex_n    = _sum_prefixes(balance_n,  amor_except)

                brut_n1  = _sum_prefixes(balance_n1, brut_cpts)
                amor_n1  = _sum_prefixes(balance_n1, amor_cpts)
                nex_n1   = _sum_prefixes(balance_n1, net_except)

                row["brut_solde_n"] = brut_n  + bex_n
                row["amor_solde_n"] = amor_n  + aex_n
                row["net_solde_n"]  = row["brut_solde_n"] + row["amor_solde_n"]
                row["net_solde_n1"] = brut_n1 + amor_n1 + nex_n1

            else:
                # Cas net seul
                net_n    = _sum_prefixes(balance_n,  net_cpts)
                nex_n    = _sum_prefixes(balance_n,  net_except)
                net_n1   = _sum_prefixes(balance_n1, net_cpts)
                nex_n1   = _sum_prefixes(balance_n1, net_except)

                row["net_solde_n"]  = net_n  + nex_n
                row["net_solde_n1"] = net_n1 + nex_n1

            # Champ legacy (chaîne lisible des préfixes)
            all_cpts = _collect_prefixes(brut_cpts, amor_cpts, net_cpts,
                                        brut_except, amor_except, net_except)
            row["compte_to_be_used"] = ",".join(all_cpts)

            datum[nature].append(row)

        # ------------------------------------------------------------------
        # Recalcul des lignes agrégées (XA, XB, CP, DD, AD, AI)
        # ------------------------------------------------------------------
        datum = self.compute_etats_financiers(datum)

        return datum


    def compute_etats_financiers(self, efi):
        """
        Recalcule les lignes agrégées SYSCOHADA en appliquant les formules
        officielles sur les valeurs déjà calculées par prod_efi().

        Formules :
        ACTIF  :  AD=AE+AF+AG+AH  |  AI=AJ+AK+AL+AM+AN  |  AQ=AR+AS
                    AZ=AD+AI+AQ      |  BG=BH+BI+BJ  |  BK=BA+BB+BG
                    BT=BQ+BR+BS     |  BZ=AZ+BK+BT+BU
        PASSIF :  CP=CA+CB+CD+CE+CF+CG+CH+CJ+CL+CM
                    DD=DA+DB+DC     |  DF=CP+DD
                    DP=DH+DI+DJ+DK+DM+DN  |  DT=DQ+DR
                    DZ=DF+DP+DT+DV
        PNL    :  XA = -TA - RA +/- RB
                    XB = -TA - TB - TC - TD
                    XC = (XB - RA +/- RB) + (+/- TE - TF - TG - TH - TI - RC +/- RD - RE +/- RF - RG - RH - RI - RJ)
                    XD = XC - RK
                    XE = XD - TJ - RL
                    XF = -TK - TL - TM - RM - RN
                    XG = XE + XF
                    XH = -TN - TO - RO - RP
                    XI = XG + XH - RQ - RS

        """

        def _g(idx, ref, key="net_solde_n"):
            row = idx.get(ref)
            if not row:
                return 0
            v = row.get(key, 0)
            return v if v is not None else 0

        def _set(idx, ref, val_n, val_n1):
            if ref in idx:
                idx[ref]["net_solde_n"]  = val_n
                idx[ref]["net_solde_n1"] = val_n1

        # def _sum(idx, refs, key="net_solde_n"):
        #     return sum(_g(idx, r, key) for r in refs)

        for section in ("actif", "passif", "pnl"):
            rows = efi.get(section, [])
            idx  = {r["ref"]: r for r in rows if r.get("ref")}
            K    = "net_solde_n1"  # alias court

            if section == "actif":
                # Ligne feuilles déjà calculées par prod_ef on recalcule
                # uniquement les agrégats pour qu'ils soient cohérents.
                # for agg_ref, leaf_refs in [
                #     ("AD", ["AE","AF","AG","AH"]),
                #     ("AI", ["AJ","AK","AL","AM","AN"]),
                #     ("AQ", ["AR","AS"]),
                #     ("BG", ["BH","BI","BJ"]),
                #     ("AZ", ["AD","AI","AP","AQ"]),
                #     ("BK", ["BA","BB","BG"]),
                #     ("BT", ["BQ","BR","BS"]),
                #     ("BZ", ["AZ","BK","BT","BU"]),
                # ]:
                #     _set(idx, agg_ref,
                #         _sum(idx, leaf_refs),
                #         _sum(idx, leaf_refs, K))

                g  = lambda r:   _g(idx, r)
                g1 = lambda r:   _g(idx, r, K)

                # ("AD", ["AE","AF","AG","AH"])
                ad  = g("AE") + g("AF") + g("AG") + g("AH")
                ad1 = g1("AE") + g1("AF") + g1("AG") + g1("AH")
                _set(idx, "AD", ad, ad1)

                # ("AI", ["AJ","AK","AL","AM","AN"])
                ai  = g("AJ") + g("AK") + g("AL") + g("AM") + g("AN")
                ai1 = g1("AJ") + g1("AK") + g1("AL") + g1("AM") + g1("AN")
                _set(idx, "AI", ai, ai1)

                # ("AQ", ["AR","AS"])
                aq  = g("AR") + g("AS")
                aq1 = g1("AR") + g1("AS")
                _set(idx, "AQ", aq, aq1)

                # ("BG", ["BH","BI","BJ"])
                bg  = g("BH") + g("BI") + g("BJ")
                bg1 = g1("BH") + g1("BI") + g1("BJ")
                _set(idx, "BG", bg, bg1)

                # ("AZ", ["AD","AI","AQ"])
                az  = g("AD") + g("AI") + g("AQ")
                az1 = g1("AD") + g1("AI") + g1("AQ")
                _set(idx, "AZ", az, az1)

                # ("BK", ["BA","BB","BG"])
                bk  = g("BA") + g("BB") + g("BG")
                bk1 = g1("BA") + g1("BB") + g1("BG")
                _set(idx, "BK", bk, bk1)

                # ("BT", ["BQ","BR","BS"])
                bt  = g("BQ") + g("BR") + g("BS")
                bt1 = g1("BQ") + g1("BR") + g1("BS")
                _set(idx, "BT", bt, bt1)

                # ("BZ", ["AZ","BK","BT","BU"])
                bz  = g("AZ") + g("BK") + g("BT") + g("BU")
                bz1 = g1("AZ") + g1("BK") + g1("BT") + g1("BU")
                _set(idx, "BZ", bz, bz1)

            elif section == "passif":
                # for agg_ref, leaf_refs in [
                #     ("CP", ["CA","CB","CD","CE","CF","CG","CH","CJ","CL","CM"]),
                #     ("DD", ["DA","DB","DC"]),
                #     ("DF", ["CP","DD"]),
                #     ("DP", ["DH","DI","DJ","DK","DM","DN"]),
                #     ("DT", ["DQ","DR"]),
                #     ("DZ", ["DF","DP","DT","DV"]),
                # ]:
                #     _set(idx, agg_ref,
                #         _sum(idx, leaf_refs),
                #         _sum(idx, leaf_refs, K))

                g  = lambda r:   _g(idx, r)
                g1 = lambda r:   _g(idx, r, K)

                # ("CP", ["CA","CB","CD","CE","CF","CG","CH","CJ","CL","CM"])
                cp  = g("CA") + g("CB") + g("CD") + g("CE") + g("CF") + g("CG") + g("CH") + g("CJ") + g("CL") + g("CM")
                cp1 = g1("CA") + g1("CB") + g1("CD") + g1("CE") + g1("CF") + g1("CG1") + g1("CH") + g1("CJ") + g1("CL") + g1("CM")
                _set(idx, "CP", cp, cp1)

                # ("DD", ["DA","DB","DC"])
                dd  = g("DA") + g("DB") + g("DC")
                dd1 = g1("DA") + g1("DB") + g1("DC")
                _set(idx, "DD", dd, dd1)

                # ("DF", ["CP","DD"])
                df  = g("CP") + g("DD")
                df1 = g1("CP") + g1("DD")
                _set(idx, "DF", df, df1)

                # ("DP", ["DH","DI","DJ","DK","DM","DN"])
                dp  = g("DH") + g("DI") + g("BJ") + g("DK") + g("DM") + g("BN")
                dp1 = g1("DH") + g1("DI") + g1("BJ") + g1("DK") + g1("DM") + g1("BN")
                _set(idx, "DP", dp, dp1)

                # ("DT", ["DQ","DR"])
                dt  = g("DQ") + g("DR")
                dt1 = g1("DQ") + g1("DR")
                _set(idx, "DT", dt, dt1)

                # ("DZ", ["DF","DP","DT","DV"])
                dz  = g("DF") + g("DP") + g("DT") + g("DV")
                dz1 = g1("DF") + g1("DP") + g1("DT") + g1("DV")
                _set(idx, "DZ", dz, dz1)

            elif section == "pnl":
                g  = lambda r:   _g(idx, r)
                g1 = lambda r:   _g(idx, r, K)

                # XA = TA - RA + RB
                xa  = - g("TA") - g("RA") + g("RB")
                xa1 = - g1("TA") - g1("RA") + g1("RB")
                _set(idx, "XA", xa, xa1)

                # XB = TA + TB + TC + TD
                xb  = - g("TA") - g("TB") - g("TC") - g("TD")
                xb1 = - g1("TA") - g1("TB") - g1("TC") - g1("TD")
                _set(idx, "XB", xb, xb1)

                # XC
                xc = (xb
                    - g("RA") + g("RB")
                    + g("TE") - g("TF") - g("TG") - g("TH") - g("TI")
                    - g("RC") + g("RD")
                    - g("RE") + g("RF")
                    - g("RG") - g("RH") - g("RI") - g("RJ"))
                xc1 = (xb1
                    - g1("RA") + g1("RB")
                    + g1("TE") - g1("TF") - g1("TG") - g1("TH") - g1("TI")
                    - g1("RC") + g1("RD")
                    - g1("RE") + g1("RF")
                    - g1("RG") - g1("RH") - g1("RI") - g1("RJ"))
                _set(idx, "XC", xc, xc1)

                # XD = XC - RK
                xd  = xc  - g("RK");   xd1 = xc1 - g1("RK")
                _set(idx, "XD", xd, xd1)

                # XE = XD + TJ - RL
                xe  = xd  - g("TJ")  - g("RL");   xe1 = xd1 - g1("TJ") - g1("RL")
                _set(idx, "XE", xe, xe1)

                # XF = TK + TL + TM - RM - RN
                xf  = - g("TK")  - g("TL")  - g("TM")  - g("RM")  - g("RN")
                xf1 = - g1("TK") - g1("TL") - g1("TM") - g1("RM") - g1("RN")
                _set(idx, "XF", xf, xf1)

                # XG = XE + XF
                xg  = xe  + xf;   xg1 = xe1 + xf1
                _set(idx, "XG", xg, xg1)

                # XH = TN + TO - RO - RP
                xh  = - g("TN")  - g("TO")  - g("RO")  - g("RP")
                xh1 = - g1("TN") - g1("TO") - g1("RO") - g1("RP")
                _set(idx, "XH", xh, xh1)

                # XI = XG + XH - RQ - RS
                xi  = xg  + xh  - g("RQ")  - g("RS")
                xi1 = xg1 + xh1 - g1("RQ") - g1("RS")
                _set(idx, "XI", xi, xi1)

        return efi

    # ---------- Piste d'audit ----------
    def audit_trail(self, id_mission):
        db = get_db()  # Obtenir la connexion à la base de données
        # Créer un fichier Excel pour la piste d'audit
        wb = openpyxl.Workbook()
        sheet = wb.active

        columns = ['A', 'B', 'C', 'D', 'E']
        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Code COMMENTAIRE']

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        balances = mission['balance_variation']
        actif = mission['efi']['actif']
        passif = mission['efi']['passif']
        pnl = mission['efi']['pnl']

        efi = actif + passif + pnl

        for i in range(len(columns)):
            sheet[columns[i] + '1'] = headers[i]

        for iteration, data in enumerate(balances):
            new_iteration = str(iteration + 2)
            sheet["A" + new_iteration] = data.get("numero_compte")
            sheet["B" + new_iteration] = data.get("solde_n")
            sheet["C" + new_iteration] = data.get("solde_n1")
            sheet["D" + new_iteration] = data.get("numero_compte")[0:2]

            list_code_efi = []
            for obj in efi:
                for elt in obj['compte_to_be_used_off']:
                    if data['numero_compte'].startswith(elt):
                        list_code_efi.append(obj['ref'])
            list_code_efi = list(set(list_code_efi))
            sheet["E" + new_iteration] = ','.join(list_code_efi)

        namefile = "piste_audit.xlsx"
        wb.save(namefile)

    # ---------- Extract grouping Excel ----------
    def extract_grouping(self, id_mission):
        db = get_db()  # Obtenir la connexion à la base de données
        # Créer un fichier Excel pour l'export du grouping
        wb = openpyxl.Workbook()
        sheet = wb.active

        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Variation', 'Variation %', 'Compte qualitatif', 'Compte quantitatif', 'Compte significatif']

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        balances = mission['balance_variation']
        grouping = mission['grouping']
        materiality = next(item for item in mission['materiality'] if item['choice'] is True)

        for i in range(len(columns)):
            sheet[columns[i] + '1'] = headers[i]

        for iteration, data in enumerate(balances):
            new_iteration = str(iteration + 2)
            sheet["A" + new_iteration] = data.get("numero_compte")
            sheet["B" + new_iteration] = data.get("solde_n")
            sheet["C" + new_iteration] = data.get("solde_n1")

            value_grouping = data.get("numero_compte")[0:2]
            variation = data.get("solde_n") - data.get("solde_n1")

            if variation == 0:
                variation_percent = 0
            elif data.get("solde_n1") == 0:
                variation_percent = 100
            else:
                variation_percent = (variation / data.get("solde_n1")) * 100

            sheet["D" + new_iteration] = value_grouping
            sheet["E" + new_iteration] = variation
            sheet["F" + new_iteration] = variation_percent
            sheet["G" + new_iteration] = next(item['significant'] for item in grouping if item['compte'] == value_grouping)
            sheet["H" + new_iteration] = next(item['materiality'] for item in grouping if item['compte'] == value_grouping)
            sheet["I" + new_iteration] = next(item['mat_sign'] for item in grouping if item['compte'] == value_grouping)

        second_sheet = wb.create_sheet(title="Seuil de matérialité")
        second_headers = ['materiality', 'performance materiality', 'trivial misstatements']
        second_sheet["A1"] = second_headers[0]
        second_sheet["B1"] = second_headers[1]
        second_sheet["C1"] = second_headers[2]

        second_sheet["A2"] = materiality['materiality']
        second_sheet["B2"] = materiality['performance_materiality']
        second_sheet["C2"] = materiality['trivial_misstatements']

        excel_io = BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        return excel_io

    # ==============================
    #  CONTROLES Cohérence & Intangibilité
    # ==============================
    def _load_balance(self, balance_id):
        local_db = get_db()
        if local_db is None:
            raise RuntimeError("_load_balance: base de données non connectée")

        bal = local_db.Balance.find_one({"_id": ObjectId(balance_id)})
        return bal.get("balance", []) if bal else []


    def _get_sens_attendu_compte(self, numero_compte):
        """
        Détermine le sens attendu (Débit/Crédit) d'un compte selon le plan comptable SYSCOHADA
        Retourne: ('D', 'C', ou 'BOTH' pour les deux sens possibles)
        """
        if not numero_compte or len(numero_compte) < 2:
            return 'BOTH'
        
        prefixe = numero_compte[:2]
        classe = numero_compte[0] if numero_compte[0].isdigit() else None
        
        # Classe 1 - Capitaux : CRÉDITEUR (sens normal)
        if prefixe.startswith('10') or prefixe.startswith('11') or prefixe.startswith('12') or \
           prefixe.startswith('13') or prefixe.startswith('14') or prefixe.startswith('15') or \
           prefixe.startswith('16') or prefixe.startswith('17') or prefixe.startswith('18') or \
           prefixe.startswith('19'):
            # Exceptions : 
            # - 105 : peut être débiteur ou créditeur
            # - 12 (Report à nouveau) : peut être débiteur ou créditeur selon le résultat
            # - 109 (Associés non appelé) : peut être débiteur
            # - 129 (Report à nouveau débiteur) : peut être débiteur
            # - 1309 (Primes de remboursement des obligations) : peut être débiteur
            if numero_compte.startswith('105') or numero_compte.startswith('12') or numero_compte.startswith('109') or numero_compte.startswith('129') or numero_compte.startswith('1309'):
                return 'BOTH'  # Peut être débiteur ou créditeur selon le cas
            return 'C'  # Créditeur normal pour la classe 1
        
        # Classe 2 - Immobilisations : DÉBITEUR (sens normal)
        # Exceptions : 
        # - 28 (Amortissements) : CRÉDITEUR
        # - 29 (Dépréciations) : CRÉDITEUR
        if prefixe.startswith('2'):
            if prefixe.startswith('28') or prefixe.startswith('29'):
                return 'C'  # Amortissements et dépréciations sont créditeurs
            return 'D'  # Immobilisations brutes sont débitrices (sens normal)
        
        # Classe 3 - Stocks : DÉBITEUR (sens normal)
        # Exception : 
        # - 39 (Dépréciations des stocks) : CRÉDITEUR (le stock peut se déprécier)
        if prefixe.startswith('3'):
            if prefixe.startswith('39'):
                return 'C'  # Dépréciations des stocks sont créditrices
            return 'D'  # Stocks sont débitrices (sens normal)
        
        # Classe 4 - Tiers : Variable selon sous-classe
        # Règles détaillées :
        # - 40 (Fournisseurs) : CRÉDITEUR normal (dettes). Sens débiteur acceptable pour 4091 (avances), avoirs, trop-versé
        # - 41 (Clients) : DÉBITEUR normal (créances). Sens créditeur acceptable pour 4191 (avances), avoirs, trop-perçu
        # - 42 (Personnel) : généralement CRÉDITEUR (salaires à payer)
        # - 43 (Organismes sociaux) : CRÉDITEUR (charges sociales à payer)
        # - 44 (État) : TVA collectée créditeur, TVA déductible débiteur, impôts généralement créditeur, crédit TVA débiteur
        # - 45 (Groupe et associés) : variable selon la situation
        # - 46, 47 (Débiteurs et créditeurs divers) : selon la nature
        # - 48 (Comptes de régularisation) : variable selon le type
        # - 49 (Dépréciations tiers) : CRÉDITEUR
        if prefixe.startswith('4'):
            if prefixe.startswith('40'):  # Fournisseurs : CRÉDITEUR
                if numero_compte.startswith('4091'):  # Avances fournisseurs : exception
                    return 'BOTH'  # Peut être débiteur (avances, avoirs, trop-versé)
                return 'C'  # CRÉDITEUR (sens normal)
            elif prefixe.startswith('41'):  # Clients : DÉBITEUR
                if numero_compte.startswith('4191'):  # Avances clients : exception
                    return 'BOTH'  # Peut être créditeur (avances, avoirs, trop-perçu)
                return 'D'  # DÉBITEUR (sens normal)
            elif prefixe.startswith('42'):  # Personnel : généralement CRÉDITEUR
                return 'C'  # CRÉDITEUR (peut être soldé)
            elif prefixe.startswith('43'):  # Organismes sociaux : CRÉDITEUR
                return 'C'  # CRÉDITEUR
            elif prefixe.startswith('44'):  # État : variable selon le type
                # TVA collectée créditeur, TVA déductible débiteur, impôts généralement créditeur, crédit TVA débiteur
                return 'BOTH'  # Variable selon le sous-compte
            elif prefixe.startswith('45'):  # Groupe et associés : variable
                return 'BOTH'  # Variable selon la situation
            elif prefixe.startswith('46') or prefixe.startswith('47'):  # Débiteurs et créditeurs divers
                return 'BOTH'  # Selon la nature
            elif prefixe.startswith('48'):  # Comptes de régularisation
                return 'BOTH'  # Variable selon le type
            elif prefixe.startswith('49'):  # Dépréciations tiers : CRÉDITEUR
                return 'C'
            return 'BOTH'
        
        # Classe 5 - Financiers : DÉBITEUR (sens normal)
        # Exception : 
        # - 519 (Découverts bancaires) : peut être CRÉDITEUR
        if prefixe.startswith('5'):
            if prefixe.startswith('519'):  # Découverts bancaires : exception
                return 'BOTH'  # Peut être créditeur
            elif prefixe.startswith('58'):  # Virements internes
                return 'BOTH'  # Doit être soldé
            elif prefixe.startswith('59'):  # Dépréciations financiers
                return 'C'  # CRÉDITEUR
            return 'D'  # DÉBITEUR (sens normal)
        
        # Classe 6 - Charges : DÉBITEUR (sens normal)
        if prefixe.startswith('6'):
            return 'D'  # DÉBITEUR
        
        # Classe 7 - Produits : CRÉDITEUR (sens normal)
        if prefixe.startswith('7'):
            return 'C'  # CRÉDITEUR
        
        # Classe 8 - Résultats : selon le résultat
        if prefixe.startswith('8'):
            return 'BOTH'
        
        return 'BOTH'  # Par défaut, les deux sens sont possibles
    
    def _comptes_doivent_etre_soldes(self, numero_compte):
        """
        Vérifie si un compte doit être soldé en fin d'exercice selon les règles SYSCOHADA
        Retourne: (True/False, 'CRITIQUE'/'MOYENNE'/'FAIBLE', message)
        """
        if not numero_compte:
            return False, None, None
        
        # Comptes CRITIQUES qui doivent absolument être soldés
        comptes_critiques = {
            '471': ('CRITIQUE', 'Compte transitoire ou d\'attente (471) doit être soldé en fin d\'exercice. Tout solde indique des écritures en suspens non régularisées. Analyser ligne par ligne et régulariser chaque écriture.'),
            '58': ('CRITIQUE', 'Virements internes (58) doivent être soldés immédiatement. Utilisé pour les virements entre banques/caisses. Tout solde indique une erreur de lettrage ou un virement non comptabilisé des deux côtés.'),
        }
        
        # Comptes MOYENS à surveiller (fin d'exercice)
        comptes_moyens_fin_exercice = {
            '109': ('MOYENNE', 'Associés - capital souscrit non appelé (109) doit être soldé si tout le capital a été appelé. Si non soldé, vérifier la cohérence avec le capital social.'),
            '475': ('MOYENNE', 'Comptes de régularisation divers (475) doivent être soldés si toutes les régularisations ont été effectuées. Analyser les soldes résiduels.'),
        }
        
        # Comptes MOYENS à surveiller (en cours d'exercice)
        comptes_moyens_cours = {
            '422': ('MOYENNE', 'Personnel - Rémunérations dues (422) doivent être soldées lors de la paie suivante. Des soldes anciens peuvent indiquer des erreurs de lettrage ou des paiements non régularisés.'),
            '467': ('MOYENNE', 'Autres comptes débiteurs/créditeurs (467) - compte de passage temporaire. Ne doit pas avoir de solde ancien. Vérifier l\'ancienneté des montants non soldés.'),
            '468': ('MOYENNE', 'Charges à payer et produits à recevoir (468) doivent être soldés lors du paiement/encaissement effectif. Vérifier que les charges/produits provisionnés sont bien régularisés.'),
            '4387': ('MOYENNE', 'Organismes sociaux - Charges à payer (4387) doivent être soldées lors de la déclaration sociale suivante. Vérifier que les provisions sont bien régularisées.'),
            '4386': ('MOYENNE', 'Organismes sociaux - Autres charges à payer (4386) doivent être soldées lors du paiement effectif.'),
            '431': ('MOYENNE', 'Sécurité sociale (431) doit normalement être soldé chaque mois/trimestre lors du paiement. Un solde ancien indique un retard ou une erreur.'),
            '437': ('MOYENNE', 'Autres organismes sociaux (437) - même logique que 431. Normalement soldé chaque mois/trimestre lors du paiement.'),
        }
        
        # Comptes de TVA à solder à chaque déclaration
        comptes_tva = {
            '44551': ('CRITIQUE', 'TVA à décaisser (44551) doit être soldée lors du paiement de la TVA. Un solde ancien indique une TVA non payé risque fiscal.'),
            '44567': ('MOYENNE', 'Crédit de TVA à reporter (44567) doit être soldé lorsque le crédit est imputé ou remboursé. Vérifier la cohérence avec les déclarations.'),
            '44558': ('MOYENNE', 'TVA à régulariser ou en attente (44558) - compte temporaire à solder rapidement. Tout solde doit être justifié précisément.'),
        }
        
        # Vérifier les comptes critiques
        for compte_key, (gravite, message) in comptes_critiques.items():
            if numero_compte.startswith(compte_key):
                return True, gravite, message
        
        # Vérifier les comptes TVA
        for compte_key, (gravite, message) in comptes_tva.items():
            if numero_compte.startswith(compte_key):
                return True, gravite, message
        
        # Vérifier les comptes moyens (fin d'exercice)
        for compte_key, (gravite, message) in comptes_moyens_fin_exercice.items():
            if numero_compte.startswith(compte_key):
                return True, gravite, message
        
        # Vérifier les comptes moyens (en cours)
        for compte_key, (gravite, message) in comptes_moyens_cours.items():
            if numero_compte.startswith(compte_key):
                return True, gravite, message
        
        # Comptes de gestion (classe 6 et 7) : doivent être soldés en fin d'exercice
        if numero_compte[0] in ['6', '7']:
            return True, 'MOYENNE', f'Compte de gestion (classe {numero_compte[0]}) doit être soldé en fin d\'exercice par le jeu des écritures de clôture (virement au compte de résultat 12). Aucun compte de classe {numero_compte[0]} ne doit avoir de solde à l\'ouverture de l\'exercice N+1. Si solde présent en début d\'exercice : erreur de clôture, oubli du virement au résultat, ou reports à nouveau incorrects.'
        
        return False, None, None

    def _generer_structure_vraisemblance(self, erreurs_detectees=None):
        """
        Génère la structure complète du contrôle de vraisemblance avec :
        - Résumé explicatif (objectif, principes généraux)
        - Tableau par classe (1 à 7) avec nature, sens normal, exceptions, anomalies détectées
        - Tableau des comptes à solder obligatoirement
        - Liste des contrôles essentiels à effectuer
        
        Args:
            erreurs_detectees: Liste des erreurs détectées pour cette balance (optionnel)
        """
        # Regrouper les erreurs de sens par classe pour affichage dans le tableau
        anomalies_par_classe = {}
        if erreurs_detectees:
            for erreur in erreurs_detectees:
                if erreur.get("type") == "signe" and erreur.get("numero_compte"):
                    numero_compte = erreur.get("numero_compte")
                    if numero_compte and numero_compte[0].isdigit():
                        classe = numero_compte[0]
                        if classe not in anomalies_par_classe:
                            anomalies_par_classe[classe] = []
                        # Extraire le message complet et le formater de manière détaillée
                        message_original = erreur.get("message", "")
                        
                        # Construire un message détaillé et structuré
                        message_detaille = message_original
                        
                        # Extraire les informations clés du message original
                        try:
                            # Chercher le sens attendu et le sens actuel
                            if "devrait être" in message_original and "mais le solde est" in message_original:
                                # Extraire le sens attendu
                                if "devrait être" in message_original:
                                    partie_attendu = message_original.split("devrait être")[1]
                                    sens_attendu = partie_attendu.split(",")[0].strip() if "," in partie_attendu else partie_attendu.split(".")[0].strip()
                                
                                # Extraire le sens actuel
                                if "mais le solde est" in message_original:
                                    partie_actuel = message_original.split("mais le solde est")[1]
                                    sens_actuel = partie_actuel.split("(")[0].strip()
                                
                                # Extraire le montant
                                montant = ""
                                if "(" in message_original and "FCFA" in message_original:
                                    try:
                                        montant_part = message_original.split("(")[1].split(")")[0]
                                        montant = montant_part.strip()
                                    except:
                                        pass
                                
                                # Construire le message détaillé
                                message_detaille = f"ANOMALIE DÉTECTÉE : Le compte {numero_compte} présente un solde {sens_actuel.lower()}, alors que selon les règles SYSCOHADA pour la classe {classe}, il devrait être {sens_attendu.lower()}."
                                if montant:
                                    message_detaille += f" Le solde anormal s'élève à {montant}."
                                
                                # Ajouter des détails selon la classe
                                if classe == "3":
                                    message_detaille += " Un stock créditeur est ANORMAL et nécessite une investigation immédiate. Vérifiez les écritures d'inventaire, les erreurs de saisie ou les problèmes de valorisation."
                                elif classe == "5":
                                    if "53" in numero_compte:
                                        message_detaille += " Une caisse créditrice est IMPOSSIBLE physiquement. Il s'agit d'une erreur certaine qui doit être corrigée immédiatement."
                                    else:
                                        message_detaille += " Vérifiez la cohérence avec les relevés bancaires et les écritures de trésorerie."
                                elif classe == "6":
                                    message_detaille += " Un compte de charges créditeur indique généralement une erreur de comptabilisation ou un avoir mal enregistré. Vérifiez les écritures comptables et les avoirs clients."
                                elif classe == "7":
                                    message_detaille += " Un compte de produits débiteur indique généralement une erreur (annulation mal comptabilisée, avoir sur vente). Vérifiez les écritures d'annulation et les avoirs accordés."
                                elif classe == "2":
                                    message_detaille += " Un compte d'immobilisation créditeur est généralement anormal et peut indiquer une erreur ou une cession non comptabilisée correctement. Vérifiez les écritures de cession d'immobilisations."
                                else:
                                    message_detaille += " Vérifiez la nature du compte et les écritures comptables pour identifier l'origine de cette anomalie."
                        except Exception as e:
                            # En cas d'erreur d'extraction, utiliser le message original
                            message_detaille = message_original
                        
                        anomalies_par_classe[classe].append({
                            "compte": numero_compte,
                            "message": message_detaille
                        })
        structure = {
            "resume": {
                "objectif": "Le contrôle de vraisemblance vérifie la cohérence des soldes comptables selon les règles du plan comptable SYSCOHADA. Il permet de détecter les anomalies de sens, les comptes non soldés et les erreurs de comptabilisation.",
                "principes_generaux": [
                    "Chaque classe de compte a un sens normal de solde (débit ou crédit) selon sa nature",
                    "Certains comptes doivent obligatoirement être soldés en fin d'exercice",
                    "Des exceptions existent pour certains comptes spécifiques",
                    "Les anomalies détectées doivent être justifiées ou corrigées avant clôture"
                ]
            },
            "tableau_classes": [
                {
                    "classe": "1",
                    "nom": "Comptes de capitaux",
                    "nature": "Capital social, réserves, report à nouveau créditeur, subventions d'investissement, provisions réglementées, emprunts et dettes assimilées",
                    "sens_normal": "CRÉDITEUR",
                    "exceptions": [
                        "Report à nouveau débiteur (pertes antérieures) : débiteur accepté",
                        "Associés - capital souscrit non appelé (109) : débiteur (vient en diminution du capital)",
                        "Primes de remboursement des obligations : débiteur (charges à répartir)",
                        "Écarts de réévaluation (105) : peut être débiteur ou créditeur"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("1", [])
                },
                {
                    "classe": "2",
                    "nom": "Comptes d'immobilisations",
                    "nature": "Tous les comptes d'immobilisations brutes (20, 21, 23, etc.), immobilisations en cours, avances et acomptes sur immobilisations",
                    "sens_normal": "DÉBITEUR",
                    "exceptions": [
                        "Comptes d'amortissements (28x) : CRÉDITEUR",
                        "Comptes de dépréciations (29x) : CRÉDITEUR",
                        "Un compte d'immobilisation créditeur est généralement anormal (peut indiquer une erreur ou une cession non comptabilisée correctement)"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("2", [])
                },
                {
                    "classe": "3",
                    "nom": "Comptes de stocks",
                    "nature": "Tous les stocks (matières, marchandises, produits), en-cours de production",
                    "sens_normal": "DÉBITEUR",
                    "exceptions": [
                        "Comptes de dépréciation des stocks (39x) : CRÉDITEUR",
                        "Un stock créditeur est anormal et doit être investigué"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("3", [])
                },
                {
                    "classe": "4",
                    "nom": "Comptes de tiers",
                    "nature": "Comptes clients (41x), fournisseurs (40x), personnel (42x), sécurité sociale (43x), État et collectivités (44x), groupe et associés (45x), débiteurs et créditeurs divers (46x, 47x)",
                    "sens_normal": "Variable selon sous-classe",
                    "exceptions": [
                        "Comptes clients (41x) : DÉBITEUR normal (créances). Sens créditeur acceptable pour avances et acomptes reçus (4191), avoirs à établir, trop-perçu",
                        "Comptes fournisseurs (40x) : CRÉDITEUR normal (dettes). Sens débiteur acceptable pour avances et acomptes versés (4091), avoirs à obtenir, trop-versé",
                        "Personnel (42x) : généralement CRÉDITEUR (salaires à payer)",
                        "Sécurité sociale (43x) : CRÉDITEUR (charges sociales à payer)",
                        "État (44x) : TVA collectée créditeur, TVA déductible débiteur, impôts sur bénéfices généralement créditeur, crédit de TVA débiteur",
                        "Groupe et associés (45x) : variable selon la situation",
                        "Débiteurs et créditeurs divers (46x, 47x) : selon la nature",
                        "Dépréciations (49x) : CRÉDITEUR"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("4", [])
                },
                {
                    "classe": "5",
                    "nom": "Comptes financiers",
                    "nature": "Comptes de trésorerie active : Banques (512), Caisse (53x), Valeurs mobilières de placement (50x). Autres comptes financiers : Charges constatées d'avance (486), Produits constatés d'avance (487)",
                    "sens_normal": "DÉBITEUR",
                    "exceptions": [
                        "Concours bancaires courants (519) : CRÉDITEUR (découverts). Une banque créditrice indique un découvert",
                        "Caisse (53x) : TOUJOURS DÉBITEUR (un solde créditeur de caisse est impossible physiquement !)",
                        "Produits constatés d'avance (487) : CRÉDITEUR"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("5", [])
                },
                {
                    "classe": "6",
                    "nom": "Comptes de charges",
                    "nature": "Tous les comptes de charges (60 à 69) doivent être débiteurs en cours d'exercice",
                    "sens_normal": "DÉBITEUR",
                    "exceptions": [
                        "En fin d'exercice, après inventaire, certains comptes peuvent être soldés",
                        "Un compte de charges créditeur indique généralement une erreur de comptabilisation ou un avoir mal enregistré"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("6", [])
                },
                {
                    "classe": "7",
                    "nom": "Comptes de produits",
                    "nature": "Tous les comptes de produits (70 à 79) doivent être créditeurs en cours d'exercice",
                    "sens_normal": "CRÉDITEUR",
                    "exceptions": [
                        "En fin d'exercice, après inventaire, certains comptes peuvent être soldés",
                        "Un compte de produits débiteur indique généralement une erreur (annulation mal comptabilisée, avoir sur vente)"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("7", [])
                }
            ],
            "comptes_a_solder": self._generer_tableau_comptes_a_solder(),
            "controles_essentiels": [
                "Vérifier que tous les comptes de classe 6 et 7 sont soldés en fin d'exercice (virement au compte de résultat 12)",
                "Solder les comptes transitoires (471) - tout solde indique des écritures en suspens",
                "Solder les virements internes (58) - tout solde indique une erreur de lettrage",
                "Vérifier que la caisse (53) n'est jamais créditrice (impossible physiquement)",
                "Contrôler les stocks créditeurs (anormal - investigation obligatoire)",
                "Solder les comptes de régularisation (468, 475) en fin d'exercice",
                "Vérifier la cohérence des comptes de TVA (44551, 44567, 44558) avec les déclarations",
                "Contrôler les comptes de charges à payer et produits à recevoir (468) - doivent être soldés lors du paiement/encaissement",
                "Vérifier les comptes sociaux (431, 437) - normalement soldés chaque mois/trimestre",
                "Contrôler les comptes de rémunérations dues (422) - doivent être soldés lors de la paie suivante",
                "Contrôler la cohérence des amortissements (28) et dépréciations (29) avec les valeurs brutes",
                "Vérifier que les comptes de gestion (6 et 7) n'ont aucun solde à l'ouverture de l'exercice N+1",
                "Effectuer la concordance bancaire pour tous les comptes bancaires",
                "Vérifier la justification de tous les comptes débiteurs/créditeurs divers (467) avec soldes anciens"
            ]
        }
        
        return structure

    def _generer_tableau_comptes_a_solder(self):
        """
        Génère le tableau des comptes à solder obligatoirement avec numéro, moment, gravité
        """
        return [
            {
                "numero": "471",
                "libelle": "Comptes transitoires ou d'attente",
                "moment": "Fin d'exercice",
                "gravite": "CRITIQUE",
                "raison": "Tout solde indique des écritures en suspens non régularisées. Analyser ligne par ligne et régulariser chaque écriture."
            },
            {
                "numero": "58",
                "libelle": "Virements internes",
                "moment": "Immédiat",
                "gravite": "CRITIQUE",
                "raison": "Utilisé pour les virements entre banques/caisses. Tout solde indique une erreur de lettrage ou un virement non comptabilisé des deux côtés."
            },
            {
                "numero": "422",
                "libelle": "Personnel - Rémunérations dues",
                "moment": "Selon cycle paie",
                "gravite": "MOYENNE",
                "raison": "Doivent être soldées lors de la paie suivante. Des soldes anciens peuvent indiquer des erreurs de lettrage ou des paiements non régularisés."
            }
        ]

    def _verifier_presence_comptes_obligatoires(self, numeros_comptes_presents):
        """
        Vérifie la complétude par classe (1 à 6) selon SYSCOHADA.
        Retourne: (liste d'erreurs, liste détaillée des classes avec comptes présents/attendus).
        """
        erreurs = []
        classes_presence = []
        classes_attendues = {
            '1': {
                'libelle': "Classe 1 - Capitaux",
                'description': "Comptes 10 à 19 (capital, réserves, reports à nouveau, subventions, provisions réglementées, emprunts)",
                'exemples': ["101", "105", "12", "13", "16"]
            },
            '2': {
                'libelle': "Classe 2 - Immobilisations",
                'description': "Comptes 20 à 29 (immobilisations incorporelles, corporelles, financières, avances, amortissements, dépréciations)",
                'exemples': ["203", "213", "218", "281", "291"]
            },
            '3': {
                'libelle': "Classe 3 - Stocks",
                'description': "Comptes 30 à 39 (stocks de marchandises, matières, en-cours, produits, dépréciations)",
                'exemples': ["31", "32", "33", "37", "39"]
            },
            '4': {
                'libelle': "Classe 4 - Comptes de tiers",
                'description': "Comptes 40 à 49 (fournisseurs, clients, personnel, organismes sociaux, État, comptes divers, dépréciations)",
                'exemples': ["401", "4091", "411", "4191", "431", "445"]
            },
            '5': {
                'libelle': "Classe 5 - Comptes financiers",
                'description': "Comptes 50 à 59 (banques, caisse, VMP, virements internes, dépréciations)",
                'exemples': ["512", "53", "518", "519", "581"]
            },
            '6': {
                'libelle': "Classe 6 - Comptes de charges",
                'description': "Comptes 60 à 69 (achats, charges externes, impôts, charges de personnel, dotations, charges financières/exceptionnelles)",
                'exemples': ["601", "606", "62", "64", "68"]
            }
        }
        
        # Indexer les comptes présents par classe
        comptes_par_classe = {classe: [] for classe in classes_attendues.keys()}
        for numero in numeros_comptes_presents:
            if numero and numero[0] in comptes_par_classe:
                comptes_par_classe[numero[0]].append(numero)
        
        # Vérifier la présence de chaque classe
        for classe, infos in classes_attendues.items():
            comptes_classe = sorted(comptes_par_classe.get(classe, []))
            if comptes_classe:
                comptes_existants = ", ".join(comptes_classe)
                comptes_inexistants = infos['description']
                details = (
                    f"{len(comptes_classe)} compte(s) recensé(s) dans la classe {classe}. "
                    f"Exemples attendus : {', '.join(infos['exemples'])}."
                )
                status = "Présente"
            else:
                comptes_existants = "Aucun compte recensé dans la balance"
                comptes_inexistants = infos['description']
                details = (
                    f"{infos['libelle']} absente. Exemples attendus : {', '.join(infos['exemples'])}."
                )
                status = "Absente"
                message = (
                    f"{infos['libelle']} absente : aucun compte de cette classe n'est présent dans la balance. "
                    f"Attendu : {infos['description']}."
                )
                erreurs.append({
                    "type": "completude",
                    "numero_compte": "-",
                    "message": message,
                    "comptes_existants": comptes_existants,
                    "comptes_inexistants": comptes_inexistants,
                    "details": details,
                    "classe": infos['libelle']
                })

            classes_presence.append({
                "classe": infos['libelle'],
                "code_classe": classe,
                "status": status,
                "comptes_existants": comptes_existants,
                "comptes_inexistants": infos['description'],
                "details": details,
                "exemples": infos['exemples']
            })

        return erreurs, classes_presence

    def _valider_numerotation_syscohada(self, numero_compte):
        """
        Valide que la numérotation du compte suit le plan comptable SYSCOHADA
        Retourne: (True/False, message d'erreur si False)
        """
        if not numero_compte:
            return False, "Numéro de compte vide"
        
        # Nettoyer le numéro de compte (enlever espaces, caractères spéciaux)
        numero_clean = ''.join(c for c in numero_compte if c.isdigit())
        
        if not numero_clean:
            return False, f"Le numéro de compte '{numero_compte}' ne contient aucun chiffre valide"
        
        # Vérifier la longueur minimale (au moins 2 chiffres pour identifier la classe / sous-classe)
        if len(numero_clean) < 2:
            return False, f"Le numéro de compte '{numero_compte}' est trop court (minimum 2 chiffres requis)"
        
        # Vérifier que le premier chiffre correspond à une classe valide (1-9)
        premiere_classe = numero_clean[0]
        if premiere_classe not in ['1', '2', '3', '4', '5', '6', '7', '8', '9']:
            return False, f"Le numéro de compte '{numero_compte}' commence par '{premiere_classe}' qui n'est pas une classe valide (1-9 selon SYSCOHADA)"
        
        # Vérifier les plages de comptes selon SYSCOHADA
        # Les deux premiers chiffres doivent être dans la plage appropriée pour la classe
        if len(numero_clean) >= 2:
            deux_premiers = numero_clean[:2]
            try:
                deux_premiers_int = int(deux_premiers)
            except ValueError:
                return False, f"Le numéro de compte '{numero_compte}' contient des caractères non numériques invalides"
            
            classe = premiere_classe
            
            # Classe 1 : 10-19 (permet les sous-comptes comme 101, 102, etc.)
            if classe == '1' and not (10 <= deux_premiers_int <= 19):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 1 doit commencer par 10-19 selon SYSCOHADA"
            
            # Classe 2 : 20-29
            if classe == '2' and not (20 <= deux_premiers_int <= 29):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 2 doit commencer par 20-29 selon SYSCOHADA"
            
            # Classe 3 : 30-39
            if classe == '3' and not (30 <= deux_premiers_int <= 39):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 3 doit commencer par 30-39 selon SYSCOHADA"
            
            # Classe 4 : 40-49
            if classe == '4' and not (40 <= deux_premiers_int <= 49):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 4 doit commencer par 40-49 selon SYSCOHADA"
            
            # Classe 5 : 50-59
            if classe == '5' and not (50 <= deux_premiers_int <= 59):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 5 doit commencer par 50-59 selon SYSCOHADA"
            
            # Classe 6 : 60-69
            if classe == '6' and not (60 <= deux_premiers_int <= 69):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 6 doit commencer par 60-69 selon SYSCOHADA"
            
            # Classe 7 : 70-79
            if classe == '7' and not (70 <= deux_premiers_int <= 79):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 7 doit commencer par 70-79 selon SYSCOHADA"
            
            # Classe 8 : 80-89
            if classe == '8' and not (80 <= deux_premiers_int <= 89):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 8 doit commencer par 80-89 selon SYSCOHADA"
            
            # Classe 9 : 90-99
            if classe == '9' and not (90 <= deux_premiers_int <= 99):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 9 doit commencer par 90-99 selon SYSCOHADA"
        
        return True, None

    def _coherence_checks_for_year(self, lines):
        report = {"equilibre_global": True}
        erreurs = []
        numeros_comptes_vus = set()
        comptes_dupliques = []
        numeros_comptes_valides = []  # Pour vérifier la présence des comptes obligatoires
        # Dictionnaire pour compter les occurrences de chaque compte
        occurrences_comptes = {}  # {numero_compte: [indices des lignes]}

        # ===== CONTRÔLES ARITHMÉTIQUES =====
        # IMPORTANT : Les deux contrôles arithmétiques sont ESSENTIELS et doivent être effectués pour TOUS les comptes
        
        # 1) PREMIER CONTRÔLE ARITHMÉTIQUE : Vérifier que le total des débits est strictement égal au total des crédits
        # Ce test vérifie l'équilibre global de la balance en additionnant TOUS les débits finaux et TOUS les crédits finaux
        
        # Vérifier que lines n'est pas vide
        if not lines or len(lines) == 0:
            erreurs.append({
                "type": "equilibre",
                "numero_compte": "-",
                "message": "Aucune ligne de balance à analyser. Le premier contrôle arithmétique ne peut pas être effectué."
            })
            report["equilibre_global"] = False
            return report
        
        # Compter le nombre total de lignes analysées (toutes les lignes, même sans numéro de compte)
        nb_comptes_analyses = len([x for x in lines if x])  # Toutes les lignes non vides
        
        # Calculer les totaux : somme de TOUS les débits finaux et crédits finaux de TOUTES les lignes
        # Ce calcul inclut TOUTES les lignes, même celles sans numéro de compte valide
        sum_deb_fin = 0
        sum_cre_fin = 0
        
        # PREMIER CONTRÔLE ARITHMÉTIQUE : TOUJOURS EXÉCUTÉ
        # Calculer la somme de tous les débits finaux et crédits finaux
        lignes_avec_valeurs = 0
        for x in lines:
            try:
                deb = int(x.get("debit_fin", 0) or 0)
                cre = int(x.get("credit_fin", 0) or 0)
                sum_deb_fin += deb
                sum_cre_fin += cre
                if deb > 0 or cre > 0:
                    lignes_avec_valeurs += 1
            except (ValueError, TypeError):
                # Ignorer les valeurs invalides mais continuer le calcul
                continue
        
        # Le premier contrôle est maintenant exécuté : sum_deb_fin et sum_cre_fin sont calculés
        # DEBUG : Afficher les totaux calculés et les premières lignes pour vérifier
        print(f"[PREMIER CONTRÔLE] ========================================")
        print(f"[PREMIER CONTRÔLE] RÉSULTATS EXACTS DU CALCUL :")
        print(f"[PREMIER CONTRÔLE] Total débits finaux : {sum_deb_fin:,} FCFA")
        print(f"[PREMIER CONTRÔLE] Total crédits finaux : {sum_cre_fin:,} FCFA")
        print(f"[PREMIER CONTRÔLE] ÉCART : {abs(sum_deb_fin - sum_cre_fin):,} FCFA")
        print(f"[PREMIER CONTRÔLE] Nombre de lignes analysées : {nb_comptes_analyses}")
        print(f"[PREMIER CONTRÔLE] Lignes avec valeurs > 0 : {lignes_avec_valeurs}")
        print(f"[PREMIER CONTRÔLE] ========================================")
        
        # Afficher les 5 premières lignes avec valeurs pour déboguer
        if len(lines) > 0:
            print(f"[PREMIER CONTRÔLE] Exemple de données (5 premières lignes avec valeurs) :")
            count = 0
            for i, x in enumerate(lines):
                if count >= 5:
                    break
                deb = x.get("debit_fin", 0) or 0
                cre = x.get("credit_fin", 0) or 0
                num = x.get("numero_compte", "N/A")
                lib = str(x.get("libelle", ""))[:30]
                if int(deb) > 0 or int(cre) > 0:
                    print(f"  Ligne {i+1}: compte={num}, libelle={lib}, debit_fin={deb:,}, credit_fin={cre:,}, solde={abs(int(deb)-int(cre)):,}")
                    count += 1
        
        # TOUJOURS exécuter le premier contrôle arithmétique
        # Vérifier si les totaux sont égaux
        if sum_deb_fin != sum_cre_fin:
            report["equilibre_global"] = False
            ecart = abs(sum_deb_fin - sum_cre_fin)
            
            # Identifier TOUS les comptes pour l'affichage
            comptes_anormaux = []
            for x in lines:
                numero_compte = str(x.get("numero_compte", "")).strip()
                if numero_compte and numero_compte != "None" and numero_compte.lower() != "nan":
                    df = int(x.get("debit_fin", 0) or 0)
                    cf = int(x.get("credit_fin", 0) or 0)
                    solde = abs(df - cf)
                    libelle = x.get("libelle", "") or x.get("libelle_compte", "") or "Sans libellé"
                    # Afficher TOUS les comptes, même ceux avec solde = 0
                    comptes_anormaux.append({
                        "compte": numero_compte,
                        "libelle": libelle,
                        "debit_fin": df,
                        "credit_fin": cf,
                        "solde": solde
                    })
            
            # Trier par solde décroissant - AFFICHER TOUS LES COMPTES
            comptes_anormaux.sort(key=lambda x: x["solde"], reverse=True)
            
            # Message d'alerte justifié et clair
            message = f"ALERTE : Déséquilibre détecté dans la balance\n\n"
            message += f"ÉCART DÉTECTÉ : {ecart:,} FCFA\n\n"
            
            if sum_deb_fin > sum_cre_fin:
                message += f"Justification : Le total des DÉBITS ({sum_deb_fin:,} FCFA) est supérieur au total des CRÉDITS ({sum_cre_fin:,} FCFA).\n"
                message += f"Il manque {ecart:,} FCFA au crédit (ou il y a {ecart:,} FCFA en trop au débit).\n\n"
            else:
                message += f"Justification : Le total des CRÉDITS ({sum_cre_fin:,} FCFA) est supérieur au total des DÉBITS ({sum_deb_fin:,} FCFA).\n"
                message += f"Il manque {ecart:,} FCFA au débit (ou il y a {ecart:,} FCFA en trop au crédit).\n\n"
            
            message += f"Principe comptable : En comptabilité en partie double, le total des débits DOIT être STRICTEMENT ÉGAL au total des crédits.\n"
            message += f"Toute différence, même de 1 FCFA, indique FORCÉMENT une erreur dans les écritures comptables.\n\n"
            
            if comptes_anormaux:
                message += f"Comptes concernés ({len(comptes_anormaux)} compte(s)) :\n"
                for compte_info in comptes_anormaux:
                    message += f"{compte_info['compte']} - {compte_info['libelle']} : Débit {compte_info['debit_fin']:,} | Crédit {compte_info['credit_fin']:,} | Solde {compte_info['solde']:,} FCFA\n"
            
            # Ajouter l'erreur d'équilibre à la liste des erreurs
            erreur_equilibre = {
                "type": "equilibre",
                "numero_compte": "-",
                "message": message
            }
            erreurs.append(erreur_equilibre)
            print(f"[PREMIER CONTRÔLE] Erreur d'équilibre ajoutée : type={erreur_equilibre['type']}, message_length={len(message)}")
            
            # Ajouter l'écart dans le rapport pour affichage dans le frontend
            report["ecart_equilibre"] = ecart
            report["total_debits"] = sum_deb_fin
            report["total_credits"] = sum_cre_fin
            
            # Créer la structure verification_equilibre même en cas de déséquilibre pour afficher les résultats exacts
            if sum_deb_fin > sum_cre_fin:
                explication = f"Le système a détecté un DÉSÉQUILIBRE : le total des débits ({sum_deb_fin:,} FCFA) est supérieur au total des crédits ({sum_cre_fin:,} FCFA). Il manque {ecart:,} FCFA au crédit (ou il y a {ecart:,} FCFA en trop au débit)."
            else:
                explication = f"Le système a détecté un DÉSÉQUILIBRE : le total des crédits ({sum_cre_fin:,} FCFA) est supérieur au total des débits ({sum_deb_fin:,} FCFA). Il manque {ecart:,} FCFA au débit (ou il y a {ecart:,} FCFA en trop au crédit)."
            
            report["verification_equilibre"] = {
                "statut": "ERREUR",
                "total_debits": sum_deb_fin,
                "total_credits": sum_cre_fin,
                "ecart": ecart,
                "nb_comptes_analyses": nb_comptes_analyses,
                "explication": explication
            }
            
            print(f"[PREMIER CONTRÔLE] Rapport mis à jour : equilibre_global=False, ecart={ecart:,} FCFA, total_debits={sum_deb_fin:,} FCFA, total_credits={sum_cre_fin:,} FCFA")
        else:
            # Les totaux sont égaux - équilibre OK
            report["equilibre_global"] = True
            print(f"[PREMIER CONTRÔLE] Équilibre OK : Total débits = Total crédits = {sum_deb_fin:,} FCFA")
            
            # Ajouter une information de vérification réussie dans le rapport
            report["verification_equilibre"] = {
                "statut": "OK",
                "total_debits": sum_deb_fin,
                "total_credits": sum_cre_fin,
                "ecart": 0,
                "nb_comptes_analyses": nb_comptes_analyses,
                "explication": f"Le système a vérifié que le total des débits ({sum_deb_fin:,} FCFA) est strictement égal au total des crédits ({sum_cre_fin:,} FCFA) en additionnant les colonnes 'Débit fin' et 'Crédit fin' de {nb_comptes_analyses} compte(s)."
            }
            report["ecart_equilibre"] = 0  # Pas d'écart si équilibré
            report["total_debits"] = sum_deb_fin
            report["total_credits"] = sum_cre_fin
        
        # IMPORTANT : Le premier contrôle arithmétique est TOUJOURS exécuté
        # Il vérifie que sum_deb_fin == sum_cre_fin
        # Si les totaux diffèrent, une erreur de type "equilibre" est ajoutée à la liste des erreurs
        # Même si les totaux sont égaux, le contrôle a été effectué et doit être documenté dans le rapport
        
        # S'assurer que les informations du premier contrôle sont TOUJOURS dans le rapport
        # (même si elles ont déjà été créées dans le bloc if/else ci-dessus)
        if "verification_equilibre" not in report:
            # Si le contrôle n'a pas créé verification_equilibre, le créer maintenant
            report["verification_equilibre"] = {
                "statut": "OK" if report.get("equilibre_global", True) else "ERREUR",
                "total_debits": sum_deb_fin,
                "total_credits": sum_cre_fin,
                "nb_comptes_analyses": nb_comptes_analyses,
                "explication": f"Le système a vérifié que le total des débits ({sum_deb_fin:,} FCFA) est strictement égal au total des crédits ({sum_cre_fin:,} FCFA) en additionnant les colonnes 'Débit fin' et 'Crédit fin' de {nb_comptes_analyses} compte(s)."
            }
        
        # Confirmer que le premier contrôle a été exécuté
        print(f"[PREMIER CONTRÔLE ARITHMÉTIQUE] Exécuté : {nb_comptes_analyses} ligne(s) analysée(s), Total débits = {sum_deb_fin:,} FCFA, Total crédits = {sum_cre_fin:,} FCFA, Équilibre = {'OK' if report.get('equilibre_global', True) else 'DÉSÉQUILIBRÉ'}")

        # 2) Détection des erreurs d'identité, de signe, arithmétiques et des comptes à solder
        nb_erreurs_identite = 0
        nb_erreurs_signe = 0
        nb_erreurs_arithmetique = 0
        nb_erreurs_comptes_soldes = 0
        
        # Compteurs pour le contrôle de vraisemblance
        nb_comptes_verifies_vraisemblance = 0
        nb_comptes_vraisemblance_ok = 0
        nb_comptes_erreur_signe = 0
        nb_comptes_erreur_soldes = 0
        
        # Compteurs pour le second contrôle arithmétique (formule solde d'ouverture + mouvements = solde de clôture)
        nb_comptes_verifies_formule = 0
        nb_comptes_formule_ok = 0
        nb_comptes_formule_erreur = 0
        
        print(f"[SECOND CONTRÔLE] Début de la vérification de la formule pour {len(lines)} ligne(s) de balance")
        
        # Première passe : collecter tous les numéros de comptes (on ne signale plus les doublons ici)
        for idx, x in enumerate(lines):
            numero_compte = str(x.get("numero_compte", "")).strip()
            if not numero_compte or numero_compte == "None" or numero_compte.lower() == "nan":
                continue
            numero_compte_normalise = numero_compte.strip()
            if numero_compte_normalise not in occurrences_comptes:
                occurrences_comptes[numero_compte_normalise] = []
            occurrences_comptes[numero_compte_normalise].append(idx)
        
        # Deuxième passe : vérifier les autres erreurs
        print(f"[SECOND CONTRÔLE] Début de la deuxième passe : vérification de la formule pour tous les comptes")
        for idx, x in enumerate(lines):
            numero_compte = str(x.get("numero_compte", "")).strip()
            if not numero_compte or numero_compte == "None" or numero_compte.lower() == "nan":
                # Pour les lignes sans numéro de compte, on utilise un identifiant temporaire
                numero_compte = f"LIGNE_SANS_NUMERO_{len(erreurs)}"
            
            # 2.1) Vérifier la numérotation SYSCOHADA (pour information, mais ne bloque pas)
            # On ne bloque plus sur la numérotation SYSCOHADA dans ce contrôle :
            # le contrôle de complétude/exhaustivité a été désactivé à la demande de l'utilisateur.
            is_valid, error_msg = self._valider_numerotation_syscohada(numero_compte)
            # On continue même si la validation SYSCOHADA échoue, car la vérification de la formule
            # est une vérification arithmétique fondamentale qui doit être faite pour TOUS les comptes
            
            # 2.2) Ajouter aux comptes valides pour vérification (si pas déjà en doublon)
            if numero_compte not in comptes_dupliques:
                if numero_compte not in numeros_comptes_vus:
                    numeros_comptes_vus.add(numero_compte)
                    if not numero_compte.startswith("LIGNE_SANS_NUMERO_"):
                        numeros_comptes_valides.append(numero_compte)
                
            di = int(x.get("debit_initial", 0) or 0)
            ci = int(x.get("credit_initial", 0) or 0)
            df = int(x.get("debit_fin", 0) or 0)
            cf = int(x.get("credit_fin", 0) or 0)
            
            # Calcul des soldes
            solde_initial = di - ci  # Solde d'ouverture
            solde_fin = df - cf  # Solde de clôture
            
            # ===== SECOND CONTRÔLE ARITHMÉTIQUE =====
            # IMPORTANT : Cette vérification doit être effectuée pour TOUS les comptes,
            # indépendamment de la validation SYSCOHADA, car c'est une vérification arithmétique fondamentale
            # 
            # Ce test vérifie la cohérence des soldes de CHAQUE compte individuellement :
            # Vérifier que : Solde de clôture = Solde d'ouverture + Mouvements de période
            # Formule : Solde de clôture = Solde d'ouverture + (Mouvement débit - Mouvement crédit)
            # Où :
            #   - Solde d'ouverture = Débit initial - Crédit initial
            #   - Solde de clôture = Débit fin - Crédit fin
            #   - Mouvements = (Débit fin - Débit initial) - (Crédit fin - Crédit initial)
            
            # Compter ce compte comme vérifié
            nb_comptes_verifies_formule += 1
            
            # Log pour tous les comptes (pour debug)
            if numero_compte == "46210010":
                print(f"[SECOND CONTRÔLE] COMPTE 46210010 DÉTECTÉ - Début de la vérification")
            
            # Récupérer les mouvements explicites s'ils sont fournis
            # Les mouvements peuvent être stockés sous différents noms selon le format d'import
            # IMPORTANT : Vérifier si les colonnes de mouvements EXISTENT dans les données
            # (même si elles sont vides/0, si elles existent, on doit les utiliser)
            # Cela permet de détecter les erreurs où mouvements=0 mais solde final != solde initial
            
            # Vérifier si les clés de mouvements existent dans le dictionnaire
            debit_explicite_present = "mouvement_debit" in x or "debit_mvt" in x
            credit_explicite_present = "mouvement_credit" in x or "credit_mvt" in x
            
            # Récupérer les valeurs (même si elles sont 0 ou None)
            if debit_explicite_present:
                mouvement_debit_explicite = x.get("mouvement_debit") or x.get("debit_mvt") or 0
                try:
                    mouvement_debit_explicite = int(mouvement_debit_explicite or 0)
                except (ValueError, TypeError):
                    mouvement_debit_explicite = 0
            else:
                mouvement_debit_explicite = 0
            
            if credit_explicite_present:
                mouvement_credit_explicite = x.get("mouvement_credit") or x.get("credit_mvt") or 0
                try:
                    mouvement_credit_explicite = int(mouvement_credit_explicite or 0)
                except (ValueError, TypeError):
                    mouvement_credit_explicite = 0
            else:
                mouvement_credit_explicite = 0
            
            # Calculer les mouvements à partir des débits/crédits (pour vérification)
            mouvement_debit_calcule = df - di
            mouvement_credit_calcule = cf - ci
            
            # IMPORTANT : Si les mouvements explicites sont PRÉSENTS dans les données (même à 0),
            # on DOIT les utiliser pour la vérification. Sinon, on utilise les mouvements calculés.
            # Cela permet de détecter les erreurs où les mouvements sont à 0 mais le solde final n'est pas à 0.
            if debit_explicite_present or credit_explicite_present:
                # Utiliser les mouvements explicites (même s'ils sont à 0)
                mouvement_debit = mouvement_debit_explicite
                mouvement_credit = mouvement_credit_explicite
                source_mouvements = "explicites"
            else:
                # Utiliser les mouvements calculés (quand les colonnes de mouvements n'existent pas)
                mouvement_debit = mouvement_debit_calcule
                mouvement_credit = mouvement_credit_calcule
                source_mouvements = "calculés"
            
            # Calculer le solde de clôture attendu selon la formule
            # Solde de clôture = Solde d'ouverture + (Mouvement débit - Mouvement crédit)
            solde_cloture_attendu = solde_initial + (mouvement_debit - mouvement_credit)
            
            # Vérifier la cohérence avec une tolérance STRICTE de 0 FCFA (pas d'arrondi accepté)
            # En comptabilité, la formule doit être STRICTEMENT respectée
            ecart = abs(solde_fin - solde_cloture_attendu)
            
            # Log détaillé pour TOUS les comptes avec écart > 0 (même très petit)
            if ecart > 0:
                print(f"[SECOND CONTRÔLE] Compte {numero_compte}:")
                print(f"  Solde initial: {solde_initial:,} (Débit init {di:,} - Crédit init {ci:,})")
                print(f"  Mouvements: Débit {mouvement_debit:,} | Crédit {mouvement_credit:,} | Net {mouvement_debit - mouvement_credit:,} (source: {source_mouvements})")
                print(f"  Solde attendu: {solde_cloture_attendu:,} FCFA")
                print(f"  Solde réel: {solde_fin:,} FCFA (Débit fin {df:,} - Crédit fin {cf:,})")
                print(f"  ÉCART: {ecart:,} FCFA")
                print(f"  Erreur détectée: {'OUI' if ecart > 0 else 'NON'}")
            
            # Tolérance STRICTE : tout écart > 0 est une erreur
            if ecart > 0:
                nb_comptes_formule_erreur += 1
                nb_erreurs_arithmetique += 1
                print(f"[SECOND CONTRÔLE] Erreur détectée pour compte {numero_compte}: écart={ecart:,} FCFA (solde_fin={solde_fin:,}, attendu={solde_cloture_attendu:,})")
                
                # Récupérer le libellé du compte
                libelle = x.get("libelle", "") or x.get("libelle_compte", "") or "Sans libellé"
                
                # Message d'alerte justifié et clair
                compte_display = "[SANS NUMÉRO]" if numero_compte.startswith("LIGNE_SANS_NUMERO_") else numero_compte
                message = f"ALERTE : Formule non respectée pour le compte {compte_display}\n"
                message += f"Libellé : {libelle}\n\n"
                message += f"ÉCART DÉTECTÉ : {ecart:,} FCFA\n\n"
                message += f"Formule attendue : Solde de clôture = Solde d'ouverture + Mouvements de période\n"
                message += f"Solde attendu : {solde_cloture_attendu:,} FCFA\n"
                message += f"Solde réel : {solde_fin:,} FCFA\n"
                message += f"Écart : {ecart:,} FCFA\n\n"
                message += f"Détails des valeurs :\n"
                message += f"Solde d'ouverture : {solde_initial:,} FCFA (Débit initial {di:,} - Crédit initial {ci:,})\n"
                message += f"Mouvements : Débit {mouvement_debit:,} | Crédit {mouvement_credit:,} | Net {mouvement_debit - mouvement_credit:,} FCFA\n"
                message += f"Solde de clôture : {solde_fin:,} FCFA (Débit fin {df:,} - Crédit fin {cf:,})\n\n"
                message += f"Justification : La formule comptable fondamentale n'est pas respectée. "
                message += f"Le solde de clôture devrait être {solde_cloture_attendu:,} FCFA mais il est {solde_fin:,} FCFA. "
                message += f"Cet écart de {ecart:,} FCFA indique une erreur dans les écritures comptables de ce compte."
                
                erreurs.append({
                    "type": "arithmetique",
                    "numero_compte": numero_compte,
                    "message": message
                })
            else:
                # La formule est respectée pour ce compte
                nb_comptes_formule_ok += 1
                # Log pour le compte 46210010 même si pas d'erreur
                if numero_compte == "46210010":
                    print(f"[SECOND CONTRÔLE] Compte 46210010: Formule respectée (écart={ecart:.2f} <= tolérance 0.01)")
            # ===== FIN DU SECOND CONTRÔLE ARITHMÉTIQUE =====
            
            # ===== CONTRÔLE DE VRAISEMBLANCE =====
            # Ce contrôle vérifie la cohérence des soldes selon les règles comptables SYSCOHADA
            # 1. Vérification du sens du solde (débit/crédit) selon la classe de compte
            # 2. Vérification des cas critiques (caisse créditrice, stock créditeur)
            # 3. Vérification des comptes qui doivent être soldés en fin d'exercice
            
            # Compter ce compte comme vérifié pour le contrôle de vraisemblance
            nb_comptes_verifies_vraisemblance += 1
            
            # Variable pour suivre si ce compte a des erreurs de vraisemblance
            compte_erreur_vraisemblance = False
            
            # Récupérer le solde réel pour les autres vérifications (vraisemblance)
            solde_reel = int(x.get("solde_reel", 0) or 0)

            # Vérifier le sens du solde selon la classe de compte
            sign = x.get("sign_solde")
            expect_sign = "D" if solde_reel >= 0 else "C"
            
            # Vérifier le sens attendu selon le plan comptable SYSCOHADA
            sens_attendu = self._get_sens_attendu_compte(numero_compte)
            
            # Vérifications spéciales pour cas critiques
            classe = numero_compte[0] if numero_compte[0].isdigit() else '?'
            
            # CAS CRITIQUE 1 : Caisse créditrice (impossible physiquement)
            if numero_compte.startswith('53') and solde_reel < 0:
                nb_erreurs_signe += 1
                nb_comptes_erreur_signe += 1
                compte_erreur_vraisemblance = True
                erreurs.append({
                    "type": "signe",
                    "numero_compte": numero_compte,
                    "message": f"ERREUR CRITIQUE : Compte de caisse {numero_compte} avec solde CRÉDITEUR ({abs(solde_reel):,} FCFA). C'est IMPOSSIBLE physiquement ! Une caisse ne peut pas avoir un solde créditeur. Vérifiez immédiatement les écritures comptables - erreur certaine."
                })
                continue  # Ne pas continuer avec les autres vérifications
            
            # CAS CRITIQUE 2 : Stock créditeur (anormal - investigation obligatoire)
            if classe == '3' and not numero_compte.startswith('39') and solde_reel < 0:
                nb_erreurs_signe += 1
                nb_comptes_erreur_signe += 1
                compte_erreur_vraisemblance = True
                erreurs.append({
                    "type": "signe",
                    "numero_compte": numero_compte,
                    "message": f"ANOMALIE CRITIQUE : Compte de stock {numero_compte} avec solde CRÉDITEUR ({abs(solde_reel):,} FCFA). Un stock créditeur est ANORMAL et doit être investigué. Vérifiez les écritures d'inventaire, les erreurs de saisie ou les problèmes de valorisation."
                })
                continue  # Ne pas continuer avec les autres vérifications
            
            if sens_attendu != 'BOTH':
                # Le compte a un sens attendu spécifique selon sa classe
                sens_classe = {
                    '1': 'CRÉDITEUR',
                    '2': 'DÉBITEUR',
                    '3': 'DÉBITEUR',
                    '4': 'Variable selon sous-classe',
                    '5': 'DÉBITEUR',
                    '6': 'DÉBITEUR',
                    '7': 'CRÉDITEUR',
                    '8': 'Variable'
                }
                sens_classe_attendu = sens_classe.get(classe, 'Variable')
                
                if sens_attendu == 'D' and solde_reel < 0:
                    # Compte devrait être débiteur mais est créditeur
                    nb_erreurs_signe += 1
                    nb_comptes_erreur_signe += 1
                    compte_erreur_vraisemblance = True
                    # Messages spécifiques selon le type de compte
                    if classe == '3':
                        message = f"ANOMALIE CRITIQUE : Compte de stock {numero_compte} avec solde CRÉDITEUR ({abs(solde_reel):,} FCFA). Un stock créditeur est ANORMAL et doit être investigué. Vérifiez les écritures d'inventaire, les erreurs de saisie ou les problèmes de valorisation."
                    elif classe == '6':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Charges). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être DÉBITEUR, mais le solde est CRÉDITEUR ({abs(solde_reel):,} FCFA). Un compte de charges créditeur indique généralement une erreur de comptabilisation ou un avoir mal enregistré. Charge créditrice : Généralement anormal, sauf cas spécifiques d'ajustements. Vérifiez les écritures comptables."
                    elif classe == '2':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Immobilisations). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être DÉBITEUR, mais le solde est CRÉDITEUR ({abs(solde_reel):,} FCFA). Un compte d'immobilisation créditeur est généralement anormal (peut indiquer une erreur ou une cession non comptabilisée correctement). Vérifiez les écritures comptables."
                    else:
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe}). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être DÉBITEUR, mais le solde est CRÉDITEUR ({abs(solde_reel):,} FCFA). Vérifiez la nature du compte et les écritures comptables."
                    
                    erreurs.append({
                        "type": "signe",
                        "numero_compte": numero_compte,
                        "message": message
                    })
                elif sens_attendu == 'C' and solde_reel > 0:
                    # Compte devrait être créditeur mais est débiteur
                    nb_erreurs_signe += 1
                    nb_comptes_erreur_signe += 1
                    compte_erreur_vraisemblance = True
                    # Messages spécifiques selon le type de compte
                    if classe == '1':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Capitaux). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être CRÉDITEUR, mais le solde est DÉBITEUR ({solde_reel:,} FCFA). Vérifiez la nature du compte (report à nouveau débiteur, associés non appelé, primes de remboursement) et les écritures comptables."
                    elif classe == '2':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Immobilisations). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être CRÉDITEUR, mais le solde est DÉBITEUR ({solde_reel:,} FCFA). Un compte d'immobilisation créditeur est généralement anormal (peut indiquer une erreur ou une cession non comptabilisée correctement). Vérifiez les écritures comptables."
                    elif classe == '6':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Charges). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être DÉBITEUR, mais le solde est CRÉDITEUR ({abs(solde_reel):,} FCFA). Un compte de charges créditeur indique généralement une erreur de comptabilisation ou un avoir mal enregistré. Vérifiez les écritures comptables."
                    elif classe == '7':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Produits). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être CRÉDITEUR, mais le solde est DÉBITEUR ({solde_reel:,} FCFA). Un compte de produits débiteur indique généralement une erreur (annulation mal comptabilisée, avoir sur vente). Produit débiteur : Généralement anormal, sauf cas spécifiques d'ajustements. Vérifiez les écritures comptables."
                    else:
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe}). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être CRÉDITEUR, mais le solde est DÉBITEUR ({solde_reel:,} FCFA). Vérifiez la nature du compte et les écritures comptables."
                    
                    erreurs.append({
                        "type": "signe",
                        "numero_compte": numero_compte,
                        "message": message
                    })
            
            # Vérifier aussi le signe enregistré
            if sign not in ("D", "C") or sign != expect_sign:
                nb_erreurs_signe += 1
                nb_comptes_erreur_signe += 1
                compte_erreur_vraisemblance = True
                erreurs.append({
                    "type": "signe",
                    "numero_compte": numero_compte,
                    "message": f"Erreur de signe pour le compte {numero_compte}. Solde réel: {solde_reel:,}, Signe attendu: {expect_sign}, Signe trouvé: {sign if sign else 'N/A'}."
                })
            
            # Vérifier si le compte doit être soldé
            doit_etre_solde, gravite, message_solde = self._comptes_doivent_etre_soldes(numero_compte)
            if doit_etre_solde and abs(solde_fin) > 0.01:
                nb_erreurs_comptes_soldes += 1
                nb_comptes_erreur_soldes += 1
                compte_erreur_vraisemblance = True
                erreurs.append({
                    "type": "compte_non_solde",
                    "numero_compte": numero_compte,
                    "gravite": gravite,
                    "message": f"[{gravite}] {message_solde} Compte {numero_compte} a un solde de {solde_fin:,} FCFA. Vérifiez et régularisez."
                })
            
            # Si aucune erreur de vraisemblance détectée, compter comme OK
            if not compte_erreur_vraisemblance:
                nb_comptes_vraisemblance_ok += 1
            # ===== FIN DU CONTRÔLE DE VRAISEMBLANCE =====

        # 3) Vérifier la présence des comptes obligatoires (désactivé : contrôle de complétude retiré)
        # erreurs_presence, classes_presence = self._verifier_presence_comptes_obligatoires(numeros_comptes_valides)
        # if erreurs_presence:
        #     erreurs.extend(erreurs_presence)
        # report["classes_presence"] = classes_presence

        # Ajouter les statistiques de vérification de la formule (second contrôle arithmétique)
        if nb_comptes_verifies_formule > 0:
            statut_formule = "OK" if nb_comptes_formule_erreur == 0 else "ERREUR"
            
            if nb_comptes_formule_erreur == 0:
                explication = f"Le système a vérifié la formule 'Solde de clôture = Solde d'ouverture + Mouvements de période' pour {nb_comptes_verifies_formule} compte(s). Tous les comptes ({nb_comptes_formule_ok}) respectent correctement la formule comptable fondamentale."
            else:
                explication = f"Le système a vérifié la formule 'Solde de clôture = Solde d'ouverture + Mouvements de période' pour {nb_comptes_verifies_formule} compte(s). {nb_comptes_formule_ok} compte(s) respectent la formule, mais {nb_comptes_formule_erreur} compte(s) présentent des ERREURS. Les comptes en erreur sont listés ci-dessous avec les détails de l'écart détecté."
            
            report["verification_formule"] = {
                "statut": statut_formule,
                "nb_comptes_verifies": nb_comptes_verifies_formule,
                "nb_comptes_ok": nb_comptes_formule_ok,
                "nb_comptes_erreur": nb_comptes_formule_erreur,
                "explication": explication
            }
            
            print(f"[SECOND CONTRÔLE] Rapport créé : statut={statut_formule}, vérifiés={nb_comptes_verifies_formule}, OK={nb_comptes_formule_ok}, Erreurs={nb_comptes_formule_erreur}")

        # Ajouter les statistiques de vérification de vraisemblance avec structure complète
        # IMPORTANT : La structure (tableaux des classes, comptes à solder) est la même pour toutes les balances
        # car ce sont des règles générales SYSCOHADA. Cependant, les statistiques (erreurs détectées) 
        # sont calculées individuellement pour chaque balance et doivent être différentes.
        if nb_comptes_verifies_vraisemblance > 0:
            statut_vraisemblance = "OK" if (nb_comptes_erreur_signe == 0 and nb_comptes_erreur_soldes == 0) else "ERREUR"
            
            if nb_comptes_erreur_signe == 0 and nb_comptes_erreur_soldes == 0:
                explication = f"Le système a vérifié la vraisemblance des soldes pour {nb_comptes_verifies_vraisemblance} compte(s). Tous les comptes ({nb_comptes_vraisemblance_ok}) respectent les règles comptables SYSCOHADA (sens des soldes et comptes à solder)."
            else:
                explication = f"Le système a vérifié la vraisemblance des soldes pour {nb_comptes_verifies_vraisemblance} compte(s). {nb_comptes_vraisemblance_ok} compte(s) sont conformes, mais {nb_comptes_erreur_signe} compte(s) présentent des anomalies de sens et {nb_comptes_erreur_soldes} compte(s) devraient être soldés. Les comptes en erreur sont listés ci-dessous."
            
            # Générer la structure complète du contrôle de vraisemblance avec les erreurs détectées pour cette balance
            # NOTE : La structure de base (règles générales) est identique pour toutes les balances,
            # mais les anomalies détectées sont spécifiques à cette balance
            structure_vraisemblance = self._generer_structure_vraisemblance(erreurs)
            
            report["verification_vraisemblance"] = {
                "statut": statut_vraisemblance,
                "nb_comptes_verifies": nb_comptes_verifies_vraisemblance,
                "nb_comptes_ok": nb_comptes_vraisemblance_ok,
                "nb_comptes_erreur_signe": nb_comptes_erreur_signe,
                "nb_comptes_erreur_soldes": nb_comptes_erreur_soldes,
                "explication": explication,
                "structure": structure_vraisemblance
            }
            
            print(f"[CONTRÔLE VRAISEMBLANCE] Rapport créé pour cette balance : statut={statut_vraisemblance}, vérifiés={nb_comptes_verifies_vraisemblance}, OK={nb_comptes_vraisemblance_ok}, Erreurs signe={nb_comptes_erreur_signe}, Erreurs soldes={nb_comptes_erreur_soldes}")
            print(f"[CONTRÔLE VRAISEMBLANCE] NOTE : La structure (tableaux des classes, comptes à solder) est identique pour toutes les balances car ce sont des règles générales SYSCOHADA.")
            print(f"[CONTRÔLE VRAISEMBLANCE] Les statistiques ci-dessus sont spécifiques à cette balance et peuvent différer d'une balance à l'autre.")

        # Récap totaux pour UI
        report["totaux"] = {
            "debit_fin": sum_deb_fin,
            "credit_fin": sum_cre_fin,
            "nb_erreurs_identite": nb_erreurs_identite,
            "nb_erreurs_signe": nb_erreurs_signe,
            "nb_erreurs_arithmetique": nb_erreurs_arithmetique,
            "nb_erreurs_comptes_soldes": nb_erreurs_comptes_soldes,
            "nb_erreurs_total": nb_erreurs_identite + nb_erreurs_signe + nb_erreurs_arithmetique + nb_erreurs_comptes_soldes,
            "nb_comptes_dupliques": len(comptes_dupliques),
            "nb_comptes_verifies_formule": nb_comptes_verifies_formule,
            "nb_comptes_formule_ok": nb_comptes_formule_ok,
            "nb_comptes_formule_erreur": nb_comptes_formule_erreur
        }
        
        # Détails des erreurs avec numéros de comptes
        report["erreurs"] = erreurs
        
        # DEBUG : Afficher le résumé final
        print(f"\n{'='*60}")
        print(f"[RAPPORT FINAL] RÉSUMÉ DU CONTRÔLE DE COHÉRENCE")
        print(f"{'='*60}")
        print(f"Nombre total d'erreurs détectées : {len(erreurs)}")
        print(f"Nombre de comptes vérifiés (formule) : {nb_comptes_verifies_formule}")
        print(f"  - Comptes OK : {nb_comptes_formule_ok}")
        print(f"  - Comptes en erreur : {nb_comptes_formule_erreur}")
        print(f"\nErreurs par type :")
        erreurs_par_type = {}
        for e in erreurs:
            t = e.get("type", "autre")
            erreurs_par_type[t] = erreurs_par_type.get(t, 0) + 1
        for t, count in erreurs_par_type.items():
            print(f"  - {t}: {count}")
        print(f"\nEquilibre global : {report.get('equilibre_global', 'N/A')}")
        print(f"{'='*60}\n")

        return report

    def controle_coherence(self, id_mission):
        print(f"=== DÉBUT contrôle_coherence pour mission {id_mission} ===")

        local_db = get_db()
        if local_db is None:
            raise RuntimeError("Base de données non connectée")

        mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
        if not mission:
            return {"ok": False, "message": "Mission non trouvée", "data": []}

        bal_ids = mission.get("balances", [])
        if not bal_ids:
            return {"ok": False, "message": "Aucune balance trouvée", "data": []}

        annee_auditee = int(mission.get("annee_auditee", 0))
        resultats = []

        for idx, bal_id in enumerate(bal_ids):
            annee = annee_auditee - idx
            lines = self._load_balance(bal_id)

            rapport = self._coherence_checks_for_year(lines)

            resultats.append({
                "annee": annee,
                "rapport": rapport
            })

        payload = {
            "ok": True,
            "data": resultats
        }

        local_db.Mission1.update_one(
            {"_id": ObjectId(id_mission)},
            {"$set": {"controle_coherence": payload}}
        )

        print("=== FIN contrôle_coherence ===")
        return payload


    def _index_by_compte(self, lines):
        # Filtrer les lignes qui ont un numero_compte valide (non vide, non None)
        index = {}
        lignes_ignorees = 0
        lignes_ignorees_none = 0
        lignes_ignorees_vide = 0

        def _normalize_compte_key(raw_value):
            if raw_value is None:
                return None
            s = str(raw_value).strip()
            if not s or s == "None" or s.lower() == "nan":
                return None
            # Normaliser les formats fréquents d'import Excel: "401.0" -> "401"
            if s.endswith(".0"):
                s = s[:-2]
            # Supprimer les espaces internes accidentels
            s = s.replace(" ", "")
            return s or None

        def _to_float(value):
            try:
                if value is None or value == "":
                    return 0.0
                return float(value)
            except Exception:
                return 0.0
        
        if not lines:
            print("[INDEX_BY_COMPTE] liste de lignes vide")
            return index
        
        print(f"[INDEX_BY_COMPTE] traitement de {len(lines)} lignes")
        
        for idx, x in enumerate(lines):
            if not x:
                lignes_ignorees += 1
                continue
                
            num_compte = x.get("numero_compte")
            
            # Debug pour les premières lignes
            if idx < 3:
                print(f"Ligne {idx}: numero_compte={num_compte} (type: {type(num_compte)}, bool: {bool(num_compte)})")
            
            # Gérer le cas où num_compte pourrait être 0 (qui est False mais valide comme numéro)
            if num_compte is None:
                lignes_ignorees_none += 1
                continue
            
            num_str = _normalize_compte_key(num_compte)
            if not num_str:
                lignes_ignorees_vide += 1
                continue

            # Agréger les doublons du même compte au lieu d'écraser la ligne précédente.
            # Cela évite de "perdre" des comptes lorsque la balance contient plusieurs lignes par numéro.
            if num_str not in index:
                row = dict(x)
                row["numero_compte"] = num_str
                index[num_str] = row
            else:
                row = index[num_str]
                for fld in (
                    "debit_initial",
                    "credit_initial",
                    "debit_mouvement",
                    "credit_mouvement",
                    "debit_fin",
                    "credit_fin",
                ):
                    row[fld] = _to_float(row.get(fld, 0)) + _to_float(x.get(fld, 0))
                if not row.get("libelle") and x.get("libelle"):
                    row["libelle"] = x.get("libelle")
        
        print(f"[INDEX_BY_COMPTE] {len(index)} comptes indexes")
        if lignes_ignorees > 0:
            print(f"   Lignes None ignorées: {lignes_ignorees}")
        if lignes_ignorees_none > 0:
            print(f"   Numéros None ignorés: {lignes_ignorees_none}")
        if lignes_ignorees_vide > 0:
            print(f"   Numéros vides ignorés: {lignes_ignorees_vide}")
        
        return index

    def controle_intangibilite(self, id_mission):
        try:
            print(f"[CONTROLE_INTANGIBILITE] ========== DEBUT ==========")
            print(f"[CONTROLE_INTANGIBILITE] Mission ID: {id_mission}")
            
            # Assurer la connexion DB (éviter les NoneType sur db)
            try:
                local_db = get_db()
                if local_db is None:
                    from src.utils.database import ensure_connection
                    print("controle_intangibilite: get_db() a renvoyé None, tentative de reconnexion...")
                    ensure_connection()
                    local_db = get_db()
                if local_db is None:
                    raise RuntimeError("controle_intangibilite: Base de données non connectée (get_db a renvoyé None)")
            except Exception as e:
                print(f"controle_intangibilite: impossible d'obtenir la connexion DB: {e}")
                raise
            
            # Essayer de trouver la mission avec ObjectId d'abord
            mission = None
            try:
                mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
                print(f"Recherche avec ObjectId: Mission trouvée = {mission is not None}")
            except Exception as e:
                print(f"Erreur lors de la conversion en ObjectId: {e}")
                # Essayer avec l'ID tel quel (string)
                try:
                    mission = local_db.Mission1.find_one({"_id": id_mission})
                    print(f"Recherche avec ID string: Mission trouvée = {mission is not None}")
                except Exception as e2:
                    print(f"Erreur lors de la recherche avec ID string: {e2}")
            
            if not mission:
                print(f"[CONTROLE_INTANGIBILITE] ERREUR: Mission non trouvee avec ID: {id_mission}")
                # Aider l'utilisateur en listant quelques missions existantes
                print(f"Recherche de missions existantes pour aider au debug...")
                try:
                    sample_missions = list(local_db.Mission1.find({}).limit(5))
                    if sample_missions:
                        print(f"Exemples de missions existantes:")
                        for m in sample_missions:
                            print(f"   - ID: {m['_id']}, id_client: {m.get('id_client', 'N/A')}, année: {m.get('annee_auditee', 'N/A')}")
                    else:
                        print(f"Aucune mission trouvée dans la base de données")
                except Exception as e:
                    print(f"Erreur lors de la recherche d'exemples: {e}")
                
                return {
                    "ok": False, 
                    "message": f"Mission non trouvée avec l'ID: {id_mission}. Vérifiez que l'ID est correct et que la mission existe dans la base de données.",
                    "comptes": []
                }
                
            bal_ids = mission.get("balances", [])
            if len(bal_ids) < 2:
                print(f"[CONTROLE_INTANGIBILITE] ERREUR: Pas assez de balances ({len(bal_ids)})")
                return {"ok": False, "message": "Il faut au moins N et N-1.", "comptes": []}
            
            print(f"[CONTROLE_INTANGIBILITE] Nombre de balances: {len(bal_ids)}")
            print(f"[CONTROLE_INTANGIBILITE] Balance IDs: {bal_ids}")

            # Récupérer l'année auditée pour déterminer les périodes
            annee_auditee = int(mission.get("annee_auditee", 0)) if mission.get("annee_auditee") else 0
            
            # Charger les balances avec leurs métadonnées
            balance_docs = []
            for bal_id in bal_ids:
                bal_doc = local_db.Balance.find_one({"_id": ObjectId(bal_id)})
                if bal_doc:
                    # Utiliser annee_balance au lieu de periode
                    annee_balance = bal_doc.get("annee_balance") or bal_doc.get("periode") or ""
                    balance_docs.append({
                        "id": bal_id,
                        "annee": annee_balance,
                        "data": bal_doc.get("balance", [])
                    })
            
            # Si aucune année n'est trouvée, utiliser l'ordre des balances et calculer les années
            if all(not bd.get("annee") for bd in balance_docs) and annee_auditee:
                # Calculer les années basées sur annee_auditee
                for idx, bd in enumerate(balance_docs):
                    bd["annee"] = annee_auditee - idx
            
            # Trier les balances par année (du plus ancien au plus récent)
            # en gérant les années stockées en string ou formats mixtes.
            def _year_sort_key(v):
                y = v.get("annee", 0)
                try:
                    return int(str(y).strip())
                except Exception:
                    return -10**9

            balance_docs.sort(key=_year_sort_key)
            
            # Identifier N et N-1
            if len(balance_docs) < 2:
                return {"ok": False, "message": "Il faut au moins N et N-1.", "periodes": {}, "comptes": []}
            
            bal_N = balance_docs[-1]["data"]  # Dernière année = N
            bal_N1 = balance_docs[-2]["data"]  # Avant-dernière = N-1
            periode_N = str(balance_docs[-1]["annee"]) if balance_docs[-1]["annee"] else "N"
            periode_N1 = str(balance_docs[-2]["annee"]) if balance_docs[-2]["annee"] else "N-1"
            
            # Vérifier que les balances contiennent des données
            if not bal_N or len(bal_N) == 0:
                return {"ok": False, "message": f"La balance N ({periode_N}) ne contient aucune donnée.", "periodes": {"N": periode_N, "N-1": periode_N1}, "comptes": []}
            if not bal_N1 or len(bal_N1) == 0:
                return {"ok": False, "message": f"La balance N-1 ({periode_N1}) ne contient aucune donnée.", "periodes": {"N": periode_N, "N-1": periode_N1}, "comptes": []}
            
            print(f"Lignes brutes dans balance N: {len(bal_N)}")
            print(f"Lignes brutes dans balance N-1: {len(bal_N1)}")
            
            # Vérifier les premières lignes pour debug
            if len(bal_N) > 0:
                print(f"Exemple ligne N (première): {list(bal_N[0].keys()) if bal_N[0] else 'vide'}")
                if bal_N[0] and 'numero_compte' in bal_N[0]:
                    num_compte = bal_N[0].get('numero_compte')
                    print(f"Premier numero_compte N: '{num_compte}' (type: {type(num_compte)}, value: {repr(num_compte)})")
                else:
                    print(f"Première ligne N ne contient pas 'numero_compte': {bal_N[0].keys() if bal_N[0] else 'ligne vide'}")
            
            if len(bal_N1) > 0:
                if bal_N1[0] and 'numero_compte' in bal_N1[0]:
                    num_compte = bal_N1[0].get('numero_compte')
                    print(f"Premier numero_compte N-1: '{num_compte}' (type: {type(num_compte)}, value: {repr(num_compte)})")
            
            print(f"[CONTROLE_INTANGIBILITE] Indexation des comptes...")
            print(f"[CONTROLE_INTANGIBILITE] Avant indexation: bal_N contient {len(bal_N)} lignes, bal_N1 contient {len(bal_N1)} lignes")
            
            idxN = self._index_by_compte(bal_N)
            idxN1 = self._index_by_compte(bal_N1)
            
            print(f"[CONTROLE_INTANGIBILITE] APRES indexation:")
            print(f"[CONTROLE_INTANGIBILITE]    - Comptes indexes dans N: {len(idxN)}")
            print(f"[CONTROLE_INTANGIBILITE]    - Comptes indexes dans N-1: {len(idxN1)}")
            print(f"[CONTROLE_INTANGIBILITE] idxN type: {type(idxN)}, idxN1 type: {type(idxN1)}")

            # Limiter aux classes 1 à 5 uniquement
            def _is_class_1_to_5(account_number):
                try:
                    s = str(account_number).strip()
                    return len(s) > 0 and s[0] in ("1", "2", "3", "4", "5")
                except Exception:
                    return False

            idxN = {k: v for k, v in idxN.items() if _is_class_1_to_5(k)}
            idxN1 = {k: v for k, v in idxN1.items() if _is_class_1_to_5(k)}

            print(f"[CONTROLE_INTANGIBILITE] Filtrage classes 1-5:")
            print(f"[CONTROLE_INTANGIBILITE]    - Comptes retenus en N: {len(idxN)}")
            print(f"[CONTROLE_INTANGIBILITE]    - Comptes retenus en N-1: {len(idxN1)}")
            
            # DIAGNOSTIC: Vérifier si l'indexation a échoué inopinément
            if len(bal_N) > 0 and len(idxN) == 0:
                print(f"PROBLÈME: bal_N contient {len(bal_N)} lignes mais idxN est vide!")
                print(f"   - Type de bal_N[0]: {type(bal_N[0])}")
                if len(bal_N) > 0 and isinstance(bal_N[0], dict):
                    print(f"   - Clés de la première ligne: {list(bal_N[0].keys())}")
                    print(f"   - numero_compte de la première ligne: {bal_N[0].get('numero_compte')}")
            
            if len(bal_N1) > 0 and len(idxN1) == 0:
                print(f"PROBLÈME: bal_N1 contient {len(bal_N1)} lignes mais idxN1 est vide!")
                print(f"   - Type de bal_N1[0]: {type(bal_N1[0])}")
                if len(bal_N1) > 0 and isinstance(bal_N1[0], dict):
                    print(f"   - Clés de la première ligne: {list(bal_N1[0].keys())}")
                    print(f"   - numero_compte de la première ligne: {bal_N1[0].get('numero_compte')}")
            
            if len(idxN) > 0:
                print(f"Exemple compte indexé N: {list(idxN.keys())[:3]}")
            if len(idxN1) > 0:
                print(f"Exemple compte indexé N-1: {list(idxN1.keys())[:3]}")
            
            # Vérifier que les index contiennent des comptes
            # MODIFICATION: Ne pas retourner si au moins une balance a des comptes
            # On peut faire le contrôle même si une seule balance a des comptes
            if len(idxN) == 0 and len(idxN1) == 0:
                print("Aucun compte trouvé dans les balances")
                print(f"   - Balance N: {len(bal_N)} lignes brutes, mais 0 comptes indexés")
                print(f"   - Balance N-1: {len(bal_N1)} lignes brutes, mais 0 comptes indexés")
                print(f"   - Vérifiez que les lignes contiennent bien un champ 'numero_compte' avec une valeur non vide")
                print(f"   - DEBUG: Type de bal_N[0] si existe: {type(bal_N[0]) if bal_N and len(bal_N) > 0 else 'N/A'}")
                print(f"   - DEBUG: Type de bal_N1[0] si existe: {type(bal_N1[0]) if bal_N1 and len(bal_N1) > 0 else 'N/A'}")
                
                # Analyser pourquoi aucun compte n'a été trouvé
                if len(bal_N) > 0:
                    premiere_ligne = bal_N[0]
                    print(f"   - Exemple ligne N: {premiere_ligne}")
                    if 'numero_compte' not in premiere_ligne:
                        print(f"   - Le champ 'numero_compte' est absent de la première ligne")
                    elif premiere_ligne.get('numero_compte') is None:
                        print(f"   - Le champ 'numero_compte' est None")
                    elif str(premiere_ligne.get('numero_compte')).strip() == "":
                        print(f"   - Le champ 'numero_compte' est une chaîne vide")
                
                return {
                    "ok": False, 
                    "message": f"Aucun compte trouvé dans les balances N et N-1. Les balances contiennent {len(bal_N)} et {len(bal_N1)} lignes brutes, mais aucun numéro de compte valide n'a été détecté. Vérifiez que les balances contiennent bien un champ 'numero_compte' avec des valeurs non vides.", 
                    "periodes": {"N": periode_N, "N-1": periode_N1}, 
                    "comptes": []
                }

            # Créer une liste de tous les comptes (sans filtrage sur les classes)
            tous_comptes = []

            def is_compte_130(numero_compte):
                # Autoriser uniquement le compte 130 (résultat en instance d'affectation),
                # pas les sous-comptes type 1309.
                digits = ''.join(ch for ch in str(numero_compte or "") if ch.isdigit())
                normalized = digits.lstrip('0') or '0'
                return normalized == "130"

            def _to_float(v):
                try:
                    if v is None or v == "":
                        return 0.0
                    return float(v)
                except Exception:
                    return 0.0

            def _get_opening_solde(row):
                # Ouverture N: priorité au couple débit/crédit initial.
                di = _to_float(row.get("debit_initial", 0))
                ci = _to_float(row.get("credit_initial", 0))
                if di != 0 or ci != 0:
                    return di - ci

                # Fallbacks selon formats de balance importés.
                if "solde_n1" in row and row.get("solde_n1") not in (None, ""):
                    return _to_float(row.get("solde_n1", 0))
                if "solde_reel" in row and row.get("solde_reel") not in (None, ""):
                    return _to_float(row.get("solde_reel", 0))
                return 0.0

            def _get_closing_solde(row):
                # Clôture N-1: priorité au couple débit/crédit fin.
                df = _to_float(row.get("debit_fin", 0))
                cf = _to_float(row.get("credit_fin", 0))
                if df != 0 or cf != 0:
                    return df - cf

                # Fallbacks selon formats de balance importés.
                if "solde_n" in row and row.get("solde_n") not in (None, ""):
                    return _to_float(row.get("solde_n", 0))
                if "solde_reel" in row and row.get("solde_reel") not in (None, ""):
                    return _to_float(row.get("solde_reel", 0))
                return 0.0
            
            print(f"Début du traitement: {len(idxN)} comptes en N, {len(idxN1)} comptes en N-1")
            
            if len(idxN) == 0:
                print(f"ATTENTION: idxN est vide même si bal_N contient {len(bal_N)} lignes")
            
            if len(idxN1) == 0:
                print(f"ATTENTION: idxN1 est vide même si bal_N1 contient {len(bal_N1)} lignes")
            
            comptes_ajoutes = 0
            comptes_erreur = 0
            
            # 1. Traiter tous les comptes présents en N
            for num, ln in idxN.items():
                try:
                    # Pour l'ouverture N, utiliser les colonnes d'ouverture avec fallback robuste
                    ouvN = _get_opening_solde(ln)

                    prev = idxN1.get(num)
                    if prev:
                        # Pour N-1, utiliser le solde de clôture avec fallback robuste
                        clotN1 = _get_closing_solde(prev)
                        # Convention métier: Écart = Ouverture N - Clôture N-1
                        ecart = ouvN - clotN1
                        
                        # Afficher les 5 premiers comptes pour debug
                        if len(tous_comptes) < 5:
                            print(f"Compte {num}: ouvN={ouvN}, clotN1={clotN1}, ecart={ecart} (ouvN - clotN1)")
                        
                        # Ajouter tous les comptes, pas seulement ceux avec des écarts
                        ecart_normal_130 = (ecart != 0 and is_compte_130(num))
                        tous_comptes.append({
                            "numero_compte": num,
                            "libelle": ln.get("libelle", ""),
                            "ouverture_n": ouvN,
                            "cloture_n1": clotN1,
                            "ecart": ecart,
                            "status": "ok_130" if ecart_normal_130 else ("ecart" if ecart != 0 else "ok"),
                            "message": (
                                f"Écart normal sur le compte 130 : Ouverture N {ouvN} vs Clôture N-1 {clotN1}"
                                if ecart_normal_130 else
                                (f"Ouverture N {ouvN} Clôture N-1 {clotN1}" if ecart != 0 else f"Ouverture N {ouvN} = Clôture N-1 {clotN1}")
                            ),
                            "justification": (
                                f"Le compte 130 (résultat en instance d'affectation) peut varier entre la clôture N-1 ({clotN1}) et l'ouverture N ({ouvN}). Écart normal: {ecart}."
                                if ecart_normal_130 else
                                (f"Écart de {ecart} entre l'ouverture de l'exercice N ({ouvN}) et la clôture de l'exercice N-1 ({clotN1})." if ecart != 0 else "Aucun écart détecté.")
                            ),
                            "conclusion_audit": (
                                "Aucune anomalie détectée: variation normale du compte 130 liée à l'affectation du résultat."
                                if ecart_normal_130 else
                                ("Écart significatif détecté - Nécessite une justification et une documentation des causes de cette variation." if ecart != 0 else "Aucune anomalie détectée.")
                            )
                        })
                        comptes_ajoutes += 1
                    else:
                        # Compte nouveau (présent en N, absent en N-1)
                        # Nouvelle convention: Écart = Ouverture N - 0
                        ecart = ouvN
                        nouveau_normal_130 = is_compte_130(num)
                        tous_comptes.append({
                            "numero_compte": num,
                            "libelle": ln.get("libelle", ""),
                            "ouverture_n": ouvN,
                            "cloture_n1": None,
                            "ecart": ecart,
                            "status": "ok_130" if nouveau_normal_130 else "nouveau",
                            "message": (
                                "Variation normale du compte 130 (présent en N, absent en N-1)"
                                if nouveau_normal_130 else
                                "Compte présent en N mais absent en N-1"
                            ),
                            "justification": (
                                f"Le compte 130 (résultat en instance d'affectation) peut légitimement varier entre N-1 et N. Solde d'ouverture N: {ouvN}."
                                if nouveau_normal_130 else
                                f"Le compte {num} est présent dans l'exercice N avec un solde d'ouverture de {ouvN}, mais n'existait pas dans l'exercice N-1. Cela peut indiquer une création de compte, un reclassement ou une erreur de saisie."
                            ),
                            "conclusion_audit": (
                                "Aucune anomalie détectée: variation normale du compte 130."
                                if nouveau_normal_130 else
                                "Compte nouvellement créé ou reclassé - Vérifier la légitimité de cette création et documenter les raisons."
                            )
                        })
                        comptes_ajoutes += 1
                except Exception as e:
                    comptes_erreur += 1
                    print(f"Erreur lors du traitement du compte {num}: {e}")
                    if comptes_erreur <= 3:
                        import traceback
                        traceback.print_exc()
                    continue
            
            print(f"Après traitement des comptes N: {comptes_ajoutes} comptes ajoutés, {comptes_erreur} erreurs")
            
            # 2. Ajouter les comptes présents en N-1 mais absents en N
            for num, ln in idxN1.items():
                try:
                    # Si le compte n'existe pas en N, c'est un écart
                    if num not in idxN:
                        clotN1 = _get_closing_solde(ln)
                        
                        # Convention métier: Écart = Ouverture N (0) - Clôture N-1
                        ecart_suppr = -clotN1

                        supprime_normal_130 = is_compte_130(num)
                        tous_comptes.append({
                            "numero_compte": num,
                            "libelle": ln.get("libelle", ""),
                            "ouverture_n": None,
                            "cloture_n1": clotN1,
                            "ecart": ecart_suppr,  # Écart = 0 - Clôture N-1 (absent en N)
                            "status": "ok_130" if supprime_normal_130 else "supprime",
                            "message": (
                                "Variation normale du compte 130 (présent en N-1, absent en N)"
                                if supprime_normal_130 else
                                "Compte présent en N-1 mais absent en N"
                            ),
                            "justification": (
                                f"Le compte 130 (résultat en instance d'affectation) peut légitimement varier entre N-1 et N. Solde de clôture N-1: {clotN1}."
                                if supprime_normal_130 else
                                f"Le compte {num} était présent dans l'exercice N-1 avec un solde de clôture de {clotN1}, mais n'existe plus dans l'exercice N. Cela peut indiquer une suppression de compte, un reclassement ou une erreur de saisie."
                            ),
                            "conclusion_audit": (
                                "Aucune anomalie détectée: variation normale du compte 130."
                                if supprime_normal_130 else
                                "Compte supprimé ou reclassé - Vérifier la légitimité de cette suppression et documenter les raisons."
                            )
                        })
                        comptes_ajoutes += 1
                except Exception as e:
                    comptes_erreur += 1
                    print(f"Erreur lors du traitement du compte N-1 {num}: {e}")
                    continue
            
            print(f"Après traitement des comptes N-1: {len(tous_comptes)} comptes au total")

            # Trier par numéro de compte
            tous_comptes.sort(key=lambda x: x["numero_compte"])
            
            # Compter les écarts
            ecarts_count = len([c for c in tous_comptes if c["status"] in ["ecart", "nouveau", "supprime"]])
            
            print(f"Résumé final: {len(tous_comptes)} comptes traités")
            print(f"   - Comptes OK: {len([c for c in tous_comptes if c['status'] == 'ok'])}")
            print(f"   - Comptes avec écart: {len([c for c in tous_comptes if c['status'] == 'ecart'])}")
            print(f"   - Comptes nouveaux: {len([c for c in tous_comptes if c['status'] == 'nouveau'])}")
            print(f"   - Comptes supprimés: {len([c for c in tous_comptes if c['status'] == 'supprime'])}")
            
            # Si aucun compte n'a été trouvé, retourner un message d'aide
            if len(tous_comptes) == 0:
                print("Aucun compte trouvé dans les balances")
                return {
                    "ok": False,
                    "message": "Aucun compte trouvé dans les balances N et N-1. Vérifiez que les balances contiennent bien des données avec des numéros de compte valides.",
                    "periodes": {"N": periode_N, "N-1": periode_N1},
                    "total_comptes": 0,
                    "ecarts_count": 0,
                    "comptes": []
                }
            
            report = {
                "ok": ecarts_count == 0, 
                "total_comptes": len(tous_comptes),
                "ecarts_count": ecarts_count,
                "periodes": {
                    "N": periode_N,
                    "N-1": periode_N1
                },
                "comptes": tous_comptes
            }
            
            print(f"Rapport généré: {len(tous_comptes)} comptes, {ecarts_count} écarts")
            print(f"Clés du rapport: {list(report.keys())}")
            print(f"Nombre de comptes dans le rapport: {len(report.get('comptes', []))}")
            
            # CRITIQUE: Vérifier que tous_comptes n'est pas vide avant de sauvegarder
            if len(tous_comptes) == 0:
                print(f"ERREUR CRITIQUE: tous_comptes est vide avant sauvegarde!")
                print(f"   - idxN avait {len(idxN)} comptes")
                print(f"   - idxN1 avait {len(idxN1)} comptes")
                print(f"   - Vérifiez les logs ci-dessus pour comprendre pourquoi tous_comptes est vide")
                # Ne pas sauvegarder un rapport vide, retourner une erreur explicite
                return {
                    "ok": False,
                    "message": f"Erreur lors du traitement des comptes: {len(idxN)} comptes indexés en N et {len(idxN1)} en N-1, mais aucun compte n'a été ajouté à la liste finale. Vérifiez les logs du serveur.",
                    "periodes": {"N": periode_N, "N-1": periode_N1},
                    "total_comptes": 0,
                    "ecarts_count": 0,
                    "comptes": []
                }
            
            local_db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"controle_intangibilite": report}})
            print(f"[CONTROLE_INTANGIBILITE] Rapport sauvegarde dans la base de donnees")
            print(f"[CONTROLE_INTANGIBILITE] Rapport retourne: total_comptes={report.get('total_comptes')}, ecarts={report.get('ecarts_count')}")
            print(f"[CONTROLE_INTANGIBILITE] ========== FIN (SUCCES) ==========")
            return report
            
        except Exception as e:
            print(f"[CONTROLE_INTANGIBILITE] ERREUR EXCEPTION: {str(e)}")
            import traceback
            traceback.print_exc()
            print(f"[CONTROLE_INTANGIBILITE] ========== FIN (ERREUR) ==========")
            return {"ok": False, "message": f"Erreur lors du contrôle d'intangibilité: {str(e)}", "comptes": []}

    def ajouter_reclassement_intangibilite(self, id_mission, reclassement_data):
        """
        Ajoute un reclassement dans le contrôle d'intangibilité
        
        Args:
            id_mission: ID de la mission
            reclassement_data: {
                "compte_source": "411001",
                "compte_destination": "411002", 
                "montant": 1000.0,
                "sens": "debit" ou "credit",
                "justification": "Reclassement suite à...",
                "date_reclassement": "2024-01-15",
                "auditeur": "Nom de l'auditeur"
            }
        """
        try:
            print(f"[RECLASSEMENT] ========== DEBUT ==========")
            print(f"[RECLASSEMENT] Mission ID: {id_mission}")
            print(f"[RECLASSEMENT] Données: {reclassement_data}")
            
            # Assurer la connexion DB
            try:
                local_db = get_db()
                if local_db is None:
                    from src.utils.database import ensure_connection
                    ensure_connection()
                    local_db = get_db()
                if local_db is None:
                    raise RuntimeError("ajouter_reclassement_intangibilite: Base de données non connectée")
            except Exception as e:
                print(f"ajouter_reclassement_intangibilite: impossible d'obtenir la connexion DB: {e}")
                raise
            
            # Vérifier que la mission existe
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée"}
            
            # Valider les données du reclassement
            required_fields = ["compte_source", "compte_destination", "montant", "sens", "justification"]
            for field in required_fields:
                if field not in reclassement_data or not reclassement_data[field]:
                    return {"ok": False, "message": f"Le champ '{field}' est requis"}
            
            # Valider le montant
            try:
                montant = float(reclassement_data["montant"])
                if montant <= 0:
                    return {"ok": False, "message": "Le montant doit être positif"}
            except (ValueError, TypeError):
                return {"ok": False, "message": "Le montant doit être un nombre valide"}
            
            # Valider le sens
            if reclassement_data["sens"] not in ["debit", "credit"]:
                return {"ok": False, "message": "Le sens doit être 'debit' ou 'credit'"}
            
            # Créer l'objet reclassement avec un ID unique
            import datetime
            reclassement = {
                "id": str(ObjectId()),
                "compte_source": str(reclassement_data["compte_source"]).strip(),
                "compte_destination": str(reclassement_data["compte_destination"]).strip(),
                "montant": montant,
                "sens": reclassement_data["sens"],
                "justification": reclassement_data["justification"],
                "date_reclassement": reclassement_data.get("date_reclassement", datetime.datetime.now().strftime("%Y-%m-%d")),
                "auditeur": reclassement_data.get("auditeur", "Système"),
                "date_creation": datetime.datetime.now().isoformat(),
                "statut": "actif"
            }
            
            # Récupérer les reclassements existants
            reclassements_existants = mission.get("reclassements_intangibilite", [])
            reclassements_existants.append(reclassement)
            
            # Sauvegarder dans la mission
            local_db.Mission1.update_one(
                {"_id": ObjectId(id_mission)}, 
                {"$set": {"reclassements_intangibilite": reclassements_existants}}
            )
            
            print(f"[RECLASSEMENT] Reclassement ajouté avec succès: {reclassement['id']}")
            print(f"[RECLASSEMENT] ========== FIN (SUCCES) ==========")
            
            return {
                "ok": True, 
                "message": "Reclassement ajouté avec succès",
                "reclassement_id": reclassement["id"],
                "reclassement": reclassement
            }
            
        except Exception as e:
            print(f"[RECLASSEMENT] ERREUR: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"ok": False, "message": f"Erreur lors de l'ajout du reclassement: {str(e)}"}

    def supprimer_reclassement_intangibilite(self, id_mission, reclassement_id):
        """
        Supprime un reclassement du contrôle d'intangibilité
        """
        try:
            print(f"[RECLASSEMENT] Suppression du reclassement {reclassement_id} pour mission {id_mission}")
            
            # Assurer la connexion DB
            try:
                local_db = get_db()
                if local_db is None:
                    from src.utils.database import ensure_connection
                    ensure_connection()
                    local_db = get_db()
                if local_db is None:
                    raise RuntimeError("supprimer_reclassement_intangibilite: Base de données non connectée")
            except Exception as e:
                print(f"supprimer_reclassement_intangibilite: impossible d'obtenir la connexion DB: {e}")
                raise
            
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée"}
            
            reclassements = mission.get("reclassements_intangibilite", [])
            reclassements_filtres = [r for r in reclassements if r.get("id") != reclassement_id]
            
            if len(reclassements_filtres) == len(reclassements):
                return {"ok": False, "message": "Reclassement non trouvé"}
            
            # Sauvegarder
            local_db.Mission1.update_one(
                {"_id": ObjectId(id_mission)}, 
                {"$set": {"reclassements_intangibilite": reclassements_filtres}}
            )
            
            return {"ok": True, "message": "Reclassement supprimé avec succès"}
            
        except Exception as e:
            print(f"[RECLASSEMENT] ERREUR lors de la suppression: {str(e)}")
            return {"ok": False, "message": f"Erreur lors de la suppression: {str(e)}"}

    def lister_reclassements_intangibilite(self, id_mission):
        """
        Liste tous les reclassements d'une mission
        """
        try:
            # Assurer la connexion DB
            try:
                local_db = get_db()
                if local_db is None:
                    from src.utils.database import ensure_connection
                    ensure_connection()
                    local_db = get_db()
                if local_db is None:
                    raise RuntimeError("lister_reclassements_intangibilite: Base de données non connectée")
            except Exception as e:
                print(f"lister_reclassements_intangibilite: impossible d'obtenir la connexion DB: {e}")
                raise
            
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "reclassements": []}
            
            reclassements = mission.get("reclassements_intangibilite", [])
            reclassements_actifs = [r for r in reclassements if r.get("statut", "actif") == "actif"]
            
            return {
                "ok": True, 
                "message": f"{len(reclassements_actifs)} reclassement(s) trouvé(s)",
                "reclassements": reclassements_actifs
            }
            
        except Exception as e:
            print(f"[RECLASSEMENT] ERREUR lors du listage: {str(e)}")
            return {"ok": False, "message": f"Erreur: {str(e)}", "reclassements": []}

    def controle_intangibilite_avec_reclassements(self, id_mission):
        """
        Version du contrôle d'intangibilité qui prend en compte les reclassements
        """
        try:
            print(f"[CONTROLE_AVEC_RECLASSEMENTS] ========== DEBUT ==========")
            
            # D'abord, exécuter le contrôle standard
            rapport_standard = self.controle_intangibilite(id_mission)
            if not rapport_standard.get("ok", False):
                return rapport_standard
            
            # Récupérer les reclassements
            reclassements_result = self.lister_reclassements_intangibilite(id_mission)
            if not reclassements_result.get("ok", False):
                # Pas de reclassements, retourner le rapport standard
                rapport_standard["reclassements_appliques"] = []
                return rapport_standard
            
            reclassements = reclassements_result.get("reclassements", [])
            if not reclassements:
                rapport_standard["reclassements_appliques"] = []
                return rapport_standard
            
            print(f"[CONTROLE_AVEC_RECLASSEMENTS] {len(reclassements)} reclassement(s) à appliquer")
            
            # Appliquer les reclassements aux comptes
            comptes_ajustes = []
            reclassements_appliques = []
            
            for compte in rapport_standard.get("comptes", []):
                compte_ajuste = compte.copy()
                numero_compte = compte["numero_compte"]
                digits = ''.join(ch for ch in str(numero_compte or "") if ch.isdigit())
                normalized = digits.lstrip('0') or '0'
                is_compte_130 = normalized == "130"
                
                # Chercher les reclassements qui affectent ce compte
                for reclassement in reclassements:
                    if (reclassement["compte_source"] == numero_compte or 
                        reclassement["compte_destination"] == numero_compte):
                        
                        # Calculer l'ajustement
                        montant = reclassement["montant"]
                        sens = reclassement["sens"]
                        
                        if reclassement["compte_source"] == numero_compte:
                            # Ce compte est la source du reclassement (on retire)
                            ajustement = -montant if sens == "debit" else montant
                        else:
                            # Ce compte est la destination du reclassement (on ajoute)
                            ajustement = montant if sens == "debit" else -montant
                        
                        # Appliquer l'ajustement à l'ouverture N
                        if compte_ajuste["ouverture_n"] is not None:
                            compte_ajuste["ouverture_n"] += ajustement
                        
                        # Recalculer l'écart
                        if (compte_ajuste["ouverture_n"] is not None and 
                            compte_ajuste["cloture_n1"] is not None):
                            # Conserver la même convention que le contrôle standard:
                            # écart = ouverture N - clôture N-1
                            compte_ajuste["ecart"] = compte_ajuste["ouverture_n"] - compte_ajuste["cloture_n1"]
                        
                        # Mettre à jour le statut
                        if is_compte_130 and abs(compte_ajuste["ecart"]) >= 0.01:
                            compte_ajuste["status"] = "ok_130"
                            compte_ajuste["message"] = "Écart normal maintenu sur le compte 130"
                            compte_ajuste["conclusion_audit"] = "Aucune anomalie détectée: variation normale du compte 130."
                        elif abs(compte_ajuste["ecart"]) < 0.01:  # Tolérance pour les erreurs d'arrondi
                            compte_ajuste["status"] = "ok_reclasse"
                            compte_ajuste["message"] = f"Écart résolu par reclassement"
                            compte_ajuste["conclusion_audit"] = "Écart justifié par reclassement documenté."
                        else:
                            compte_ajuste["status"] = "ecart_partiel"
                            compte_ajuste["message"] = f"Écart partiellement résolu par reclassement"
                        
                        # Ajouter l'info du reclassement
                        if "reclassements" not in compte_ajuste:
                            compte_ajuste["reclassements"] = []
                        compte_ajuste["reclassements"].append({
                            "id": reclassement["id"],
                            "type": "source" if reclassement["compte_source"] == numero_compte else "destination",
                            "montant": montant,
                            "sens": sens,
                            "ajustement": ajustement,
                            "justification": reclassement["justification"]
                        })
                        
                        reclassements_appliques.append(reclassement["id"])
                
                comptes_ajustes.append(compte_ajuste)
            
            # Recalculer les statistiques
            ecarts_count = len([c for c in comptes_ajustes if c["status"] in ["ecart", "nouveau", "supprime", "ecart_partiel"]])
            
            rapport_final = rapport_standard.copy()
            rapport_final.update({
                "comptes": comptes_ajustes,
                "ecarts_count": ecarts_count,
                "ok": ecarts_count == 0,
                "reclassements_appliques": list(set(reclassements_appliques)),
                "nombre_reclassements": len(reclassements)
            })
            
            print(f"[CONTROLE_AVEC_RECLASSEMENTS] Rapport final: {len(comptes_ajustes)} comptes, {ecarts_count} écarts restants")
            print(f"[CONTROLE_AVEC_RECLASSEMENTS] ========== FIN (SUCCES) ==========")
            
            return rapport_final
            
        except Exception as e:
            print(f"[CONTROLE_AVEC_RECLASSEMENTS] ERREUR: {str(e)}")
            import traceback
            traceback.print_exc()
            return {"ok": False, "message": f"Erreur lors du contrôle avec reclassements: {str(e)}", "comptes": []}

    def classement_bilan(self, id_mission):
        try:
            #1. Charger la mission
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "classement": []}

            balances_ids = mission.get("balances", [])
            if len(balances_ids) < 2:
                return {"ok": False, "message": "Balances manquantes", "classement": []}

            #2. Charger balances N et N-1
            balance_n = db.Balance.find_one(
                {"_id": ObjectId(balances_ids[0])},
                {"balance": 1}
            )
            balance_n1 = db.Balance.find_one(
                {"_id": ObjectId(balances_ids[1])},
                {"balance": 1}
            )

            if not balance_n or not balance_n1:
                return {"ok": False, "message": "Balances introuvables", "classement": []}

            balance_n_data = balance_n.get("balance", [])
            balance_n1_data = balance_n1.get("balance", [])

            #3. Index balance N-1
            balance_n1_index = {
                str(item.get("numero_compte")): item
                for item in balance_n1_data
            }

            #4. Détection des préfixes depuis les balances
            prefixes = set()
            for item in balance_n_data:
                numero = str(item.get("numero_compte", "")).strip()
                if numero.isdigit() and len(numero) >= 2:
                    prefixes.add(numero[:2])

            if not prefixes:
                return {"ok": False, "message": "Aucun compte valide trouvé", "classement": []}

            #5. Création dynamique des groupes
            table_grouping = []
            for prefixe in sorted(prefixes):
                table_grouping.append({
                    "compte": prefixe,
                    "nature": "bilan" if prefixe[0] in "12345" else "pnl",
                    "libelle": f"AUTRES - COMPTE {prefixe}"
                })

            #6. Calcul du classement
            classement = []

            for group in table_grouping:
                prefixe = group["compte"]

                solde_n = 0
                solde_n1 = 0
                comptes_detaille = []

                for item_n in balance_n_data:
                    numero = str(item_n.get("numero_compte", ""))
                    if numero.startswith(prefixe):
                        item_n1 = balance_n1_index.get(numero, {})

                        solde_item_n = (item_n.get("debit_fin", 0) or 0) - (item_n.get("credit_fin", 0) or 0)
                        solde_item_n1 = (item_n1.get("debit_fin", 0) or 0) - (item_n1.get("credit_fin", 0) or 0)

                        solde_n += solde_item_n
                        solde_n1 += solde_item_n1

                        comptes_detaille.append({
                            "numero_compte": numero,
                            "libelle": item_n.get("libelle", ""),
                            "solde_n": solde_item_n,
                            "solde_n1": solde_item_n1,
                            "variation": solde_item_n - solde_item_n1
                        })

                variation = solde_n - solde_n1
                variation_percent = 0
                if solde_n1 != 0:
                    variation_percent = (variation / abs(solde_n1)) * 100

                classement.append({
                    "compte": prefixe,
                    "libelle": group["libelle"],
                    "nature": group["nature"],
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": variation,
                    "variation_percent": variation_percent,
                    "comptes_detaille": comptes_detaille
                })

            #7. Sauvegarde propre en base
            report = {
                "ok": True,
                "message": f"Classement généré ({len(classement)} groupes)",
                "classement": classement
            }

            db.Mission1.update_one(
                {"_id": ObjectId(id_mission)},
                {"$set": {"classement_bilan": report}}
            )

            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "classement": []}


    def etats_financiers_preliminaires(self, id_mission):
        try:
            local_db = get_db()
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "efi": {}}

            balances = mission.get("balances", [])
            if len(balances) < 2:
                return {"ok": False, "message": "Balances manquantes", "efi": {}}

            balance_n = local_db.Balance.find_one({"_id": ObjectId(balances[0])}, {"balance": 1})
            balance_n1 = local_db.Balance.find_one({"_id": ObjectId(balances[1])}, {"balance": 1})

            if not balance_n or not balance_n1:
                return {"ok": False, "message": "Balances introuvables", "efi": {}}

            # Charger le fichier de mapping EFI
            mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")
            with open(mapping_path, 'r', encoding='utf-8') as file:
                mapping_data = json.load(file)
            
            mapping = mapping_data.get("structure", [])

            # Récupérer les données de balance
            balance_n_data = balance_n.get("balance", [])
            balance_n1_data = balance_n1.get("balance", [])

            # Garantir une balance de variation exploitable pour l'alignement EFI/grouping
            balance_variation = mission.get("balance_variation", []) or []
            if not balance_variation:
                balance_variation = self.rapprochement_des_balances(balance_n_data, balance_n1_data)

            # Générer les états financiers en utilisant la méthode existante prod_efi
            efi_data = self.prod_efi(balance_n_data, balance_n1_data, balance_variation)

            # prod_efi retourne déjà un dictionnaire organisé par nature (actif, passif, pnl)
            # Utiliser directement les données organisées
            efi_organized = {
                "actif": efi_data.get("actif", []),
                "passif": efi_data.get("passif", []),
                "pnl": efi_data.get("pnl", [])
            }
            
            # Log pour vérifier le nombre de lignes générées
            print(f"États financiers générés - Actif: {len(efi_organized.get('actif', []))} lignes, Passif: {len(efi_organized.get('passif', []))} lignes, PNL: {len(efi_organized.get('pnl', []))} lignes")

            # Sauvegarder les états financiers dans la mission
            report = {
                "ok": True,
                "message": f"États financiers générés avec succès",
                "efi": efi_organized,
                "annee_auditee": mission.get("annee_auditee", [])[0] if mission.get("annee_auditee") else None
            }
            local_db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"etats_financiers_preliminaires": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "efi": {}}

    def analyse_quantitative(self, id_mission):
        try:
            local_db = get_db()
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "analyse": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "analyse": []}

            # Récupérer les données de matérialité
            materiality = mission.get("materiality", [])
            selected_materiality = next((mat for mat in materiality if mat.get('choice')), None)

            if not selected_materiality:
                return {"ok": False, "message": "Aucune matérialité sélectionnée", "analyse": []}

            # Analyser chaque compte du grouping
            analyse_data = []
            for item in grouping:
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                materiality_threshold = selected_materiality.get('materiality', 0)
                
                # Déterminer si le compte est significatif
                is_significant = solde_n >= materiality_threshold
                
                # Calculer le pourcentage par rapport au seuil
                percentage_of_threshold = 0
                if materiality_threshold > 0:
                    percentage_of_threshold = (solde_n / materiality_threshold) * 100

                analyse_data.append({
                    "compte": item.get('compte', ''),
                    "libelle": item.get('libelle', ''),
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": solde_n - solde_n1,
                    "materiality_threshold": materiality_threshold,
                    "is_significant": is_significant,
                    "percentage_of_threshold": percentage_of_threshold,
                    "status": "À tester" if is_significant else "Ne pas tester"
                })

            # Trier par solde décroissant
            analyse_data.sort(key=lambda x: x['solde_n'], reverse=True)

            # Statistiques
            total_accounts = len(analyse_data)
            significant_accounts = len([a for a in analyse_data if a['is_significant']])
            total_significant_amount = sum([a['solde_n'] for a in analyse_data if a['is_significant']])

            report = {
                "ok": True,
                "message": f"Analyse quantitative générée avec succès",
                "analyse": analyse_data,
                "statistics": {
                    "total_accounts": total_accounts,
                    "significant_accounts": significant_accounts,
                    "non_significant_accounts": total_accounts - significant_accounts,
                    "total_significant_amount": total_significant_amount,
                    "materiality_threshold": selected_materiality.get('materiality', 0),
                    "materiality_benchmark": selected_materiality.get('benchmark', '')
                }
            }

            # Sauvegarder l'analyse dans la mission
            local_db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"analyse_quantitative": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "analyse": []}

    def analyse_qualitative(self, id_mission):
        try:
            local_db = get_db()
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "analyse": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "analyse": []}

            # Questions Q1-Q8 pour l'analyse qualitative
            questions = [
                "Volume d'activité, complexité et homogénéité des transactions enregistrées, existence de transactions significatives inhabituelles ou anormales dans le COTABD",
                "Changements identifiés dans le COTABD et détermination si un ou de nouveaux risque(s) sont apparus du fait de changement au sein de l'entité ou de son environnement (économique, légal, réglementaire, normatif ou méthodes comptables)",
                "Sensibilité de l'entité aux anomalies issues de fraudes (Si oui, le risque est obligatoirement Significant)",
                "Niveau de complexité des normes, règles, méthodes comptables, notes annexes, estimations ou jugements liées aux comptes ou aux notes annexes",
                "Exposition du COTABD à des pertes (charges ou dépréciations)",
                "Probabilité que des passifs éventuels significatifs (procès, contentieux, litiges etc) puissent être issus des transactions enregistrées dans le COTABD",
                "Existence de transactions avec des parties liées dans le COTABD",
                "Niveau de contrôle interne et fiabilité des systèmes d'information liés aux comptes"
            ]

            # Récupérer les réponses existantes si disponibles
            qualitative_responses = mission.get("qualitative_responses", {})
            
            # Préparer les données d'analyse
            analyse_data = []
            for item in grouping:
                compte = item.get('compte', '')
                libelle = item.get('libelle', '')
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                
                # Récupérer les réponses pour ce compte (initialiser à False si pas de réponses)
                compte_responses = qualitative_responses.get(compte, {})
                
                # S'assurer que toutes les questions Q1-Q8 sont initialisées
                for q in range(1, len(questions) + 1):
                    if f'Q{q}' not in compte_responses:
                        compte_responses[f'Q{q}'] = False
                
                # Calculer le score qualitatif
                total_questions = len(questions)
                positive_responses = sum(1 for q in range(1, total_questions + 1) if compte_responses.get(f'Q{q}', False))
                qualitative_score = (positive_responses / total_questions) * 100
                
                # Déterminer si le compte est significatif qualitativement (logique à 3 niveaux)
                if qualitative_score >= 50:
                    is_qualitatively_significant = True
                elif qualitative_score >= 25:
                    is_qualitatively_significant = False  # À évaluer mais pas significatif
                else:
                    is_qualitatively_significant = False
                
                # Préparer les réponses détaillées
                responses_detail = []
                for i, question in enumerate(questions, 1):
                    responses_detail.append({
                        "question_id": f"Q{i}",
                        "question_text": question,
                        "response": compte_responses.get(f'Q{i}', False),
                        "is_positive": compte_responses.get(f'Q{i}', False)
                    })

                analyse_data.append({
                    "compte": compte,
                    "libelle": libelle,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": solde_n - solde_n1,
                    "qualitative_score": qualitative_score,
                    "positive_responses": positive_responses,
                    "total_questions": total_questions,
                    "is_qualitatively_significant": is_qualitatively_significant,
                    "responses_detail": responses_detail,
                    "status": "À tester" if is_qualitatively_significant else "Ne pas tester"
                })

            # Trier par score qualitatif décroissant
            analyse_data.sort(key=lambda x: x['qualitative_score'], reverse=True)

            # Statistiques
            total_accounts = len(analyse_data)
            significant_accounts = len([a for a in analyse_data if a['is_qualitatively_significant']])
            total_positive_responses = sum([a['positive_responses'] for a in analyse_data])

            report = {
                "ok": True,
                "message": f"Analyse qualitative générée avec succès",
                "analyse": analyse_data,
                "questions": questions,
                "statistics": {
                    "total_accounts": total_accounts,
                    "significant_accounts": significant_accounts,
                    "non_significant_accounts": total_accounts - significant_accounts,
                    "total_positive_responses": total_positive_responses,
                    "average_score": sum([a['qualitative_score'] for a in analyse_data]) / total_accounts if total_accounts > 0 else 0
                }
            }

            # Sauvegarder l'analyse dans la mission
            local_db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"analyse_qualitative": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "analyse": []}

    def synthese_comptes_significatifs(self, id_mission):
        try:
            local_db = get_db()
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "synthese": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "synthese": []}

            # Récupérer les analyses quantitative et qualitative
            analyse_quantitative = mission.get("analyse_quantitative", {})
            analyse_qualitative = mission.get("analyse_qualitative", {})
            
            # Créer un index des analyses pour un accès rapide
            quant_index = {}
            if analyse_quantitative.get("analyse"):
                for item in analyse_quantitative["analyse"]:
                    quant_index[item["compte"]] = item
            
            qual_index = {}
            if analyse_qualitative.get("analyse"):
                for item in analyse_qualitative["analyse"]:
                    qual_index[item["compte"]] = item

            # Préparer les données de synthèse
            synthese_data = []
            for item in grouping:
                compte = item.get('compte', '')
                libelle = item.get('libelle', '')
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                
                # Récupérer les données quantitatives
                quant_data = quant_index.get(compte, {})
                is_quantitatively_significant = quant_data.get('is_significant', False)
                materiality_threshold = quant_data.get('materiality_threshold', 0)
                percentage_of_threshold = quant_data.get('percentage_of_threshold', 0)
                
                # Récupérer les données qualitatives
                qual_data = qual_index.get(compte, {})
                is_qualitatively_significant = qual_data.get('is_qualitatively_significant', False)
                qualitative_score = qual_data.get('qualitative_score', 0)
                positive_responses = qual_data.get('positive_responses', 0)
                
                # Déterminer le statut final
                final_status = self.determine_final_status(
                    is_quantitatively_significant, 
                    is_qualitatively_significant
                )
                
                # Calculer le niveau de risque
                risk_level = self.calculate_risk_level(
                    is_quantitatively_significant,
                    is_qualitatively_significant,
                    percentage_of_threshold,
                    qualitative_score
                )

                synthese_data.append({
                    "compte": compte,
                    "libelle": libelle,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": solde_n - solde_n1,
                    "is_quantitatively_significant": is_quantitatively_significant,
                    "is_qualitatively_significant": is_qualitatively_significant,
                    "materiality_threshold": materiality_threshold,
                    "percentage_of_threshold": percentage_of_threshold,
                    "qualitative_score": qualitative_score,
                    "positive_responses": positive_responses,
                    "final_status": final_status,
                    "risk_level": risk_level,
                    "recommendation": self.get_recommendation(final_status, risk_level)
                })

            # Trier par niveau de risque décroissant
            risk_order = {"Très élevé": 4, "Élevé": 3, "Modéré": 2, "Faible": 1}
            synthese_data.sort(key=lambda x: risk_order.get(x['risk_level'], 0), reverse=True)

            # Statistiques
            total_accounts = len(synthese_data)
            accounts_to_test = len([a for a in synthese_data if a['final_status'] == "À tester"])
            accounts_not_to_test = len([a for a in synthese_data if a['final_status'] == "Ne pas tester"])
            
            risk_distribution = {}
            for item in synthese_data:
                risk = item['risk_level']
                risk_distribution[risk] = risk_distribution.get(risk, 0) + 1

            report = {
                "ok": True,
                "message": f"Synthèse des comptes significatifs générée avec succès",
                "synthese": synthese_data,
                "statistics": {
                    "total_accounts": total_accounts,
                    "accounts_to_test": accounts_to_test,
                    "accounts_not_to_test": accounts_not_to_test,
                    "risk_distribution": risk_distribution,
                    "materiality_threshold": materiality_threshold,
                    "total_amount_to_test": sum([a['solde_n'] for a in synthese_data if a['final_status'] == "À tester"])
                }
            }

            # Sauvegarder la synthèse dans la mission
            local_db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"synthese_comptes_significatifs": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "synthese": []}

    def determine_final_status(self, is_quantitatively_significant, is_qualitatively_significant):
        """Détermine le statut final d'un compte basé sur les analyses quantitative et qualitative"""
        if is_quantitatively_significant and is_qualitatively_significant:
            return "À tester (Quantitatif + Qualitatif)"
        elif is_quantitatively_significant:
            return "À tester (Quantitatif)"
        elif is_qualitatively_significant:
            return "À tester (Qualitatif)"
        else:
            return "Ne pas tester"

    def calculate_risk_level(self, is_quantitatively_significant, is_qualitatively_significant, percentage_of_threshold, qualitative_score):
        """Calcule le niveau de risque d'un compte"""
        if is_quantitatively_significant and is_qualitatively_significant:
            if percentage_of_threshold >= 200 or qualitative_score >= 75:
                return "Très élevé"
            elif percentage_of_threshold >= 150 or qualitative_score >= 50:
                return "Élevé"
            else:
                return "Modéré"
        elif is_quantitatively_significant:
            if percentage_of_threshold >= 200:
                return "Très élevé"
            elif percentage_of_threshold >= 150:
                return "Élevé"
            else:
                return "Modéré"
        elif is_qualitatively_significant:
            if qualitative_score >= 75:
                return "Élevé"
            elif qualitative_score >= 50:
                return "Modéré"
            else:
                return "Faible"
        else:
            return "Faible"

    def get_recommendation(self, final_status, risk_level):
        """Génère une recommandation basée sur le statut final et le niveau de risque"""
        if final_status.startswith("À tester"):
            if risk_level == "Très élevé":
                return "Tests d'audit approfondis obligatoires"
            elif risk_level == "Élevé":
                return "Tests d'audit substantiels recommandés"
            elif risk_level == "Modéré":
                return "Tests d'audit de base suffisants"
            else:
                return "Tests d'audit minimaux"
        else:
            return "Aucun test d'audit requis"

    def presentation_comptes_significatifs(self, id_mission):
        try:
            local_db = get_db()
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "presentation": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "presentation": []}

            # Récupérer les analyses quantitative et qualitative
            analyse_quantitative = mission.get("analyse_quantitative", {})
            analyse_qualitative = mission.get("analyse_qualitative", {})
            
            # Les analyses quantitative/qualitative ne sont plus obligatoires
            # La presentation peut se baser sur les flags du grouping.

            
            # Créer un index des analyses pour un accès rapide
            quant_index = {}
            if analyse_quantitative.get("analyse"):
                for item in analyse_quantitative["analyse"]:
                    quant_index[item["compte"]] = item
            qual_index = {}
            if analyse_qualitative.get("analyse"):
                for item in analyse_qualitative["analyse"]:
                    qual_index[item["compte"]] = item
            # Préparer les données de présentation
            presentation_data = []
            for item in grouping:
                compte = item.get('compte', '')
                libelle = item.get('libelle', '')
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                variation = solde_n - solde_n1
                
                # Récupérer les données quantitatives
                quant_data = quant_index.get(compte, {})
                is_quantitatively_significant = quant_data.get('is_significant', bool(item.get('materiality', False)))
                materiality_threshold = quant_data.get('materiality_threshold', 0)
                percentage_of_threshold = quant_data.get('percentage_of_threshold', 0)
                
                # Récupérer les données qualitatives
                qual_data = qual_index.get(compte, {})
                is_qualitatively_significant = qual_data.get('is_qualitatively_significant', bool(item.get('significant', False)))
                qualitative_score = qual_data.get('qualitative_score', 100 if is_qualitatively_significant else 0)
                positive_responses = qual_data.get('positive_responses', 0)
                
                # Calculer le pourcentage de variation
                variation_percent = 0
                if solde_n1 != 0:
                    variation_percent = (variation / abs(solde_n1)) * 100
                
                # Déterminer le statut de significativité
                if is_quantitatively_significant and is_qualitatively_significant:
                    significativite_status = "Significatif (Quantitatif + Qualitatif)"
                    priorite = "Haute"
                elif is_quantitatively_significant:
                    significativite_status = "Significatif (Quantitatif)"
                    priorite = "Moyenne"
                elif is_qualitatively_significant:
                    significativite_status = "Significatif (Qualitatif)"
                    priorite = "Moyenne"
                else:
                    significativite_status = "Non significatif"
                    priorite = "Faible"
                
                # Générer une recommandation d'audit
                if significativite_status.startswith("Significatif"):
                    if priorite == "Haute":
                        recommandation_audit = "Tests d'audit approfondis obligatoires"
                    else:
                        recommandation_audit = "Tests d'audit substantiels recommandés"
                else:
                    recommandation_audit = "Tests d'audit minimaux ou aucun test"

                presentation_data.append({
                    "compte": compte,
                    "libelle": libelle,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": variation,
                    "variation_percent": variation_percent,
                    "is_quantitatively_significant": is_quantitatively_significant,
                    "is_qualitatively_significant": is_qualitatively_significant,
                    "materiality_threshold": materiality_threshold,
                    "percentage_of_threshold": percentage_of_threshold,
                    "qualitative_score": qualitative_score,
                    "positive_responses": positive_responses,
                    "significativite_status": significativite_status,
                    "priorite": priorite,
                    "recommandation_audit": recommandation_audit
                })

            # Trier par priorité et montant décroissant
            priorite_order = {"Haute": 3, "Moyenne": 2, "Faible": 1}
            presentation_data.sort(key=lambda x: (priorite_order.get(x['priorite'], 0), x['solde_n']), reverse=True)

            # Statistiques
            total_accounts = len(presentation_data)
            significant_accounts = len([a for a in presentation_data if a['significativite_status'].startswith("Significatif")])
            non_significant_accounts = total_accounts - significant_accounts
            high_priority_accounts = len([a for a in presentation_data if a['priorite'] == "Haute"])
            total_significant_amount = sum([a['solde_n'] for a in presentation_data if a['significativite_status'].startswith("Significatif")])

            report = {
                "ok": True,
                "message": f"Présentation des comptes significatifs générée avec succès",
                "presentation": presentation_data,
                "statistics": {
                    "total_accounts": total_accounts,
                    "significant_accounts": significant_accounts,
                    "non_significant_accounts": non_significant_accounts,
                    "high_priority_accounts": high_priority_accounts,
                    "total_significant_amount": total_significant_amount,
                    "materiality_threshold": materiality_threshold
                }
            }

            # Sauvegarder la présentation dans la mission
            local_db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"presentation_comptes_significatifs": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "presentation": []}

    def revue_analytique_finale(self, id_mission):
        try:
            local_db = get_db()
            mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "revue": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "revue": []}

            # Récupérer les analyses précédentes (optionnelles)
            synthese = mission.get("synthese_comptes_significatifs", {})
            analyse_quantitative = mission.get("analyse_quantitative", {})
            analyse_qualitative = mission.get("analyse_qualitative", {})
            
            # Créer un index des analyses pour un accès rapide
            synthese_index = {}
            if synthese and synthese.get("synthese"):
                for item in synthese["synthese"]:
                    synthese_index[item["compte"]] = item
            
            # Créer un index des analyses quantitatives si disponibles
            quant_index = {}
            if analyse_quantitative and analyse_quantitative.get("analyse"):
                for item in analyse_quantitative["analyse"]:
                    quant_index[item["compte"]] = item
            
            # Créer un index des analyses qualitatives si disponibles
            qual_index = {}
            if analyse_qualitative and analyse_qualitative.get("analyse"):
                for item in analyse_qualitative["analyse"]:
                    qual_index[item["compte"]] = item

            # Liste des tableaux de revue analytique
            list_tableaux = [
                "CR", "BI", "ERE", "MA", "APE", "TSE", "IT", "PE", "ADP", "ACE",
                "RF", "PA", "IC", "AAVI", "IF", "ST", "AC", "TR", "CP", "DF",
                "DCRA", "FCR", "DFS", "AD", "COMMENTAIRE"
            ]

            # Préparer les données de revue analytique
            revue_data = []
            for item in grouping:
                compte = item.get('compte', '')
                libelle = item.get('libelle', '')
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                variation = solde_n - solde_n1
                
                # Récupérer les données de synthèse si disponibles
                synthese_data = synthese_index.get(compte, {})
                final_status = synthese_data.get('final_status', 'Non évalué')
                risk_level = synthese_data.get('risk_level', 'Non évalué')
                recommendation = synthese_data.get('recommendation', 'Non évalué')
                
                # Si pas de synthèse, essayer d'utiliser les analyses individuelles
                if final_status == 'Non évalué':
                    # Vérifier les analyses quantitatives
                    quant_data = quant_index.get(compte, {})
                    is_quantitatively_significant = quant_data.get('is_significant', False)
                    
                    # Vérifier les analyses qualitatives
                    qual_data = qual_index.get(compte, {})
                    is_qualitatively_significant = qual_data.get('is_qualitatively_significant', False)
                    
                    # Déterminer le statut basé sur les analyses disponibles
                    if is_quantitatively_significant and is_qualitatively_significant:
                        final_status = 'À tester (Quantitatif + Qualitatif)'
                        risk_level = 'Élevé'
                        recommendation = 'Tests d\'audit substantiels recommandés'
                    elif is_quantitatively_significant:
                        final_status = 'À tester (Quantitatif)'
                        risk_level = 'Modéré'
                        recommendation = 'Tests d\'audit de base suffisants'
                    elif is_qualitatively_significant:
                        final_status = 'À tester (Qualitatif)'
                        risk_level = 'Modéré'
                        recommendation = 'Tests d\'audit de base suffisants'
                    else:
                        # Aucune analyse disponible, utiliser les données de grouping
                        # Suppression du statut "À évaluer" : considérer comme non significatif
                        final_status = 'Ne pas tester'
                        risk_level = 'Faible'
                        recommendation = 'Aucun test d\'audit requis'
                
                # Calculer le pourcentage de variation
                variation_percent = 0
                if solde_n1 != 0:
                    variation_percent = (variation / abs(solde_n1)) * 100
                
                # Générer un commentaire automatique basé sur l'analyse
                commentaire_auto = self.generate_automatic_comment(
                    compte, solde_n, solde_n1, variation_percent, 
                    final_status, risk_level, list_tableaux
                )
                
                # Déterminer le statut de validation
                validation_status = self.determine_validation_status(
                    final_status, risk_level, variation_percent
                )

                revue_data.append({
                    "compte": compte,
                    "libelle": libelle,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": variation,
                    "variation_percent": variation_percent,
                    "final_status": final_status,
                    "risk_level": risk_level,
                    "recommendation": recommendation,
                    "commentaire_auto": commentaire_auto,
                    "commentaire_perso": "",  # À remplir par l'auditeur
                    "validation_status": validation_status,
                    "is_validated": False  # À valider par l'auditeur
                })

            # Trier par niveau de risque décroissant
            risk_order = {"Très élevé": 5, "Élevé": 4, "Modéré": 3, "Faible": 2, "À déterminer": 1, "Non évalué": 0}
            revue_data.sort(key=lambda x: risk_order.get(x['risk_level'], 0), reverse=True)

            # Statistiques de validation
            total_accounts = len(revue_data)
            accounts_to_validate = len([a for a in revue_data if a['final_status'].startswith("À tester")])
            accounts_validated = len([a for a in revue_data if a['is_validated']])
            accounts_pending = accounts_to_validate - accounts_validated

            # Calculer le pourcentage de validation
            validation_percentage = (accounts_validated / accounts_to_validate * 100) if accounts_to_validate > 0 else 100

            # Déterminer quelles analyses sont disponibles
            analyses_disponibles = []
            if synthese and synthese.get("synthese"):
                analyses_disponibles.append("Synthèse")
            if analyse_quantitative and analyse_quantitative.get("analyse"):
                analyses_disponibles.append("Quantitative")
            if analyse_qualitative and analyse_qualitative.get("analyse"):
                analyses_disponibles.append("Qualitative")
            
            if analyses_disponibles:
                message = f"Revue analytique finale générée avec succès (Analyses disponibles: {', '.join(analyses_disponibles)})"
            else:
                message = "Revue analytique finale générée avec succès (Évaluation basique basée sur les données de grouping)"

            report = {
                "ok": True,
                "message": message,
                "revue": revue_data,
                "analyses_disponibles": analyses_disponibles,
                "statistics": {
                    "total_accounts": total_accounts,
                    "accounts_to_validate": accounts_to_validate,
                    "accounts_validated": accounts_validated,
                    "accounts_pending": accounts_pending,
                    "validation_percentage": validation_percentage,
                    "audit_status": "En cours" if accounts_pending > 0 else "Validé"
                }
            }

            # Sauvegarder la revue analytique dans la mission
            local_db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"revue_analytique_finale": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "revue": []}

    def generate_automatic_comment(self, compte, solde_n, solde_n1, variation_percent, final_status, risk_level, list_tableaux):
        """Génère un commentaire automatique basé sur l'analyse"""
        # Déterminer le type de compte
        compte_type = "Compte général"
        for tableau in list_tableaux:
            if str(compte).startswith(str(tableau)):
                compte_type = f"Compte {tableau}"
                break
        
        # Commentaire basé sur la variation
        if abs(variation_percent) > 50:
            variation_comment = f"Variation significative de {variation_percent:.1f}%"
        elif abs(variation_percent) > 20:
            variation_comment = f"Variation notable de {variation_percent:.1f}%"
        else:
            variation_comment = f"Variation normale de {variation_percent:.1f}%"
        
        # Commentaire basé sur le statut
        if final_status.startswith("À tester"):
            status_comment = f"Compte significatif nécessitant des tests d'audit"
        elif final_status == "Ne pas tester":
            status_comment = f"Compte non significatif"
        else:
            status_comment = f"Compte non significatif"
        
        # Commentaire basé sur le niveau de risque
        risk_comment = f"Niveau de risque: {risk_level}"
        
        return f"{compte_type}. {variation_comment}. {status_comment}. {risk_comment}."

    def determine_validation_status(self, final_status, risk_level, variation_percent):
        """Détermine le statut de validation d'un compte"""
        if final_status.startswith("À tester"):
            if risk_level == "Très élevé":
                return "Validation obligatoire"
            elif risk_level == "Élevé":
                return "Validation recommandée"
            elif abs(variation_percent) > 30:
                return "Validation recommandée"
            else:
                return "Validation optionnelle"
        else:
            return "Validation non requise"


        return materiality



    # ---------- Analyses grouping ----------

    def make_quantitative_analysis(self, id_mission):
        local_db = get_db()
        if local_db is None:
            raise RuntimeError("Base de données non connectée (get_db a retourné None)")

        mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
        if not mission:
            return 0

        grouping = mission.get('grouping', [])
        materialities = mission.get('materiality', [])

        materiality = next((mat for mat in materialities if mat.get('choice')), None)

        if not materiality:
            return 0

        seuil = int(materiality.get('materiality', 0))

        for item in grouping:
            solde = int(item.get('solde_n', 0))
            item['materiality'] = solde >= seuil

        result = local_db.Mission1.update_one(
            {"_id": ObjectId(id_mission)},
            {"$set": {"grouping": grouping}}
        )

        return result.modified_count



    def make_qualitative_analysis(self, id_mission, significant):
        local_db = get_db()
        if local_db is None:
            raise RuntimeError("Base de données non connectée (get_db a retourné None)")

        mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})
        if not mission:
            return 0

        grouping = mission.get('grouping', [])

        for item in grouping:
            value = next(
                (elt.get('significant') for elt in significant if elt.get('compte') == item.get('compte')),
                False
            )
            item['significant'] = bool(value)

        result = local_db.Mission1.update_one(
            {"_id": ObjectId(id_mission)},
            {"$set": {"grouping": grouping}}
        )

        return result.modified_count




    def make_final_sm(self, id_mission):

        db = get_database()
        mission = db["Mission1"].find_one({"_id": ObjectId(id_mission)})
        if not mission:
            return 0, []

        existing_grouping = mission.get('grouping', []) or []
        balance_variation = mission.get('balance_variation', []) or []
        referentiel = mission.get('referentiel', 'syscohada')

        # Recalcule balance_variation depuis les 2 balances sources quand possible
        # pour inclure aussi les comptes présents uniquement en N-1.
        balances_ids = mission.get('balances', []) or []
        if len(balances_ids) >= 2:
            try:
                balance_n_doc = db.Balance.find_one({"_id": ObjectId(balances_ids[0])}, {"balance": 1}) or {}
                balance_n1_doc = db.Balance.find_one({"_id": ObjectId(balances_ids[1])}, {"balance": 1}) or {}
                balance_n_data = balance_n_doc.get("balance", []) or []
                balance_n1_data = balance_n1_doc.get("balance", []) or []
                if balance_n_data or balance_n1_data:
                    balance_variation = self.rapprochement_des_balances(balance_n_data, balance_n1_data)
                    db.Mission1.update_one(
                        {"_id": ObjectId(id_mission)},
                        {"$set": {"balance_variation": balance_variation}}
                    )
            except Exception as e:
                print(f"make_final_sm: échec recalcul balance_variation: {e}")

        # Régénérer le grouping depuis la balance variation pour refléter
        # les dernières règles (ex: compte 419 au passif), puis conserver
        # les champs d'analyse déjà calculés quand ils existent.
        if isinstance(balance_variation, list) and balance_variation:
            regenerated_grouping = self.create_grouping(balance_variation, referentiel)
            old_idx = {
                f"{item.get('libelle', '')}|{item.get('section', '')}": item
                for item in existing_grouping
            }
            for item in regenerated_grouping:
                old_item = old_idx.get(f"{item.get('libelle', '')}|{item.get('section', '')}")
                if old_item:
                    if 'materiality' in old_item:
                        item['materiality'] = old_item.get('materiality')
                    if 'significant' in old_item:
                        item['significant'] = old_item.get('significant')
                    if 'mat_sign' in old_item:
                        item['mat_sign'] = old_item.get('mat_sign')
            grouping = regenerated_grouping
        else:
            grouping = existing_grouping



        for item in grouping:

            mat = item.get('materiality', None)

            sign = item.get('significant', None)



            if (mat is not None) and (sign is not None):

                value = ""

                if mat is False and sign is True:

                    value = "non matériel significatif"

                elif mat is False and sign is False:

                    value = "non matériel non significatif"

                elif mat is True and sign is True:

                    value = "matériel significatif"

                elif mat is True and sign is False:

                    value = "matériel non significatif"

                else:

                    value = None

                item['mat_sign'] = value



        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})

        return result.modified_count, grouping



    # ---------- Récup mission ----------

    def afficher_informations_missions(self, id_client):
        db = get_database()
        _id = db["Mission1"].find_one({"_id": ObjectId(id_client)})

        return _id



    # ---------- Production COMMENTAIRE ----------

    def prod_efi(self, balance_n, balance_n1, balance_variation):
        # Fallback défensif : certains appels historiques envoient [].
        # Sans balance_variation, l'alignement EFI/grouping (notamment comptes mixtes
        # 42/43/44... selon solde débiteur/créditeur) ne peut pas s'appliquer.
        if not isinstance(balance_variation, list) or not balance_variation:
            try:
                balance_variation = self.rapprochement_des_balances(balance_n, balance_n1)
            except Exception:
                balance_variation = []

        mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")

        with open(mapping_path, 'r', encoding='utf-8') as file:

            result = json.load(file)

        mapping = result['structure']



        # nettoyer les champs *_cpt en listes

        for mapp in mapping:

            mapp['brut_cpt'] = mapp['brut_cpt'].split(',') if mapp.get('brut_cpt') is not None else mapp.get('brut_cpt')

            mapp['amor_cpt'] = mapp['amor_cpt'].split(',') if mapp.get('amor_cpt') is not None else mapp.get('amor_cpt')

            mapp['net_cpt'] = mapp['net_cpt'].split(',') if mapp.get('net_cpt') is not None else mapp.get('net_cpt')

            mapp['brut_except_cpt'] = mapp['brut_except_cpt'].split(',') if mapp.get('brut_except_cpt') is not None else mapp.get('brut_except_cpt', [])

            mapp['amor_except_cpt'] = mapp['amor_except_cpt'].split(',') if mapp.get('amor_except_cpt') is not None else mapp.get('amor_except_cpt', [])

            mapp['net_except_cpt'] = mapp['net_except_cpt'].split(',') if mapp.get('net_except_cpt') is not None else mapp.get('net_except_cpt', [])



        datum = {}

        list_efi = ['actif', 'passif', 'pnl']

        for efi in list_efi:

            structure = []

            select_mapping = [elt for elt in mapping if elt['nature'] == efi]
            
            # Log pour vérifier le nombre de lignes dans le mapping
            print(f"Mapping {efi}: {len(select_mapping)} lignes trouvées")

            for idx, data in enumerate(select_mapping):

                row = {}
                
                ref = data.get('ref', '')
                libelle = data.get('libelle', '')
                
                # Log pour debug si ref est vide
                if not ref:
                    print(f"Ligne {idx+1} sans ref dans {efi}: {libelle[:50]}")

                # Vérifier si brut_cpt ou net_cpt existe (au moins un des deux doit être défini)
                brut_cpt = data.get('brut_cpt')
                amor_cpt = data.get('amor_cpt')
                net_cpt = data.get('net_cpt')
                
                # brut_cpt et amor_cpt peuvent être None ou une liste (après split)
                has_brut_amor = brut_cpt and amor_cpt
                if isinstance(brut_cpt, list):
                    has_brut_amor = has_brut_amor and len(brut_cpt) > 0
                if isinstance(amor_cpt, list):
                    has_brut_amor = has_brut_amor and len(amor_cpt) > 0
                
                # net_cpt peut être None ou une liste (après split)
                has_net = net_cpt is not None
                if isinstance(net_cpt, list):
                    has_net = has_net and len(net_cpt) > 0
                
                # TOUJOURS inclure la ligne, même sans comptes
                if not has_brut_amor and not has_net:
                    # Si aucune définition de compte n'existe, on crée quand même la ligne avec des valeurs nulles
                    row['ref'] = data.get('ref') or ''
                    row['libelle'] = data.get('libelle') or ''
                    row['note'] = data.get('note')  # Inclure la note si elle existe
                    row['brut_solde_n'] = 0
                    row['amor_solde_n'] = None
                    row['net_solde_n'] = 0
                    row['net_solde_n1'] = 0
                    row['compte_to_be_used'] = ''
                    row['compte_to_be_used_off'] = []
                    structure.append(row)
                    print(f"Ligne {idx+1} ({ref}) ajoutée sans comptes")
                    continue

                if has_brut_amor:

                    # N

                    brut_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))

                    amor_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))



                    brut_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('brut_except_cpt', [])))

                    amor_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('amor_except_cpt', [])))

                    net_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))



                    data['brut_solde_n'] = brut_solde_n + brut_except_n

                    data['amor_solde_n'] = amor_solde_n + amor_except_n

                    data['net_solde_n'] = data['brut_solde_n'] + data['amor_solde_n'] + net_except_n



                    # N-1

                    brut_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))

                    amor_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))

                    brut_except_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('brut_except_cpt', [])))

                    amor_except_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('amor_except_cpt', [])))

                    net_except_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))

                    data['net_solde_n1'] = (brut_n1 + brut_except_n1) + (amor_n1 + amor_except_n1) + net_except_n1

                else:

                    net_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))

                    net_solde_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))



                    net_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))

                    net_except_n1_bis = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))



                    data['net_solde_n'] = net_solde_n + net_except_n

                    data['net_solde_n1'] = net_solde_n1 + net_except_n1_bis



                row['ref'] = data.get('ref') or ''

                row['libelle'] = data.get('libelle') or ''
                
                row['note'] = data.get('note')  # Inclure la note si elle existe

                row['compte_to_be_used'] = str(data.get('brut_cpt')) + str(data.get('amor_cpt')) + str(data.get('net_cpt')) + str(data.get('brut_except_cpt')) + str(data.get('amor_except_cpt')) + str(data.get('net_except_cpt'))

                row['compte_to_be_used'] = row['compte_to_be_used'].replace('None', '')



                one = data.get('brut_cpt', []) or []

                two = data.get('amor_cpt', []) or []

                three = data.get('net_cpt', []) or []

                four = data.get('brut_except_cpt', []) or []

                five = data.get('amor_except_cpt', []) or []

                six = data.get('net_except_cpt', []) or []



                row['compte_to_be_used_off'] = list(set(one + two + three + four + five + six))



                row['brut_solde_n'] = data.get('brut_solde_n')

                row['amor_solde_n'] = data.get('amor_solde_n')

                row['net_solde_n'] = data.get('net_solde_n')

                row['net_solde_n1'] = data.get('net_solde_n1')



                structure.append(row)
                print(f"Ligne {idx+1} ({ref}) ajoutée avec comptes")

            # Log pour vérifier le nombre de lignes générées
            print(f"Structure {efi} générée: {len(structure)} lignes sur {len(select_mapping)} attendues")
            
            # Log des références pour vérifier
            refs = [row.get('ref', '') for row in structure]
            print(f"Références {efi} ({len(refs)}): {', '.join(refs[:30])}{'...' if len(refs) > 30 else ''}")
            
            # Vérifier si toutes les lignes sont incluses
            if len(structure) < len(select_mapping):
                missing_refs = [data.get('ref', 'NO_REF') for data in select_mapping if data.get('ref', '') not in refs]
                print(f"Lignes manquantes dans {efi}: {missing_refs}")

            datum[efi] = structure
        # Ajustement cible des comptes mixtes (actif/passif selon le signe)
        # pour conserver la structure historique des EFI.
        try:
            def _starts(numero, *prefixes):
                s = str(numero or "")
                return any(s.startswith(p) for p in prefixes)

            def _sum_mixed(prefixes, *, is_debiteur, exclude_prefixes=()):
                total_n = 0
                total_n1 = 0
                for item in (balance_variation or []):
                    numero = str(item.get("numero_compte", "") or "")
                    if not _starts(numero, *prefixes):
                        continue
                    if exclude_prefixes and _starts(numero, *exclude_prefixes):
                        continue
                    s_n = item.get("solde_n", 0) or 0
                    s_n1 = item.get("solde_n1", 0) or 0
                    if is_debiteur:
                        if s_n >= 0:
                            total_n += s_n
                        if s_n1 >= 0:
                            total_n1 += s_n1
                    else:
                        if s_n < 0:
                            total_n += s_n
                        if s_n1 < 0:
                            total_n1 += s_n1
                return total_n, total_n1

            def _set_row(section, ref, n_val, n1_val):
                rows = datum.get(section, [])
                row = next((r for r in rows if str(r.get("ref", "")).strip().upper() == ref), None)
                if not row:
                    return
                row["net_solde_n"] = n_val
                row["net_solde_n1"] = n1_val
                row["brut_solde_n"] = n_val
                row["amor_solde_n"] = 0

            # BI: 41 debiteurs (hors 419)
            bi_n, bi_n1 = _sum_mixed(("41",), is_debiteur=True, exclude_prefixes=("419",))
            _set_row("actif", "BI", bi_n, bi_n1)

            # BJ: 42/43/44/45/46/47/185 debiteurs (hors 478/479)
            bj_n, bj_n1 = _sum_mixed(
                ("42", "43", "44", "45", "46", "47", "185"),
                is_debiteur=True,
                exclude_prefixes=("478", "479"),
            )
            _set_row("actif", "BJ", bj_n, bj_n1)

            # DI: 419 (clients avances recues), quel que soit le signe
            di_n_d, di_n1_d = _sum_mixed(("419",), is_debiteur=True)
            di_n_c, di_n1_c = _sum_mixed(("419",), is_debiteur=False)
            _set_row("passif", "DI", di_n_d + di_n_c, di_n1_d + di_n1_c)

            # DK: 42/43/44 crediteurs
            dk_n, dk_n1 = _sum_mixed(("42", "43", "44"), is_debiteur=False)
            _set_row("passif", "DK", dk_n, dk_n1)

            # DM: 185/45/46/47 crediteurs (hors 478/479)
            dm_n, dm_n1 = _sum_mixed(
                ("185", "45", "46", "47"),
                is_debiteur=False,
                exclude_prefixes=("478", "479"),
            )
            _set_row("passif", "DM", dm_n, dm_n1)

            # DJ: tous les 40* sauf 409* (quel que soit le signe)
            dj_n = 0
            dj_n1 = 0
            for item in (balance_variation or []):
                numero = str(item.get("numero_compte", "") or "")
                if not _starts(numero, "40") or _starts(numero, "409"):
                    continue
                dj_n += item.get("solde_n", 0) or 0
                dj_n1 += item.get("solde_n1", 0) or 0
            _set_row("passif", "DJ", dj_n, dj_n1)

            # Classe 5 mixte:
            # - BS (actif): comptes 52/53/54/55/57/58 en solde debiteur
            # - DR (passif): comptes 52/53/54/55/57/58 en solde crediteur + 561/566
            bs_n, bs_n1 = _sum_mixed(("52", "53", "54", "55", "57", "58"), is_debiteur=True)
            _set_row("actif", "BS", bs_n, bs_n1)

            dr_mix_n, dr_mix_n1 = _sum_mixed(("52", "53", "54", "55", "57", "58"), is_debiteur=False)
            dr_561_n_d, dr_561_n1_d = _sum_mixed(("561",), is_debiteur=True)
            dr_561_n_c, dr_561_n1_c = _sum_mixed(("561",), is_debiteur=False)
            dr_566_n_d, dr_566_n1_d = _sum_mixed(("566",), is_debiteur=True)
            dr_566_n_c, dr_566_n1_c = _sum_mixed(("566",), is_debiteur=False)
            _set_row(
                "passif",
                "DR",
                dr_mix_n + dr_561_n_d + dr_561_n_c + dr_566_n_d + dr_566_n_c,
                dr_mix_n1 + dr_561_n1_d + dr_561_n1_c + dr_566_n1_d + dr_566_n1_c,
            )

            # CH (Report à nouveau): si positif (ex: 129), afficher en négatif.
            # Si déjà négatif (ex: 121), conserver tel quel.
            passif_rows = datum.get("passif", [])
            ch_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CH"), None)
            if ch_row:
                ch_n = ch_row.get("net_solde_n", 0) or 0
                ch_n1 = ch_row.get("net_solde_n1", 0) or 0
                ch_row["net_solde_n"] = -abs(ch_n)
                ch_row["net_solde_n1"] = -abs(ch_n1)
                ch_row["brut_solde_n"] = ch_row["net_solde_n"]
                ch_row["amor_solde_n"] = 0

            # CA (Capital): toujours affiché en positif au passif.
            ca_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CA"), None)
            if ca_row:
                ca_n = ca_row.get("net_solde_n", 0) or 0
                ca_n1 = ca_row.get("net_solde_n1", 0) or 0
                ca_row["net_solde_n"] = abs(ca_n)
                ca_row["net_solde_n1"] = abs(ca_n1)
                ca_row["brut_solde_n"] = ca_row["net_solde_n"]
                ca_row["amor_solde_n"] = 0

            # CB : toujours affiché en négatif au passif.
            cb_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CB"), None)
            if cb_row:
                cb_n = cb_row.get("net_solde_n", 0) or 0
                cb_n1 = cb_row.get("net_solde_n1", 0) or 0
                cb_row["net_solde_n"] = -abs(cb_n)
                cb_row["net_solde_n1"] = -abs(cb_n1)
                cb_row["brut_solde_n"] = cb_row["net_solde_n"]
                cb_row["amor_solde_n"] = 0

            # CD : toujours affiché en positif au passif.
            cd_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CD"), None)
            if cd_row:
                cd_n = cd_row.get("net_solde_n", 0) or 0
                cd_n1 = cd_row.get("net_solde_n1", 0) or 0
                cd_row["net_solde_n"] = abs(cd_n)
                cd_row["net_solde_n1"] = abs(cd_n1)
                cd_row["brut_solde_n"] = cd_row["net_solde_n"]
                cd_row["amor_solde_n"] = 0

            # CE : toujours affiché en positif au passif.
            ce_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CE"), None)
            if ce_row:
                ce_n = ce_row.get("net_solde_n", 0) or 0
                ce_n1 = ce_row.get("net_solde_n1", 0) or 0
                ce_row["net_solde_n"] = abs(ce_n)
                ce_row["net_solde_n1"] = abs(ce_n1)
                ce_row["brut_solde_n"] = ce_row["net_solde_n"]
                ce_row["amor_solde_n"] = 0

            # CF : toujours affiché en positif au passif.
            cf_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CF"), None)
            if cf_row:
                cf_n = cf_row.get("net_solde_n", 0) or 0
                cf_n1 = cf_row.get("net_solde_n1", 0) or 0
                cf_row["net_solde_n"] = abs(cf_n)
                cf_row["net_solde_n1"] = abs(cf_n1)
                cf_row["brut_solde_n"] = cf_row["net_solde_n"]
                cf_row["amor_solde_n"] = 0

            # CG : toujours affiché en positif au passif.
            cg_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CG"), None)
            if cg_row:
                cg_n = cg_row.get("net_solde_n", 0) or 0
                cg_n1 = cg_row.get("net_solde_n1", 0) or 0
                cg_row["net_solde_n"] = abs(cg_n)
                cg_row["net_solde_n1"] = abs(cg_n1)
                cg_row["brut_solde_n"] = cg_row["net_solde_n"]
                cg_row["amor_solde_n"] = 0

            # CL : toujours affiché en positif au passif.
            cl_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CL"), None)
            if cl_row:
                cl_n = cl_row.get("net_solde_n", 0) or 0
                cl_n1 = cl_row.get("net_solde_n1", 0) or 0
                cl_row["net_solde_n"] = abs(cl_n)
                cl_row["net_solde_n1"] = abs(cl_n1)
                cl_row["brut_solde_n"] = cl_row["net_solde_n"]
                cl_row["amor_solde_n"] = 0

            # CM : toujours affiché en positif au passif.
            cm_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "CM"), None)
            if cm_row:
                cm_n = cm_row.get("net_solde_n", 0) or 0
                cm_n1 = cm_row.get("net_solde_n1", 0) or 0
                cm_row["net_solde_n"] = abs(cm_n)
                cm_row["net_solde_n1"] = abs(cm_n1)
                cm_row["brut_solde_n"] = cm_row["net_solde_n"]
                cm_row["amor_solde_n"] = 0

            # DA : toujours affiché en positif au passif.
            da_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DA"), None)
            if da_row:
                da_n = da_row.get("net_solde_n", 0) or 0
                da_n1 = da_row.get("net_solde_n1", 0) or 0
                da_row["net_solde_n"] = abs(da_n)
                da_row["net_solde_n1"] = abs(da_n1)
                da_row["brut_solde_n"] = da_row["net_solde_n"]
                da_row["amor_solde_n"] = 0

            # DB : toujours affiché en positif au passif.
            db_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DB"), None)
            if db_row:
                db_n = db_row.get("net_solde_n", 0) or 0
                db_n1 = db_row.get("net_solde_n1", 0) or 0
                db_row["net_solde_n"] = abs(db_n)
                db_row["net_solde_n1"] = abs(db_n1)
                db_row["brut_solde_n"] = db_row["net_solde_n"]
                db_row["amor_solde_n"] = 0

            # DC : toujours affiché en positif au passif.
            dc_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DC"), None)
            if dc_row:
                dc_n = dc_row.get("net_solde_n", 0) or 0
                dc_n1 = dc_row.get("net_solde_n1", 0) or 0
                dc_row["net_solde_n"] = abs(dc_n)
                dc_row["net_solde_n1"] = abs(dc_n1)
                dc_row["brut_solde_n"] = dc_row["net_solde_n"]
                dc_row["amor_solde_n"] = 0

            # DH : toujours affiché en positif au passif.
            dh_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DH"), None)
            if dh_row:
                dh_n = dh_row.get("net_solde_n", 0) or 0
                dh_n1 = dh_row.get("net_solde_n1", 0) or 0
                dh_row["net_solde_n"] = abs(dh_n)
                dh_row["net_solde_n1"] = abs(dh_n1)
                dh_row["brut_solde_n"] = dh_row["net_solde_n"]
                dh_row["amor_solde_n"] = 0

            # DI : toujours affiché en positif au passif.
            di_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DI"), None)
            if di_row:
                di_n = di_row.get("net_solde_n", 0) or 0
                di_n1 = di_row.get("net_solde_n1", 0) or 0
                di_row["net_solde_n"] = abs(di_n)
                di_row["net_solde_n1"] = abs(di_n1)
                di_row["brut_solde_n"] = di_row["net_solde_n"]
                di_row["amor_solde_n"] = 0

            # DJ : toujours affiché en positif au passif.
            dj_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DJ"), None)
            if dj_row:
                dj_n = dj_row.get("net_solde_n", 0) or 0
                dj_n1 = dj_row.get("net_solde_n1", 0) or 0
                dj_row["net_solde_n"] = abs(dj_n)
                dj_row["net_solde_n1"] = abs(dj_n1)
                dj_row["brut_solde_n"] = dj_row["net_solde_n"]
                dj_row["amor_solde_n"] = 0

            # DK : toujours affiché en positif au passif.
            dk_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DK"), None)
            if dk_row:
                dk_n = dk_row.get("net_solde_n", 0) or 0
                dk_n1 = dk_row.get("net_solde_n1", 0) or 0
                dk_row["net_solde_n"] = abs(dk_n)
                dk_row["net_solde_n1"] = abs(dk_n1)
                dk_row["brut_solde_n"] = dk_row["net_solde_n"]
                dk_row["amor_solde_n"] = 0

            # DL : toujours affiché en positif au passif.
            dl_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DL"), None)
            if dl_row:
                dl_n = dl_row.get("net_solde_n", 0) or 0
                dl_n1 = dl_row.get("net_solde_n1", 0) or 0
                dl_row["net_solde_n"] = abs(dl_n)
                dl_row["net_solde_n1"] = abs(dl_n1)
                dl_row["brut_solde_n"] = dl_row["net_solde_n"]
                dl_row["amor_solde_n"] = 0

            # DM : toujours affiché en positif au passif.
            dm_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DM"), None)
            if dm_row:
                dm_n = dm_row.get("net_solde_n", 0) or 0
                dm_n1 = dm_row.get("net_solde_n1", 0) or 0
                dm_row["net_solde_n"] = abs(dm_n)
                dm_row["net_solde_n1"] = abs(dm_n1)
                dm_row["brut_solde_n"] = dm_row["net_solde_n"]
                dm_row["amor_solde_n"] = 0

            # DN : toujours affiché en positif au passif.
            dn_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DN"), None)
            if dn_row:
                dn_n = dn_row.get("net_solde_n", 0) or 0
                dn_n1 = dn_row.get("net_solde_n1", 0) or 0
                dn_row["net_solde_n"] = abs(dn_n)
                dn_row["net_solde_n1"] = abs(dn_n1)
                dn_row["brut_solde_n"] = dn_row["net_solde_n"]
                dn_row["amor_solde_n"] = 0

            # DQ : toujours affiché en positif au passif.
            dq_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DQ"), None)
            if dq_row:
                dq_n = dq_row.get("net_solde_n", 0) or 0
                dq_n1 = dq_row.get("net_solde_n1", 0) or 0
                dq_row["net_solde_n"] = abs(dq_n)
                dq_row["net_solde_n1"] = abs(dq_n1)
                dq_row["brut_solde_n"] = dq_row["net_solde_n"]
                dq_row["amor_solde_n"] = 0

            # DR : toujours affiché en positif au passif.
            dr_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DR"), None)
            if dr_row:
                dr_n = dr_row.get("net_solde_n", 0) or 0
                dr_n1 = dr_row.get("net_solde_n1", 0) or 0
                dr_row["net_solde_n"] = abs(dr_n)
                dr_row["net_solde_n1"] = abs(dr_n1)
                dr_row["brut_solde_n"] = dr_row["net_solde_n"]
                dr_row["amor_solde_n"] = 0

            # DV : toujours affiché en positif au passif.
            dv_row = next((r for r in passif_rows if str(r.get("ref", "")).strip().upper() == "DV"), None)
            if dv_row:
                dv_n = dv_row.get("net_solde_n", 0) or 0
                dv_n1 = dv_row.get("net_solde_n1", 0) or 0
                dv_row["net_solde_n"] = abs(dv_n)
                dv_row["net_solde_n1"] = abs(dv_n1)
                dv_row["brut_solde_n"] = dv_row["net_solde_n"]
                dv_row["amor_solde_n"] = 0
        except Exception as e:
            print(f"Ajustement comptes mixtes EFI ignore: {e}")
        # Recalcule des agrégats ACTIF selon les formules officielles
        # pour éviter toute dérive liée à une configuration mapping.
        actif_rows = datum.get('actif', [])
        if actif_rows:
            idx = {row.get('ref'): row for row in actif_rows if row.get('ref')}

            def _g(ref, key):
                row = idx.get(ref) or {}
                value = row.get(key, 0)
                return value if value is not None else 0

            def _sum_refs(refs, key):
                return sum(_g(r, key) for r in refs)

            def _set_from_refs(ref, refs):
                if ref in idx:
                    idx[ref]['brut_solde_n'] = _sum_refs(refs, 'brut_solde_n')
                    idx[ref]['amor_solde_n'] = _sum_refs(refs, 'amor_solde_n')
                    idx[ref]['net_solde_n'] = _sum_refs(refs, 'net_solde_n')
                    idx[ref]['net_solde_n1'] = _sum_refs(refs, 'net_solde_n1')

            _set_from_refs('AD', ['AE', 'AF', 'AG', 'AH'])
            _set_from_refs('AI', ['AJ', 'AK', 'AL', 'AM', 'AN'])
            _set_from_refs('AQ', ['AR', 'AS'])
            _set_from_refs('BG', ['BH', 'BI', 'BJ'])
            _set_from_refs('AZ', ['AD', 'AI', 'AP', 'AQ'])
            _set_from_refs('BK', ['BA', 'BB', 'BG'])
            _set_from_refs('BT', ['BQ', 'BR', 'BS'])
            _set_from_refs('BZ', ['AZ', 'BK', 'BT', 'BU'])

        # Recalcule des agrégats PASSIF selon les formules officielles.
        passif_rows = datum.get('passif', [])
        if passif_rows:
            idx = {row.get('ref'): row for row in passif_rows if row.get('ref')}

            def _g(ref, key):
                row = idx.get(ref) or {}
                value = row.get(key, 0)
                return value if value is not None else 0

            def _sum_refs(refs, key):
                return sum(_g(r, key) for r in refs)

            def _set_from_refs(ref, refs):
                if ref in idx:
                    idx[ref]['brut_solde_n'] = _sum_refs(refs, 'brut_solde_n')
                    idx[ref]['amor_solde_n'] = _sum_refs(refs, 'amor_solde_n')
                    idx[ref]['net_solde_n'] = _sum_refs(refs, 'net_solde_n')
                    idx[ref]['net_solde_n1'] = _sum_refs(refs, 'net_solde_n1')

            _set_from_refs('CP', ['CA', 'CB', 'CD', 'CE', 'CF', 'CG', 'CH', 'CJ', 'CL', 'CM'])
            _set_from_refs('DD', ['DA', 'DB', 'DC'])
            _set_from_refs('DF', ['CP', 'DD'])
            _set_from_refs('DP', ['DH', 'DI', 'DJ', 'DK', 'DM', 'DN'])
            _set_from_refs('DT', ['DQ', 'DR'])
            _set_from_refs('DZ', ['DF', 'DP', 'DT', 'DV'])

        # Recalcule des agrégats PNL selon les formules officielles.
        pnl_rows = datum.get('pnl', [])
        if pnl_rows:
            idx = {row.get('ref'): row for row in pnl_rows if row.get('ref')}

            def _g(ref, key):
                row = idx.get(ref) or {}
                value = row.get(key, 0)
                return value if value is not None else 0

            def _set(ref, n_val, n1_val):
                if ref in idx:
                    idx[ref]['net_solde_n'] = n_val
                    idx[ref]['net_solde_n1'] = n1_val

            def _pm(val):
                # Règle demandée pour les termes +/- :
                # solde négatif => on fait "+"
                # solde positif => on fait "-"
                # En pratique : contribution = val si val < 0, sinon -val
                return val if val < 0 else -val

            xa_n = -_g('TA', 'net_solde_n') - _g('RA', 'net_solde_n') + _g('RB', 'net_solde_n')
            xa_n1 = -_g('TA', 'net_solde_n1') - _g('RA', 'net_solde_n1') + _g('RB', 'net_solde_n1')
            _set('XA', xa_n, xa_n1)

            xb_n = -_g('TA', 'net_solde_n') - _g('TB', 'net_solde_n') - _g('TC', 'net_solde_n') - _g('TD', 'net_solde_n')
            xb_n1 = -_g('TA', 'net_solde_n1') - _g('TB', 'net_solde_n1') - _g('TC', 'net_solde_n1') - _g('TD', 'net_solde_n1')
            _set('XB', xb_n, xb_n1)

            xc_n = (
                xb_n
                - _g('RA', 'net_solde_n') + _pm(_g('RB', 'net_solde_n'))
                + _pm(_g('TE', 'net_solde_n')) - _g('TF', 'net_solde_n') - _g('TG', 'net_solde_n') - _g('TH', 'net_solde_n') - _g('TI', 'net_solde_n')
                - _g('RC', 'net_solde_n') + _pm(_g('RD', 'net_solde_n'))
                - _g('RE', 'net_solde_n') + _pm(_g('RF', 'net_solde_n'))
                - _g('RG', 'net_solde_n') - _g('RH', 'net_solde_n') - _g('RI', 'net_solde_n') - _g('RJ', 'net_solde_n')
            )
            xc_n1 = (
                xb_n1
                - _g('RA', 'net_solde_n1') + _pm(_g('RB', 'net_solde_n1'))
                + _pm(_g('TE', 'net_solde_n1')) - _g('TF', 'net_solde_n1') - _g('TG', 'net_solde_n1') - _g('TH', 'net_solde_n1') - _g('TI', 'net_solde_n1')
                - _g('RC', 'net_solde_n1') + _pm(_g('RD', 'net_solde_n1'))
                - _g('RE', 'net_solde_n1') + _pm(_g('RF', 'net_solde_n1'))
                - _g('RG', 'net_solde_n1') - _g('RH', 'net_solde_n1') - _g('RI', 'net_solde_n1') - _g('RJ', 'net_solde_n1')
            )
            _set('XC', xc_n, xc_n1)

            xd_n = xc_n - _g('RK', 'net_solde_n')
            xd_n1 = xc_n1 - _g('RK', 'net_solde_n1')
            _set('XD', xd_n, xd_n1)

            xe_n = xd_n - _g('TJ', 'net_solde_n') - _g('RL', 'net_solde_n')
            xe_n1 = xd_n1 - _g('TJ', 'net_solde_n1') - _g('RL', 'net_solde_n1')
            _set('XE', xe_n, xe_n1)

            xf_n = -_g('TK', 'net_solde_n') - _g('TL', 'net_solde_n') - _g('TM', 'net_solde_n') - _g('RM', 'net_solde_n') - _g('RN', 'net_solde_n')
            xf_n1 = -_g('TK', 'net_solde_n1') - _g('TL', 'net_solde_n1') - _g('TM', 'net_solde_n1') - _g('RM', 'net_solde_n1') - _g('RN', 'net_solde_n1')
            _set('XF', xf_n, xf_n1)

            xg_n = xe_n + xf_n
            xg_n1 = xe_n1 + xf_n1
            _set('XG', xg_n, xg_n1)

            xh_n = -_g('TN', 'net_solde_n') - _g('TO', 'net_solde_n') - _g('RO', 'net_solde_n') - _g('RP', 'net_solde_n')
            xh_n1 = -_g('TN', 'net_solde_n1') - _g('TO', 'net_solde_n1') - _g('RO', 'net_solde_n1') - _g('RP', 'net_solde_n1')
            _set('XH', xh_n, xh_n1)

            xi_n = xg_n + xh_n - _g('RQ', 'net_solde_n') - _g('RS', 'net_solde_n')
            xi_n1 = xg_n1 + xh_n1 - _g('RQ', 'net_solde_n1') - _g('RS', 'net_solde_n1')
            _set('XI', xi_n, xi_n1)

            # Inversion de signe des comptes "à additionner" du CR
            # (positif -> négatif, négatif -> positif), ex: TB, TD, TF...
            pnl_component_refs = {
                'TA', 'TB', 'TC', 'TD',
                'RA', 'RB', 'TE', 'TF', 'TG', 'TH', 'TI',
                'RC', 'RD', 'RE', 'RF', 'RG', 'RH', 'RI', 'RJ',
                'RK', 'TJ', 'RL',
                'TK', 'TL', 'TM', 'RM', 'RN',
                'TN', 'TO', 'RO', 'RP',
                'RQ', 'RS',
            }
            for ref in pnl_component_refs:
                row = idx.get(ref)
                if not row:
                    continue
                row['net_solde_n'] = -(row.get('net_solde_n', 0) or 0)
                row['net_solde_n1'] = -(row.get('net_solde_n1', 0) or 0)

            # Recalcul final des agrégats PNL après inversion des comptes composants
            # pour que les lignes XA..XI soient cohérentes avec les montants affichés.
            xa_n = -_g('TA', 'net_solde_n') - _g('RA', 'net_solde_n') + _g('RB', 'net_solde_n')
            xa_n1 = -_g('TA', 'net_solde_n1') - _g('RA', 'net_solde_n1') + _g('RB', 'net_solde_n1')
            _set('XA', - xa_n, - xa_n1)

            xb_n = -_g('TA', 'net_solde_n') - _g('TB', 'net_solde_n') - _g('TC', 'net_solde_n') - _g('TD', 'net_solde_n')
            xb_n1 = -_g('TA', 'net_solde_n1') - _g('TB', 'net_solde_n1') - _g('TC', 'net_solde_n1') - _g('TD', 'net_solde_n1')
            _set('XB', - xb_n, - xb_n1)

            xc_n = (
                xb_n
                - _g('RA', 'net_solde_n') + _pm(_g('RB', 'net_solde_n'))
                + _pm(_g('TE', 'net_solde_n')) - _g('TF', 'net_solde_n') - _g('TG', 'net_solde_n') - _g('TH', 'net_solde_n') - _g('TI', 'net_solde_n')
                - _g('RC', 'net_solde_n') + _pm(_g('RD', 'net_solde_n'))
                - _g('RE', 'net_solde_n') + _pm(_g('RF', 'net_solde_n'))
                - _g('RG', 'net_solde_n') - _g('RH', 'net_solde_n') - _g('RI', 'net_solde_n') - _g('RJ', 'net_solde_n')
            )
            xc_n1 = (
                xb_n1
                - _g('RA', 'net_solde_n1') + _pm(_g('RB', 'net_solde_n1'))
                + _pm(_g('TE', 'net_solde_n1')) - _g('TF', 'net_solde_n1') - _g('TG', 'net_solde_n1') - _g('TH', 'net_solde_n1') - _g('TI', 'net_solde_n1')
                - _g('RC', 'net_solde_n1') + _pm(_g('RD', 'net_solde_n1'))
                - _g('RE', 'net_solde_n1') + _pm(_g('RF', 'net_solde_n1'))
                - _g('RG', 'net_solde_n1') - _g('RH', 'net_solde_n1') - _g('RI', 'net_solde_n1') - _g('RJ', 'net_solde_n1')
            )
            _set('XC', - xc_n, - xc_n1)

            xd_n = xc_n - _g('RK', 'net_solde_n')
            xd_n1 = xc_n1 - _g('RK', 'net_solde_n1')
            _set('XD', - xd_n, - xd_n1)

            xe_n = xd_n - _g('TJ', 'net_solde_n') - _g('RL', 'net_solde_n')
            xe_n1 = xd_n1 - _g('TJ', 'net_solde_n1') - _g('RL', 'net_solde_n1')
            _set('XE', - xe_n, - xe_n1)

            xf_n = -_g('TK', 'net_solde_n') - _g('TL', 'net_solde_n') - _g('TM', 'net_solde_n') - _g('RM', 'net_solde_n') - _g('RN', 'net_solde_n')
            xf_n1 = -_g('TK', 'net_solde_n1') - _g('TL', 'net_solde_n1') - _g('TM', 'net_solde_n1') - _g('RM', 'net_solde_n1') - _g('RN', 'net_solde_n1')
            _set('XF', - xf_n, - xf_n1)

            xg_n = xe_n + xf_n
            xg_n1 = xe_n1 + xf_n1
            _set('XG', - xg_n, - xg_n1)

            xh_n = -_g('TN', 'net_solde_n') - _g('TO', 'net_solde_n') - _g('RO', 'net_solde_n') - _g('RP', 'net_solde_n')
            xh_n1 = -_g('TN', 'net_solde_n1') - _g('TO', 'net_solde_n1') - _g('RO', 'net_solde_n1') - _g('RP', 'net_solde_n1')
            _set('XH', - xh_n, - xh_n1)

            xi_n = xg_n + xh_n - _g('RQ', 'net_solde_n') - _g('RS', 'net_solde_n')
            xi_n1 = xg_n1 + xh_n1 - _g('RQ', 'net_solde_n1') - _g('RS', 'net_solde_n1')
            _set('XI', - xi_n, - xi_n1)

        # Alimentation explicite de CJ (Résultat net de l'exercice)
        # depuis la balance PNL: somme des comptes de classes 6 à 8.
        # Puis recalcul des agrégats passif impactés (CP, DF, DZ).
        passif_rows = datum.get('passif', [])
        if passif_rows:
            idxp = {row.get('ref'): row for row in passif_rows if row.get('ref')}

            cj_n = - sum(
                (item.get('solde_reel', 0) or 0)
                for item in (balance_n or [])
                if str(item.get('numero_compte', '')).startswith(('6', '7', '8'))
            )
            cj_n1 = - sum(
                (item.get('solde_reel', 0) or 0)
                for item in (balance_n1 or [])
                if str(item.get('numero_compte', '')).startswith(('6', '7', '8'))
            )

            if 'CJ' in idxp:
                idxp['CJ']['net_solde_n'] = cj_n
                idxp['CJ']['net_solde_n1'] = cj_n1
                idxp['CJ']['brut_solde_n'] = cj_n
                idxp['CJ']['amor_solde_n'] = 0

            def _gp(ref, key):
                row = idxp.get(ref) or {}
                value = row.get(key, 0)
                return value if value is not None else 0

            def _setp_from_refs(ref, refs):
                if ref in idxp:
                    idxp[ref]['brut_solde_n'] = sum(_gp(r, 'brut_solde_n') for r in refs)
                    idxp[ref]['amor_solde_n'] = sum(_gp(r, 'amor_solde_n') for r in refs)
                    idxp[ref]['net_solde_n'] = sum(_gp(r, 'net_solde_n') for r in refs)
                    idxp[ref]['net_solde_n1'] = sum(_gp(r, 'net_solde_n1') for r in refs)

            _setp_from_refs('CP', ['CA', 'CB', 'CD', 'CE', 'CF', 'CG', 'CH', 'CJ', 'CL', 'CM'])
            _setp_from_refs('DF', ['CP', 'DD'])
            _setp_from_refs('DZ', ['DF', 'DP', 'DT', 'DV'])

        return datum



    # ---------- Piste d'audit ----------

    def audit_trail(self, id_mission):
        # Assurer la connexion DB (certain code legacy utilisait le global db)
        try:
            local_db = get_db()
            if local_db is None:
                from src.utils.database import ensure_connection
                print("audit_trail: get_db() a renvoyé None, tentative de reconnexion...")
                ensure_connection()
                local_db = get_db()
            if local_db is None:
                raise RuntimeError("audit_trail: Base de données non connectée (get_db a renvoyé None)")
        except Exception as e:
            print(f"audit_trail: impossible d'obtenir la connexion DB: {e}")
            raise

        # Créer un fichier Excel pour la piste d'audit
        wb = openpyxl.Workbook()

        sheet = wb.active



        columns = ['A', 'B', 'C', 'D', 'E']

        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Code COMMENTAIRE']



        mission = local_db.Mission1.find_one({"_id": ObjectId(id_mission)})

        balances = mission['balance_variation']

        actif = mission['efi']['actif']

        passif = mission['efi']['passif']

        pnl = mission['efi']['pnl']



        efi = actif + passif + pnl



        for i in range(len(columns)):

            sheet[columns[i] + '1'] = headers[i]



        for iteration, data in enumerate(balances):

            new_iteration = str(iteration + 2)

            sheet["A" + new_iteration] = data.get("numero_compte")

            sheet["B" + new_iteration] = data.get("solde_n")

            sheet["C" + new_iteration] = data.get("solde_n1")

            sheet["D" + new_iteration] = data.get("numero_compte")[0:2]



            list_code_efi = []

            for obj in efi:

                for elt in obj['compte_to_be_used_off']:

                    if data['numero_compte'].startswith(elt):

                        list_code_efi.append(obj['ref'])

            list_code_efi = list(set(list_code_efi))

            sheet["E" + new_iteration] = ','.join(list_code_efi)



        namefile = "piste_audit.xlsx"

        wb.save(namefile)



    # ---------- Extract grouping Excel ----------

    def extract_grouping(self, id_mission):

        # Créer un fichier Excel pour l'export du grouping
        wb = openpyxl.Workbook()

        sheet = wb.active



        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Variation', 'Variation %', 'Compte qualitatif', 'Compte quantitatif', 'Compte significatif']



        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        balances = mission['balance_variation']

        grouping = mission['grouping']

        materiality = next(item for item in mission['materiality'] if item['choice'] is True)



        for i in range(len(columns)):

            sheet[columns[i] + '1'] = headers[i]



        for iteration, data in enumerate(balances):

            new_iteration = str(iteration + 2)

            sheet["A" + new_iteration] = data.get("numero_compte")

            sheet["B" + new_iteration] = data.get("solde_n")

            sheet["C" + new_iteration] = data.get("solde_n1")



            value_grouping = data.get("numero_compte")[0:2]

            variation = data.get("solde_n") - data.get("solde_n1")



            if variation == 0:

                variation_percent = 0

            elif data.get("solde_n1") == 0:

                variation_percent = 100

            else:

                variation_percent = (variation / data.get("solde_n1")) * 100



            sheet["D" + new_iteration] = value_grouping

            sheet["E" + new_iteration] = variation

            sheet["F" + new_iteration] = variation_percent

            sheet["G" + new_iteration] = next(item['significant'] for item in grouping if item['compte'] == value_grouping)

            sheet["H" + new_iteration] = next(item['materiality'] for item in grouping if item['compte'] == value_grouping)

            sheet["I" + new_iteration] = next(item['mat_sign'] for item in grouping if item['compte'] == value_grouping)



        second_sheet = wb.create_sheet(title="Seuil de matérialité")

        second_headers = ['materiality', 'performance materiality', 'trivial misstatements']

        second_sheet["A1"] = second_headers[0]

        second_sheet["B1"] = second_headers[1]

        second_sheet["C1"] = second_headers[2]



        second_sheet["A2"] = materiality['materiality']

        second_sheet["B2"] = materiality['performance_materiality']

        second_sheet["C2"] = materiality['trivial_misstatements']



        excel_io = BytesIO()

        wb.save(excel_io)

        excel_io.seek(0)

        return excel_io
