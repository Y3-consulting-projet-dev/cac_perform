"""
Importe une liste d'utilisateurs autorises dans MongoDB avec un mot de passe initial.

Usage:
    python import_authorized_users.py --file users.csv --password "MotDePasse123!"
    python import_authorized_users.py --file "C:\\mon_fichier.xlsx" --password "MotDePasse123!" --default-role "Auditeur" --default-grade "Junior"

Format CSV attendu:
    firstname,lastname,email,role,grade
    Jean,Dupont,jean.dupont@example.com,Manager,Senior
"""

import argparse
import csv
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

CURRENT_DIR = Path(__file__).resolve().parent
if str(CURRENT_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_DIR))

from openpyxl import load_workbook
from pymongo.errors import PyMongoError

from config import config as app_config, db_manager
from src.schemas.user_schemas import ROLES, GRADES, validate_email
from src.services.user_services import UserService
from src.utils.database import ensure_connection, get_db


def load_csv(csv_path: Path):
    with csv_path.open(mode="r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required_columns = {"firstname", "lastname", "email", "role", "grade"}
        missing_columns = required_columns.difference(reader.fieldnames or [])
        if missing_columns:
            missing = ", ".join(sorted(missing_columns))
            raise ValueError(f"Colonnes manquantes dans le CSV: {missing}")
        return list(reader)


def normalize_header(value):
    return (value or "").strip().lower()


def normalize_grade_value(value):
    raw = (value or "").strip()
    if not raw:
        return raw

    normalized = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")
    normalized = " ".join(normalized.lower().split())

    mapping = {
        "pre-emploi": "Pre-emploi",
        "pre emploi": "Pre-emploi",
        "assistant": "Assistant 1",
        "assistante": "Assistant 1",
        "assistant 1": "Assistant 1",
        "assistant1": "Assistant 1",
        "assistant 2": "Assistant 2",
        "assistant2": "Assistant 2",
        "senior": "Senior 3",
        "senior 3": "Senior 3",
        "senior3": "Senior 3",
        "assistant manager": "Assistant Manager",
        "manager": "Manager 1",
        "manager 1": "Manager 1",
        "manager1": "Manager 1",
        "manager 2": "Manager 2",
        "manager2": "Manager 2",
        "manager 3": "Manager 3",
        "manager3": "Manager 3",
        "senior manager": "Senior Manager",
    }
    return mapping.get(normalized, raw.strip())


def load_xlsx(xlsx_path: Path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(cell) for cell in rows[0]]
    header_map = {header: index for index, header in enumerate(headers) if header}

    normalized_rows = []
    for raw_row in rows[1:]:
        if not raw_row:
            continue

        def get_value(possible_headers):
            for header in possible_headers:
                index = header_map.get(header)
                if index is not None and index < len(raw_row):
                    value = raw_row[index]
                    return "" if value is None else str(value).strip()
            return ""

        normalized_rows.append(
            {
                "lastname": get_value(["nom", "lastname"]),
                "firstname": get_value(["prenoms", "prénoms", "firstname"]),
                "email": get_value(["mail", "email", "e-mail"]),
                "role": get_value(["role", "fonction"]),
                "grade": get_value(["grade"]),
            }
        )

    return normalized_rows


def load_rows(input_path: Path):
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return load_csv(input_path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx(input_path)
    raise ValueError("Format de fichier non supporte. Utilisez un CSV ou un XLSX.")


def build_payload(row, default_password: str, default_role: str, default_grade: str):
    firstname = (row.get("firstname") or "").strip()
    lastname = (row.get("lastname") or "").strip()
    email = (row.get("email") or "").strip().lower()
    role = ((row.get("role") or "").strip() or default_role)
    grade = ((row.get("grade") or "").strip() or default_grade)
    grade = normalize_grade_value(grade)

    if len(firstname) < 2:
        raise ValueError(f"Prenom invalide pour {email or lastname}")
    if len(lastname) < 2:
        raise ValueError(f"Nom invalide pour {email or firstname}")
    validate_email(email)
    if role not in ROLES:
        raise ValueError(f"Role invalide pour {email}: {role}")
    if grade not in GRADES:
        raise ValueError(f"Grade invalide pour {email}: {grade}")

    return {
        "firstname": firstname,
        "lastname": lastname,
        "email": email,
        "role": role,
        "grade": grade,
        "password": default_password,
    }


def import_users(
    rows,
    default_password: str,
    default_role: str,
    default_grade: str,
    reset_passwords: bool = False,
):
    if db_manager.config is None:
      default_config = app_config["default"]
      db_manager.config = {
          "MONGO_HOST": default_config.MONGO_HOST,
          "MONGO_PORT": default_config.MONGO_PORT,
          "MONGO_DB_NAME": default_config.MONGO_DB_NAME,
          "MONGO_USERNAME": default_config.MONGO_USERNAME,
          "MONGO_PASSWORD": default_config.MONGO_PASSWORD,
          "MONGO_AUTH_SOURCE": default_config.MONGO_AUTH_SOURCE,
          "MONGO_CONNECT_TIMEOUT_MS": default_config.MONGO_CONNECT_TIMEOUT_MS,
          "MONGO_SERVER_SELECTION_TIMEOUT_MS": default_config.MONGO_SERVER_SELECTION_TIMEOUT_MS,
      }
    ensure_connection()
    db = get_db()
    created = 0
    updated = 0

    for row in rows:
        validated = build_payload(row, default_password, default_role, default_grade)
        email = validated["email"]
        existing_user = db.Manager.find_one({"email": email})

        document = {
            "firstname": validated["firstname"],
            "lastname": validated["lastname"],
            "email": email,
            "role": validated["role"],
            "grade": validated["grade"],
            "is_active": True,
            "directory_visible": True,
            "login_attempts": 0,
            "locked_until": None,
        }

        if existing_user:
            if reset_passwords:
                document["mot_de_passe"] = UserService.hash_password(default_password)
            db.Manager.update_one({"_id": existing_user["_id"]}, {"$set": document})
            updated += 1
        else:
            document["mot_de_passe"] = UserService.hash_password(default_password)
            document["user_id"] = UserService.generate_user_id()
            document["created_at"] = datetime.now()
            document["last_login"] = None
            db.Manager.insert_one(document)
            created += 1

    return created, updated


def main():
    parser = argparse.ArgumentParser(description="Importer les utilisateurs autorises.")
    parser.add_argument("--file", required=True, help="Chemin du fichier CSV ou XLSX.")
    parser.add_argument("--password", required=True, help="Mot de passe initial commun.")
    parser.add_argument("--default-role", default="Auditeur", help="Role par defaut si absent du fichier.")
    parser.add_argument("--default-grade", default="Junior", help="Grade par defaut si absent du fichier.")
    parser.add_argument(
        "--reset-passwords",
        action="store_true",
        help="Reinitialise aussi le mot de passe des utilisateurs deja existants.",
    )
    args = parser.parse_args()

    input_path = Path(args.file)
    if not input_path.exists():
        raise FileNotFoundError(f"Fichier introuvable: {input_path}")

    rows = load_rows(input_path)
    rows = [row for row in rows if (row.get("email") or "").strip()]
    if not rows:
        print("Aucun utilisateur a importer.")
        return

    try:
        created, updated = import_users(
            rows,
            args.password,
            args.default_role,
            args.default_grade,
            args.reset_passwords,
        )
        print(f"Import termine. Crees: {created}. Mis a jour: {updated}.")
    except PyMongoError as exc:
        raise RuntimeError(f"Erreur MongoDB pendant l'import: {exc}") from exc


if __name__ == "__main__":
    main()
